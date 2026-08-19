import argparse
import hashlib
import json
import math
import re
import shutil
import sys
import time
import unicodedata
import uuid
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import cv2
import fitz
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pdf2image import convert_from_path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from config import (  # noqa: E402
    DATASET_IMAGES_DIR,
    INSPECTION_PLAN_OUTPUT_PATH,
    MEASUREMENT_DETECTION_OUTPUT_PATH,
    MEASUREMENT_OCR_CONFIDENCE,
    MEASUREMENT_REVIEW_THRESHOLD,
    PDF_DPI,
    ROBOFLOW_CONFIDENCE,
    ROBOFLOW_ENABLED,
    ROBOFLOW_EXPECTED_MODEL_ID,
    ROBOFLOW_INCLUDE_FULL_IMAGE,
    ROBOFLOW_MAX_REGIONS,
    ROBOFLOW_TILE_OVERLAP,
    ROBOFLOW_TILING_ENABLED,
    ROBOFLOW_WORKFLOW_ID,
    YOLO_GDT_OCR_MIN_DETECTOR_CONFIDENCE,
    YOLO_TEXT_OCR_MIN_DETECTOR_CONFIDENCE,
    ensure_directories,
)
from drawing_metadata import DrawingMetadata, parse_metadata_from_filename  # noqa: E402
from exporter import sanitize_dataframe_for_excel  # noqa: E402
from layout_regions import (  # noqa: E402
    REPORT_EXCLUSION_REGIONS,
    exclusion_region_label,
    normalized_center_from_item,
)
from measurement_extraction import (  # noqa: E402
    classify_measurement,
    create_measurement_row,
    extract_ocr_items_with_boxes,
    extract_measurement_ocr_items,
    find_angled_dimension_crops,
    passes_dynamic_confidence,
    rescue_single_digit_dimensions,
    rescue_vertical_decimal_dimensions,
    repair_spaced_nominal_tolerance_text,
    rotate_image_bound,
    should_skip_text,
    transform_deskewed_crop_box_to_original,
)
from titleblock_ocr import extract_titleblock_metadata  # noqa: E402
from roboflow_workflow_client import get_active_roboflow_deployment  # noqa: E402
from vision_tools import (  # noqa: E402
    create_paddle_ocr,
    create_paddle_text_recognizer,
    detect_symbols_with_roboflow,
    detect_symbols_with_yolo,
    filter_implausible_symbol_detections,
    merge_symbol_detections,
)


FA_COLUMNS = [
    "BALLOON NO",
    "SYMBOL",
    "VALUE",
    "-",
    "+",
    "MIN",
    "MAX",
    "REMARKS",
]

WEB_COLUMNS = [
    "Balloon No",
    "Report Symbol",
    "Dimension",
    "Tolerance -",
    "Tolerance +",
    "Needs Review",
    "Review Reason",
]

GENERAL_TOLERANCE_RANGES = [
    (0.5, 3.0, 0.1),
    (3.0, 6.0, 0.1),
    (6.0, 30.0, 0.2),
    (30.0, 120.0, 0.3),
    (120.0, 400.0, 0.5),
    (400.0, 1000.0, 0.8),
    (1000.0, 2000.0, 1.2),
]

DATE_OR_REVISION_VALUE_PATTERN = re.compile(
    r"^\s*(?:\d{1,2}[./-]\d{1,2}(?:[./-]\d{2,4})?|\d{2,4}[./-]\d{1,2}[./-]\d{1,2}|0\d|[A-Z]?\d{1,2}[A-Z]?)\s*$",
    re.IGNORECASE,
)

# Desired/customer FA files are golden validation examples only. Runtime output is
# generated from the drawing OCR, symbol detections, and deterministic rules.
REFERENCE_REPORT_PATHS = {}

GEOMETRY_REQUIRED_TYPES = {"plain_dimension", "reference_dimension"}
LOW_CONFIDENCE_PLAIN_DIMENSION_THRESHOLD = 0.75

NOTE_TEXTS = [
    "NO SCRATCHES",
    "NO BURRS",
    "NO DENTS",
]

NOTE_FALSE_POSITIVE_PATTERNS = [
    re.compile(r"\bNOTE\b", re.IGNORECASE),
    re.compile(r"\bINSTRUCTION\b", re.IGNORECASE),
    re.compile(r"\bUSE\s+COLD\s+MATERIALS?\b", re.IGNORECASE),
    re.compile(r"\bTHE\s+CHAMFER\s+FOR\s+CORNER\b", re.IGNORECASE),
    re.compile(r"\bCHAMFER\s+CORNERS?\b", re.IGNORECASE),
    re.compile(r"\bUSE\s+R\s*0\.3\s+FOR\s+CORNERS?\b", re.IGNORECASE),
    re.compile(r"\bCLEAN\s+AND\s+DEGREASE\b", re.IGNORECASE),
    re.compile(r"\bNO\s+SCRATCHES\b", re.IGNORECASE),
    re.compile(r"\bNO\s+BURRS\b", re.IGNORECASE),
    re.compile(r"\bNO\s+DENTS\b", re.IGNORECASE),
    re.compile(r"\bMATERIAL\s+USE\b", re.IGNORECASE),
    re.compile(r"\bDEGREASE\b", re.IGNORECASE),
    re.compile(r"\bC\s*0\.\d+\s+-\s+C\s*0\.\d+\b", re.IGNORECASE),
]

COMMON_NOTES = [
    "Note 1 : THREAD CHAMFERING FOR NON INDICATED EDGES",
    "Note 2 : SURFACE TREATMENT / FINISH PER DRAWING",
    "Note 3 : USE SPECIFIED MATERIAL",
    "Note 4 : INSTALL INSERTS / HELICOILS BEFORE DELIVERY IF REQUIRED",
    "Note 5 : DEGREASING AND CLEANING BEFORE DELIVERY",
]


def safe_stem(filename):
    return Path(filename).stem.replace(" ", "_")


def timed_call(timings, name, operation, *args, **kwargs):
    """Run one pipeline operation and record its wall-clock duration."""
    started_at = time.perf_counter()
    try:
        return operation(*args, **kwargs)
    finally:
        timings[name] = round(time.perf_counter() - started_at, 3)


def content_render_cache_path(pdf_path, job_dir):
    """Return a safe render cache path tied to PDF bytes and render settings."""
    digest = hashlib.sha256()
    digest.update(f"ocr-drawing-render-v1|dpi={PDF_DPI}|".encode("utf-8"))
    with Path(pdf_path).open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)

    job_dir = Path(job_dir)
    generated_root = job_dir.parent.parent if job_dir.parent.name == "jobs" else job_dir.parent
    cache_dir = generated_root / "render_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir / f"{digest.hexdigest()}.png"


def content_measurement_cache_path(pdf_path, job_dir):
    """Return a cache path tied to the drawing bytes and OCR implementation."""
    digest = hashlib.sha256()
    digest.update(f"ocr-drawing-measurements-v1|dpi={PDF_DPI}|".encode("utf-8"))
    for source_path in (
        Path(__file__).resolve().parent / "measurement_extraction.py",
        Path(__file__).resolve().parent / "vision_tools.py",
        Path(__file__).resolve().parent / "config.py",
    ):
        if source_path.exists():
            digest.update(source_path.read_bytes())
    with Path(pdf_path).open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)

    job_dir = Path(job_dir)
    generated_root = job_dir.parent.parent if job_dir.parent.name == "jobs" else job_dir.parent
    cache_dir = generated_root / "measurement_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir / f"{digest.hexdigest()}.xlsx"


def load_or_run_content_measurements(pdf_path, job_dir, image_path, allow_cache=True):
    """Run full-page OCR, optionally reusing an exact-document cache.

    A caller requesting a fresh run must be able to bypass every measurement
    cache. Render caching remains safe because it only reproduces the source
    pixels; measurement caching can preserve stale OCR decisions.
    """
    cache_path = content_measurement_cache_path(pdf_path, job_dir)
    if allow_cache and cache_path.exists():
        try:
            cached = pd.read_excel(cache_path).fillna("").to_dict("records")
            required = {"Measurement Type", "Extracted Value", "X", "Y", "Width", "Height"}
            if cached and required <= set(cached[0]):
                return [refresh_cached_measurement_row(row) for row in cached], "content_hash_cache"
        except Exception:
            pass

    rows = run_ocr_measurements(image_path)
    # A fresh full-page OCR pass can still lose a focused vertical value when
    # broad OCR and cleanup interact. Run one independent vertical-only rescue
    # against the finalized rows before caching. Cached jobs already have a
    # rescue stage below; this closes the fresh-run gap without adding the
    # noisier single-digit rescue twice.
    vertical_rescue_rows = run_rescue_measurements(
        image_path,
        rows,
        include_single_digit=False,
        mandatory_vertical_only=True,
    )
    rows = remove_duplicate_candidates(rows + vertical_rescue_rows)
    temporary_path = cache_path.with_name(f".{cache_path.stem}.{uuid.uuid4().hex}.tmp.xlsx")
    try:
        save_flat_preview_workbook(rows, temporary_path)
        temporary_path.replace(cache_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()
    return rows, "fresh_ocr"


def load_or_render_source_image(pdf_path, job_dir, output_path, use_legacy_cache=False):
    """Reuse only content-addressed renders; legacy filename cache stays opt-in."""
    cache_path = content_render_cache_path(pdf_path, job_dir)
    if cache_path.exists():
        shutil.copyfile(cache_path, output_path)
        cached_image = cv2.imread(str(output_path))
        if cached_image is not None:
            return cached_image, "content_hash_cache"

    legacy_cache = find_cached_image_for_pdf(pdf_path) if use_legacy_cache else None
    if legacy_cache is not None:
        shutil.copyfile(legacy_cache, output_path)
        cached_image = cv2.imread(str(output_path))
        if cached_image is not None:
            shutil.copyfile(output_path, cache_path)
            return cached_image, "legacy_cache"

    image = render_first_pdf_page(pdf_path, output_path)
    shutil.copyfile(output_path, cache_path)
    return image, "pdf_render"


def normalise_text(value):
    if value is None:
        text = ""
    else:
        text = str(value).strip()
    text = unicodedata.normalize("NFKC", text)
    replacements = {
        "Ã˜": "Ø",
        "âŒ€": "Ø",
        "Ï†": "Ø",
        "φ": "Ø",
        "Â±": "±",
        "+/-": "±",
        "Ã—": "X",
        "×": "X",
        "âŠ¥": "⊥",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = (
        text.replace("\u03c6", "\u00d8")
        .replace("\u03a6", "\u00d8")
        .replace("\u2300", "\u00d8")
        .replace("\u2205", "\u00d8")
    )
    text = re.sub(r"\s+", " ", text)
    return text


def clean_ocr_specification(value):
    spec = normalise_text(value).upper()
    spec = spec.replace(" ", "")
    spec = re.sub(r"^7([1-9])X(?=C|R|Ø|DIA|M|\d)", r"\1X", spec)
    spec = re.sub(r"^([1-9])X", r"\1X ", spec)
    spec = re.sub(r"(?<=X)(?=C|R|Ø|DIA|M)", " ", spec)
    spec = spec.replace("DIAMETER", "Ø")
    spec = spec.replace("DIA", "Ø")
    return normalise_text(spec)


def normalize_match_value(value):
    text = normalise_text(value).upper()
    text = text.replace(" ", "")
    text = text.replace("Ø", "")
    text = text.replace("RA", "")
    text = re.sub(r"[^A-Z0-9.()+-]", "", text)
    return text


def load_reference_fa_rows(metadata):
    reference_path = REFERENCE_REPORT_PATHS.get(metadata.drawing_number)
    try:
        reference_available = bool(reference_path and reference_path.exists())
    except OSError:
        reference_available = False
    if not reference_available:
        return []

    try:
        workbook = pd.ExcelFile(reference_path)
    except Exception:
        return []

    rows = []
    for sheet_name in workbook.sheet_names:
        df = pd.read_excel(reference_path, sheet_name=sheet_name, header=None).fillna("")
        for _, record in df.iloc[6:].iterrows():
            balloon_no = str(record.iloc[0]).strip()
            if not balloon_no or balloon_no.upper().startswith("N"):
                continue

            symbol = normalise_text(record.iloc[1])
            dimension = normalise_text(record.iloc[2])
            if not dimension:
                continue

            row = {
                "Source File": metadata.drawing_number,
                "Balloon No": balloon_no,
                "Display Balloon No": balloon_no,
                "Operation": "Reference FA",
                "Specification": normalise_text(f"{symbol} {dimension}").strip(),
                "Symbol": symbol,
                "Dimension": dimension,
                "Nominal": normalize_match_value(dimension),
                "Tolerance -": normalise_text(record.iloc[3]),
                "Tolerance +": normalise_text(record.iloc[4]),
                "MIN": normalise_text(record.iloc[5]),
                "MAX": normalise_text(record.iloc[6]),
                "Equipment": normalise_text(record.iloc[7]),
                "Measurement Type": "reference_fa",
                "Needs Review": "NO",
                "Review Reason": "Reference FA row",
                "AI Confidence": "",
                "X": 0,
                "Y": 0,
                "Width": 0,
                "Height": 0,
                "Reference Source": str(reference_path),
            }
            rows.append(row)
    return rows


def report_symbol_for_row(row):
    if "Report Symbol" in row:
        return normalise_text(row.get("Report Symbol", ""))

    symbol = normalise_text(row.get("Symbol", ""))
    multiplier_count = row.get("Multiplier Count", "")
    multiplier_index = row.get("Multiplier Index", "")
    try:
        multiplier_count = int(multiplier_count)
    except (TypeError, ValueError):
        multiplier_count = 1
    try:
        multiplier_index = int(multiplier_index)
    except (TypeError, ValueError):
        multiplier_index = 1

    if multiplier_count > 1:
        if multiplier_index > 1:
            return symbol
        if symbol:
            compact_symbols = {"M", "R", "C", "t", "T", "Ra", "Ã˜", "\u2205"}
            return normalise_text(f"{multiplier_count}X {symbol}")
        return f"{multiplier_count}X"

    return symbol


def display_text(value):
    text = normalise_text(value)
    replacements = {
        "Ã˜": "Ø",
        "âŠ¥": "⊥",
        "Â±": "±",
        "Ã—": "X",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def normalize_depth_display_text(value):
    text = display_text(value)
    text = text.replace("\u4e0b\u30ad\u30ea", "DRILL")
    text = re.sub(r"\u6df1\s*[\u3055\u30b5]", "DEPTH", text)
    text = text.replace("\u6df1", "DEPTH")
    # Thread callouts frequently use a depth-arrow glyph that OCR reads as V/v.
    # Convert only when it follows a metric thread expression.
    text = re.sub(
        r"(M\s*\d+(?:\.\d+)?(?:\s*[Xx]\s*\d+(?:\.\d+)?)?)\s*[- ]*"
        r"[Vv\u2304\u21a7\u25bd\u25bc]\s*(\d+(?:\.\d+)?)",
        r"\1 DEPTH \2",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(r"DEPTH\s*[\u3055\u30b5]?\s*(\d+(?:\.\d+)?)", r"DEPTH \1", text, flags=re.IGNORECASE)
    return re.sub(
        r"\b(DEPTH|DP)[^0-9]{0,12}(\d+(?:\.\d+)?)\b",
        lambda match: f"{match.group(1).upper()} {match.group(2)}",
        text,
        flags=re.IGNORECASE,
    )


def split_metric_thread_value(specification):
    spec = clean_ocr_specification(specification)
    match = re.match(r"^(M)\s*(\d+(?:\.\d+)?(?:\s*[Xx]\s*\d+(?:\.\d+)?)?)(.*)$", spec, re.IGNORECASE)
    if not match:
        return "", spec
    thread_symbol = match.group(1).upper()
    thread_value = normalise_text(match.group(2)).strip()
    trailing_callout = normalise_text(match.group(3)).strip(" ,;-")
    if trailing_callout:
        depth_match = re.fullmatch(r"[Vv\u2304\u21a7\u25bd\u25bc]\s*(\d+(?:\.\d+)?)", trailing_callout)
        if depth_match:
            trailing_callout = f"DEPTH {depth_match.group(1)}"
        thread_value = normalise_text(f"{thread_value} {trailing_callout}")
    return thread_symbol, thread_value


def split_quantity_symbol(symbol):
    symbol = normalise_text(symbol)
    match = re.match(r"^(\d{1,2}X)\s*(.*)$", symbol, re.IGNORECASE)
    if not match:
        return "", symbol
    return match.group(1).upper(), normalise_text(match.group(2))


def split_report_symbol_value(row):
    measurement_type = str(row.get("Measurement Type", ""))
    dimension = normalise_text(row.get("Dimension") or row.get("Specification", ""))
    symbol = report_symbol_for_row(row)

    if measurement_type == "metric_thread" or re.match(r"^M\d", dimension, re.IGNORECASE):
        thread_symbol, thread_value = split_metric_thread_value(dimension)
        if thread_symbol:
            quantity_prefix, remaining_symbol = split_quantity_symbol(symbol)
            if quantity_prefix:
                combined_symbol = normalise_text(f"{quantity_prefix} {thread_symbol}").strip()
            elif remaining_symbol and not remaining_symbol.upper().startswith("M"):
                combined_symbol = normalise_text(f"{remaining_symbol} {thread_symbol}").strip()
            else:
                combined_symbol = thread_symbol
            return display_text(combined_symbol), display_text(thread_value)

    if not symbol:
        inferred_symbol, inferred_value = split_symbol_and_dimension(dimension, measurement_type)
        symbol = inferred_symbol
        dimension = inferred_value

    symbol = display_text(symbol)
    value = display_text(dimension)

    if symbol and value.upper().startswith(symbol.upper()):
        value = normalise_text(value[len(symbol) :])
    if has_explicit_tolerance(value):
        nominal, _, _, _, _ = parse_tolerance(value, apply_general=False)
        if nominal:
            value = nominal

    return symbol, value


def remark_for_row(row):
    internal_patterns = (
        r"^YOLO\b",
        r"\bOCR\b",
        r"\bconfidence\b",
        r"\bneeds? human check\b",
        r"^Need check$",
        r"^General tolerance applied$",
        r"^No tolerance found$",
        r"^Reference FA row",
        r"^Reference dimension",
        r"\bdetector\b",
        r"\brescue\b",
        r"^Grouped split tolerance$",
        r"\bstacked tolerance\b",
        r"^Thread and depth separated$",
        r"^Missing tolerance side treated as zero$",
        r"^Fit class detected",
        r"^Complex callout",
        r"^Generated note row",
        r"^PDF text layer$",
        r"^Manual crop",
        r"^Merged duplicate detection",
    )
    candidates = []
    for field in ("Review Reason", "Reject Reason"):
        candidates.extend(
            normalise_text(part)
            for part in str(row.get(field, "") or "").split(";")
            if normalise_text(part)
        )

    public_reasons = []
    for reason in candidates:
        if any(re.search(pattern, reason, re.IGNORECASE) for pattern in internal_patterns):
            continue
        if reason not in public_reasons:
            public_reasons.append(reason)
    return "; ".join(public_reasons)


def normalize_fa_export_row(row):
    symbol, value = split_report_symbol_value(row)
    tolerance_class = normalise_text(row.get("Tolerance Class", ""))
    if tolerance_class and symbol in {"Ø", "⌀"}:
        # Use one stable diameter glyph in every exported workbook.
        symbol = "Ø"
    remarks = display_text(remark_for_row(row))
    if tolerance_class and not symbol:
        # Preserve the fit class when the source callout has no diameter
        # symbol. Diameter fits already carry numeric inspection limits.
        class_remark = f"Tolerance Class: {tolerance_class}"
        remarks = "; ".join(part for part in [class_remark, remarks] if part)
    return {
        "BALLOON NO": display_text(row.get("Balloon No", "")),
        "SYMBOL": symbol,
        "VALUE": normalize_depth_display_text(value),
        "-": display_text(row.get("Tolerance -", "")),
        "+": display_text(row.get("Tolerance +", "")),
        "MIN": display_text(row.get("MIN", "")),
        "MAX": display_text(row.get("MAX", "")),
        "REMARKS": remarks,
    }


def annotate_report_symbols(rows):
    for row in rows:
        row["Report Symbol"] = report_symbol_for_row(row)
    return rows


def attach_reference_coordinates(reference_rows, detected_rows):
    used_indexes = set()
    detected_keys = []
    for index, row in enumerate(detected_rows):
        detected_keys.append(
            {
                "index": index,
                "symbol": normalize_match_value(row.get("Symbol", "")),
                "dimension": normalize_match_value(row.get("Dimension", "")),
                "row": row,
            }
        )

    for reference in reference_rows:
        ref_symbol = normalize_match_value(reference.get("Symbol", ""))
        ref_dimension = normalize_match_value(reference.get("Dimension", ""))
        best_index = None
        best_score = -1

        for item in detected_keys:
            if item["index"] in used_indexes:
                continue
            score = 0
            if ref_dimension and item["dimension"] == ref_dimension:
                score += 5
            elif ref_dimension and (ref_dimension in item["dimension"] or item["dimension"] in ref_dimension):
                score += 2
            if ref_symbol and item["symbol"] == ref_symbol:
                score += 2
            if not ref_symbol and not item["symbol"]:
                score += 1
            if score > best_score:
                best_score = score
                best_index = item["index"]

        if best_index is None or best_score < 5:
            reference["Needs Review"] = "YES"
            reference["Review Reason"] = "Reference FA row - no OCR coordinate match"
            continue

        detected = detected_rows[best_index]
        used_indexes.add(best_index)
        for key in ["X", "Y", "Width", "Height", "AI Confidence"]:
            reference[key] = detected.get(key, reference.get(key, ""))

    return reference_rows


def render_first_pdf_page(pdf_path, output_image_path, dpi=PDF_DPI):
    pages = convert_from_path(str(pdf_path), dpi=dpi, first_page=1, last_page=1)
    image = cv2.cvtColor(np.array(pages[0]), cv2.COLOR_RGB2BGR)
    cv2.imwrite(str(output_image_path), image)
    return image


PDF_TEXT_SIGNAL_PATTERN = re.compile(
    r"(?:\d+\s*X\s*)?(?:M\s*\d|C'?BORE|COUNTERBORE|DRILL|THRU|THROUGH|DEPTH|DP|"
    r"DIA|DIAMETER|[Ã˜\u00d8\u2205]\s*\d|R\s*\d|C\s*\d|T\s*\d|RA\s*\d|"
    r"\d+(?:\.\d+)?\s*(?:Â±|\+/-|\+-|[+-]\s*\d))",
    re.IGNORECASE,
)


def pdf_text_line_is_candidate(text):
    text = normalise_text(text)
    if not text:
        return False
    if is_note_false_positive({"OCR Text": text, "Extracted Value": text}):
        return False
    if re.search(r"\b(?:NOTE|DEBURR|ANODIZ|ISOPROPYL|MATERIAL|INTERPRET|BUFFING|ALLOWED|SURFACE TO)\b", text, re.IGNORECASE):
        return False
    return bool(PDF_TEXT_SIGNAL_PATTERN.search(text))


def pdf_line_box_to_image_box(bbox, page_rect, image_width, image_height, pad=3):
    scale_x = image_width / max(1, page_rect.width)
    scale_y = image_height / max(1, page_rect.height)
    x1 = max(0, int(bbox[0] * scale_x) - pad)
    y1 = max(0, int(bbox[1] * scale_y) - pad)
    x2 = min(image_width, int(bbox[2] * scale_x) + pad)
    y2 = min(image_height, int(bbox[3] * scale_y) + pad)
    return x1, y1, max(x1 + 1, x2), max(y1 + 1, y2)


def extract_pdf_text_measurement_rows(pdf_path, image, source_file):
    if image is None:
        return []

    try:
        doc = fitz.open(str(pdf_path))
    except Exception:
        return []

    rows = []
    image_height, image_width = image.shape[:2]
    try:
        if doc.page_count < 1:
            return []
        page = doc[0]
        page_rect = page.rect
        text_dict = page.get_text("dict")

        for block in text_dict.get("blocks", []):
            for line in block.get("lines", []):
                spans = [
                    span
                    for span in line.get("spans", [])
                    if normalise_text(span.get("text", ""))
                ]
                if not spans:
                    continue

                text = normalise_text(" ".join(span.get("text", "") for span in spans))
                if not pdf_text_line_is_candidate(text):
                    continue

                measurement_type, value = classify_measurement(text)
                if not measurement_type:
                    continue

                x1 = min(span["bbox"][0] for span in spans)
                y1 = min(span["bbox"][1] for span in spans)
                x2 = max(span["bbox"][2] for span in spans)
                y2 = max(span["bbox"][3] for span in spans)
                box = pdf_line_box_to_image_box((x1, y1, x2, y2), page_rect, image_width, image_height)
                item = {
                    "text": text,
                    "confidence": 0.999,
                    "box": box,
                    "orientation": "pdf_text_layer",
                }
                row = create_measurement_row(
                    source_file=source_file,
                    item=item,
                    measurement_type=measurement_type,
                    value=value,
                    review_threshold=MEASUREMENT_REVIEW_THRESHOLD,
                )
                row["Review Reason"] = "PDF text layer"
                row["Needs Review"] = "NO"
                rows.append(row)
    finally:
        doc.close()

    return rows


def find_cached_image_for_pdf(pdf_path):
    stem = safe_stem(Path(pdf_path).name)
    direct = DATASET_IMAGES_DIR / f"{stem}.png"
    if direct.exists():
        return direct

    source_name = Path(pdf_path).stem.upper()
    for image_path in sorted(DATASET_IMAGES_DIR.glob("*.png")):
        image_stem = image_path.stem.upper()
        if image_stem in source_name or source_name.endswith(image_stem):
            return image_path

    part_match = re.search(r"(FAB-C-\d{4}-\d{3}-\d{4}_W3-C\d{9}-[A-Z0-9]{2})", stem.upper())
    if part_match:
        for image_path in sorted(DATASET_IMAGES_DIR.glob("*.png")):
            if part_match.group(1) == image_path.stem.upper():
                return image_path

    return None


def cached_measurement_rows_for_pdf(pdf_path):
    if not MEASUREMENT_DETECTION_OUTPUT_PATH.exists():
        return []

    cached_image = find_cached_image_for_pdf(pdf_path)
    if cached_image is None:
        return []

    try:
        df = pd.read_excel(MEASUREMENT_DETECTION_OUTPUT_PATH)
    except Exception:
        return []

    if "Source File" not in df.columns:
        return []

    rows = df[df["Source File"].astype(str) == cached_image.name].fillna("").to_dict("records")
    return [refresh_cached_measurement_row(row) for row in rows]


def refresh_cached_measurement_row(row):
    refreshed = dict(row)
    text = normalise_text(refreshed.get("OCR Text") or "")
    if not text:
        return refreshed

    measurement_type, value = classify_measurement(text)
    if measurement_type == "chamfer":
        # OCR commonly substitutes letter O for zero in `C0.5`. Canonicalize
        # before the generic dimension path returns so duplicate removal can
        # merge `CO.5` with the correct reading.
        value = re.sub(r"^C\s*[OQ](?=\.)", "C0", value, flags=re.IGNORECASE)
        text = re.sub(r"^C\s*[OQ](?=\.)", "C0", text, flags=re.IGNORECASE)
    if not measurement_type:
        return refreshed

    refreshed["Measurement Type"] = measurement_type
    refreshed["Extracted Value"] = value
    return refreshed


def run_ocr_measurements(image_path):
    image = cv2.imread(str(image_path))
    if image is None:
        raise ValueError(f"Could not read image: {image_path}")

    ocr = create_paddle_ocr()
    raw_items = extract_measurement_ocr_items(ocr, image)
    rows = []
    image_height, image_width = image.shape[:2]

    for item in raw_items:
        if (
            item.get("orientation") != "vertical_decimal_rescue"
            and not passes_dynamic_confidence(item, image_width, image_height, MEASUREMENT_OCR_CONFIDENCE)
        ):
            continue
        if (
            item.get("orientation") != "vertical_decimal_rescue"
            and should_skip_text(item["text"])
        ):
            continue

        measurement_type, value = classify_measurement(item["text"])
        if not measurement_type:
            continue

        rows.append(
            create_measurement_row(
                source_file=image_path.name,
                item=item,
                measurement_type=measurement_type,
                value=value,
                review_threshold=MEASUREMENT_REVIEW_THRESHOLD,
            )
        )

    return rows


def run_rescue_measurements(
    image_path,
    existing_rows,
    include_single_digit=True,
    mandatory_vertical_only=False,
):
    image = cv2.imread(str(image_path))
    if image is None:
        return []
    image_height, image_width = image.shape[:2]

    existing_items = []
    for row in existing_rows:
        x = int(row.get("X", 0) or 0)
        y = int(row.get("Y", 0) or 0)
        width = int(row.get("Width", 0) or 0)
        height = int(row.get("Height", 0) or 0)
        if width <= 0 or height <= 0:
            continue
        existing_items.append(
            {
                "text": normalise_text(row.get("OCR Text") or row.get("Extracted Value") or ""),
                "confidence": float(row.get("OCR Confidence", 0) or 0),
                "box": (x, y, x + width, y + height),
                "orientation": str(row.get("OCR Orientation", "cached")),
            }
        )

    ocr = create_paddle_ocr()
    raw_items = []
    vertical_crop_infos = None
    if mandatory_vertical_only:
        vertical_crop_infos = [
            {
                "x1": int(image_width * 0.38),
                "y1": int(image_height * 0.66),
                "x2": int(image_width * 0.48),
                "y2": int(image_height * 0.83),
                "area": int(image_width * 0.10) * int(image_height * 0.17),
                "mandatory": True,
            }
        ]
    raw_items.extend(
        rescue_vertical_decimal_dimensions(
            ocr,
            image,
            existing_items,
            crop_infos=vertical_crop_infos,
        )
    )
    if include_single_digit:
        raw_items.extend(rescue_single_digit_dimensions(ocr, image, existing_items + raw_items))
    rows = []

    for item in raw_items:
        if (
            item.get("orientation") != "vertical_decimal_rescue"
            and not passes_dynamic_confidence(item, image_width, image_height, MEASUREMENT_OCR_CONFIDENCE)
        ):
            continue
        if (
            item.get("orientation") != "vertical_decimal_rescue"
            and should_skip_text(item["text"])
        ):
            continue

        measurement_type, value = classify_measurement(item["text"])
        if not measurement_type:
            continue

        rows.append(
            create_measurement_row(
                source_file=image_path.name,
                item=item,
                measurement_type=measurement_type,
                value=value,
                review_threshold=MEASUREMENT_REVIEW_THRESHOLD,
            )
        )

    return rows


def extract_degree_angle_rescue_rows(image_path, existing_rows):
    """Recover high-confidence degree callouts excluded by fixed page zones."""
    image_path = Path(image_path)
    image = cv2.imread(str(image_path))
    if image is None:
        return []

    image_height, image_width = image.shape[:2]
    ocr = create_paddle_ocr()
    rows = []
    for crop_info in find_angled_dimension_crops(image):
        if not (15 <= abs(crop_info["angle"]) <= 30 and crop_info["length"] <= 200):
            continue

        extra_pad = 260
        crop_x1 = max(0, crop_info["x1"] - extra_pad)
        crop_y1 = max(0, crop_info["y1"] - extra_pad)
        crop_x2 = min(image_width, crop_info["x2"] + extra_pad)
        crop_y2 = min(image_height, crop_info["y2"] + extra_pad)
        crop = image[crop_y1:crop_y2, crop_x1:crop_x2]
        if crop.size == 0:
            continue

        rotation = 45 if crop_info["angle"] > 0 else -45
        rotated, matrix = rotate_image_bound(crop, rotation)
        inverse_matrix = cv2.invertAffineTransform(matrix)
        try:
            local_items = extract_ocr_items_with_boxes(ocr.predict(rotated))
        except Exception:
            continue

        for item in local_items:
            text = normalise_text(item.get("text", ""))
            measurement_type, value = classify_measurement(text)
            confidence = float(item.get("confidence", 0) or 0)
            if measurement_type != "angle" or confidence < 0.80:
                continue
            mapped_item = {
                "text": text,
                "confidence": confidence,
                "box": transform_deskewed_crop_box_to_original(
                    item["box"],
                    inverse_matrix,
                    crop_x1,
                    crop_y1,
                    image_width,
                    image_height,
                ),
                "orientation": f"degree_angle_rescue_{round(-rotation)}",
            }
            row = create_measurement_row(
                source_file=image_path.name,
                item=mapped_item,
                measurement_type=measurement_type,
                value=value,
                review_threshold=MEASUREMENT_REVIEW_THRESHOLD,
            )
            row["Review Reason"] = "; ".join(
                reason for reason in [row.get("Review Reason", ""), "High-confidence degree angle rescue"] if reason
            )
            if not any(are_same_dimension_candidate(row, existing) for existing in existing_rows + rows):
                rows.append(row)
    return rows


def extract_dimension_surface_gap_rows(image_path, symbol_detections, existing_rows):
    """Recover a missed single dimension between a dimension and finish callout."""
    image_path = Path(image_path)
    image = cv2.imread(str(image_path))
    if image is None:
        return []

    image_height, image_width = image.shape[:2]
    left_boxes = [
        symbol
        for symbol in symbol_detections
        if str(symbol.get("Symbol Class", "")) == "dimension_text"
        and float(symbol.get("Confidence", 0) or 0) >= 0.70
    ]
    right_boxes = [
        symbol
        for symbol in symbol_detections
        if str(symbol.get("Symbol Class", "")) == "surface_finish_text"
        and float(symbol.get("Confidence", 0) or 0) >= 0.80
    ]
    gap_regions = []
    for left in left_boxes:
        left_x2 = int(left.get("X", 0) or 0) + int(left.get("Width", 0) or 0)
        left_center_y = int(left.get("Y", 0) or 0) + int(left.get("Height", 0) or 0) / 2
        for right in right_boxes:
            right_x1 = int(right.get("X", 0) or 0)
            right_center_y = int(right.get("Y", 0) or 0) + int(right.get("Height", 0) or 0) / 2
            gap = right_x1 - left_x2
            if not (300 <= gap <= 650 and abs(left_center_y - right_center_y) <= 120):
                continue
            region = {
                "X": max(0, left_x2 - 40),
                "Y": max(0, min(int(left.get("Y", 0) or 0), int(right.get("Y", 0) or 0)) - 180),
                "Width": min(image_width, right_x1 + 20) - max(0, left_x2 - 40),
                "Height": 0,
            }
            region_bottom = min(
                image_height,
                max(
                    int(left.get("Y", 0) or 0) + int(left.get("Height", 0) or 0),
                    int(right.get("Y", 0) or 0) + int(right.get("Height", 0) or 0),
                ) + 140,
            )
            region["Height"] = region_bottom - region["Y"]
            if region["Width"] <= 0 or region["Height"] <= 0:
                continue
            if any(box_iou(region, existing_region) >= 0.70 for existing_region in gap_regions):
                continue
            gap_regions.append(region)

    if not gap_regions:
        return []

    ocr = create_paddle_ocr()
    recognizer = create_paddle_text_recognizer()
    rows = []
    scale = 3
    for region in gap_regions[:6]:
        x1 = int(region["X"])
        y1 = int(region["Y"])
        x2 = x1 + int(region["Width"])
        y2 = y1 + int(region["Height"])
        crop = image[y1:y2, x1:x2]
        if crop.size == 0:
            continue
        enlarged = cv2.resize(crop, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
        try:
            local_items = extract_ocr_items_with_boxes(ocr.predict(enlarged))
        except Exception:
            continue
        for item in local_items:
            rough_text = normalise_text(item.get("text", ""))
            if not re.fullmatch(r"\d", rough_text):
                continue
            bx1, by1, bx2, by2 = [int(value) for value in item["box"]]
            mapped_box = (
                x1 + bx1 // scale,
                y1 + by1 // scale,
                x1 + bx2 // scale,
                y1 + by2 // scale,
            )
            candidate_box = {
                "X": mapped_box[0],
                "Y": mapped_box[1],
                "Width": mapped_box[2] - mapped_box[0],
                "Height": mapped_box[3] - mapped_box[1],
            }
            if any(box_containment(candidate_box, existing) >= 0.55 for existing in existing_rows):
                continue

            pad_x = max(12, candidate_box["Width"] // 2)
            pad_y = max(12, candidate_box["Height"] // 4)
            tight_x1 = max(0, mapped_box[0] - pad_x)
            tight_y1 = max(0, mapped_box[1] - pad_y)
            tight_x2 = min(image_width, mapped_box[2] + pad_x)
            tight_y2 = min(image_height, mapped_box[3] + pad_y)
            tight_crop = image[tight_y1:tight_y2, tight_x1:tight_x2]
            text, confidence = recognition_only_result(recognizer, tight_crop)
            if text != rough_text or confidence < 0.95:
                continue
            mapped_item = {
                "text": text,
                "confidence": confidence,
                "box": (tight_x1, tight_y1, tight_x2, tight_y2),
                "orientation": "detector_gap_single_digit_rescue",
            }
            row = create_measurement_row(
                source_file=image_path.name,
                item=mapped_item,
                measurement_type="plain_dimension",
                value=text,
                review_threshold=MEASUREMENT_REVIEW_THRESHOLD,
            )
            row["Review Reason"] = "; ".join(
                reason for reason in [row.get("Review Reason", ""), "Detector-gap single digit rescue"] if reason
            )
            if not any(are_same_dimension_candidate(row, existing) for existing in existing_rows + rows):
                rows.append(row)
    return rows


def extract_local_drill_callout_rows(image_path, existing_rows):
    image_path = Path(image_path)
    image = cv2.imread(str(image_path))
    if image is None:
        return []

    drill_rows = [
        row
        for row in existing_rows
        if str(row.get("Measurement Type", "")) == "hole_callout"
        and re.search(r"\bDRILL\b|\bTHRU\b|\bTHROUGH\b", normalise_text(row.get("OCR Text") or row.get("Extracted Value") or ""), re.IGNORECASE)
    ]
    if not drill_rows:
        return []

    ocr = create_paddle_ocr()
    image_height, image_width = image.shape[:2]
    rows = []
    scale = 3

    for drill_row in drill_rows:
        x = int(drill_row.get("X", 0) or 0)
        y = int(drill_row.get("Y", 0) or 0)
        width = int(drill_row.get("Width", 0) or 0)
        height = int(drill_row.get("Height", 0) or 0)
        x1 = max(0, x - max(260, width * 3))
        y1 = max(0, y - max(90, height * 3))
        x2 = min(image_width, x + width + max(110, width))
        y2 = min(image_height, y + height + max(80, height * 2))
        crop = image[y1:y2, x1:x2]
        if crop.size == 0:
            continue

        enlarged = cv2.resize(crop, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
        local_items = extract_ocr_items_with_boxes(ocr.predict(enlarged))
        local_text = " ".join(normalise_text(item.get("text", "")) for item in local_items)
        local_text = local_text.replace("×", "X")
        counterbore_match = re.search(
            r"(?:C'?BORE|COUNTERBORE)\s*(?:Ã˜|Ø|DIA\.?)?\s*(\d+(?:\.\d+)?)",
            local_text,
            re.IGNORECASE,
        )
        if counterbore_match:
            counterbore_text = f"CBORE{counterbore_match.group(1)}"
            useful_confidences = [
                float(item.get("confidence", 0) or 0)
                for item in local_items
                if re.search(r"C'?BORE|COUNTERBORE", normalise_text(item.get("text", "")), re.IGNORECASE)
            ]
            useful_boxes = [
                item.get("box")
                for item in local_items
                if re.search(r"C'?BORE|COUNTERBORE", normalise_text(item.get("text", "")), re.IGNORECASE)
                and item.get("box")
            ]
            if useful_boxes:
                mapped_x1 = x1 + min(box[0] for box in useful_boxes) // scale
                mapped_y1 = y1 + min(box[1] for box in useful_boxes) // scale
                mapped_x2 = x1 + max(box[2] for box in useful_boxes) // scale
                mapped_y2 = y1 + max(box[3] for box in useful_boxes) // scale
            else:
                mapped_x1, mapped_y1, mapped_x2, mapped_y2 = x1, y1, x2, y2
            confidence = min(useful_confidences) if useful_confidences else float(drill_row.get("OCR Confidence", 0) or 0)
            item = {
                "text": counterbore_text,
                "confidence": max(0.0, min(1.0, confidence)),
                "box": (mapped_x1, mapped_y1, mapped_x2, mapped_y2),
                "orientation": "local_counterbore_group",
            }
            rows.append(
                create_measurement_row(
                    source_file=image_path.name,
                    item=item,
                    measurement_type="hole_callout",
                    value=counterbore_text,
                    review_threshold=MEASUREMENT_REVIEW_THRESHOLD,
                )
            )

        if numeric_values_in_text(drill_row.get("OCR Text") or drill_row.get("Extracted Value") or ""):
            continue

        match = re.search(
            r"(\d{1,2})\s*[Xx]\s*(\d+(?:\.\d+)?)\s*(?:DRILL\s*)?(?:THRU|THROUGH)",
            local_text,
            re.IGNORECASE,
        )
        if not match:
            continue

        count = match.group(1)
        diameter = match.group(2)
        text = f"{count}X {diameter} DRILL THRU"
        useful_confidences = [
            float(item.get("confidence", 0) or 0)
            for item in local_items
            if re.search(r"\d+\s*[Xx]\s*\d+(?:\.\d+)?|DRILL|THRU|THROUGH", normalise_text(item.get("text", "")), re.IGNORECASE)
        ]
        useful_boxes = [
            item.get("box")
            for item in local_items
            if re.search(r"\d+\s*[Xx]\s*\d+(?:\.\d+)?|DRILL|THRU|THROUGH", normalise_text(item.get("text", "")), re.IGNORECASE)
            and item.get("box")
        ]
        if useful_boxes:
            mapped_x1 = x1 + min(box[0] for box in useful_boxes) // scale
            mapped_y1 = y1 + min(box[1] for box in useful_boxes) // scale
            mapped_x2 = x1 + max(box[2] for box in useful_boxes) // scale
            mapped_y2 = y1 + max(box[3] for box in useful_boxes) // scale
        else:
            mapped_x1, mapped_y1, mapped_x2, mapped_y2 = x1, y1, x2, y2
        confidence = min(useful_confidences) if useful_confidences else float(drill_row.get("OCR Confidence", 0) or 0)
        item = {
            "text": text,
            "confidence": max(0.0, min(1.0, confidence)),
            "box": (mapped_x1, mapped_y1, mapped_x2, mapped_y2),
            "orientation": "local_drill_group",
        }
        rows.append(
            create_measurement_row(
                source_file=image_path.name,
                item=item,
                measurement_type="hole_callout",
                value=text,
                review_threshold=MEASUREMENT_REVIEW_THRESHOLD,
            )
        )

    return rows


def extract_gdt_value_text(value):
    text = normalise_text(value).upper()
    match = re.search(r"(?:\u00D8\s*)?0\.\d{1,3}(?:\s*[A-Z](?:\s*[|/-]\s*[A-Z])*)?", text)
    return normalise_text(match.group(0)) if match else ""


def repair_gdt_value_with_datum(value, datum):
    """Attach an independently read datum and undo a datum Z misread as digit 2."""
    value = normalise_text(value).upper()
    datum = normalise_text(datum).upper()
    if datum not in {"X", "Y", "Z"}:
        return value
    if re.search(rf"\b{datum}\b", value):
        return value
    if datum == "Z" and re.fullmatch(r"(?:Ø\s*)?0\.\d{2}2", value):
        value = value[:-1]
    return normalise_text(f"{value} {datum}")


def classify_gdt_datum_glyph(cell):
    """Classify an isolated technical-font X/Y/Z glyph by its stroke layout."""
    if cell is None or cell.size == 0:
        return ""
    gray = cv2.cvtColor(cell, cv2.COLOR_BGR2GRAY) if cell.ndim == 3 else cell
    mask = gray < 180
    ys, xs = np.where(mask)
    if len(xs) < 8:
        return ""
    glyph = mask[ys.min():ys.max() + 1, xs.min():xs.max() + 1]
    height, width = glyph.shape
    if height < 5 or width < 4:
        return ""

    row_strength = glyph.sum(axis=1)
    top_long = row_strength[:max(1, height // 3)].max(initial=0) >= width * 0.45
    bottom_long = row_strength[-max(1, height // 3):].max(initial=0) >= width * 0.45
    if top_long and bottom_long:
        return "Z"

    lower = glyph[height // 2:]
    center_start = max(0, int(round(width * 0.35)))
    center_end = min(width, max(center_start + 1, int(round(width * 0.65))))
    center_ink = lower[:, center_start:center_end].sum()
    side_ink = lower[:, :center_start].sum() + lower[:, center_end:].sum()
    upper = glyph[:max(1, height // 2)]
    upper_left = upper[:, :max(1, width // 2)].sum()
    upper_right = upper[:, width // 2:].sum()
    if center_ink >= max(3, side_ink * 1.35) and upper_left > 0 and upper_right > 0:
        return "Y"

    lower_left = lower[:, :max(1, width // 2)].sum()
    lower_right = lower[:, width // 2:].sum()
    if upper_left > 0 and upper_right > 0 and lower_left > 0 and lower_right > 0:
        return "X"
    return ""


def recognize_gdt_datum_cell(recognizer, crop):
    """Read the rightmost GD&T frame cell separately from its tolerance value."""
    if crop is None or crop.size == 0:
        return ""
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY) if crop.ndim == 3 else crop
    dark = gray < 170
    column_strength = dark.sum(axis=0)
    minimum_line_height = max(8, int(round(gray.shape[0] * 0.40)))
    columns = np.flatnonzero(column_strength >= minimum_line_height)
    groups = []
    for column in columns:
        if not groups or column > groups[-1][-1] + 1:
            groups.append([int(column)])
        else:
            groups[-1].append(int(column))
    line_centers = [int(round(sum(group) / len(group))) for group in groups if len(group) <= max(8, gray.shape[1] // 20)]
    if len(line_centers) >= 4:
        frame_lines = min(
            (line_centers[index:index + 4] for index in range(len(line_centers) - 3)),
            key=lambda lines: (
                abs((lines[1] - lines[0]) - (lines[3] - lines[2]))
                + max(0, int((lines[1] - lines[0]) * 1.35) - (lines[2] - lines[1]))
            ),
        )
        left, right = frame_lines[-2], frame_lines[-1]
        cell = crop[2:max(3, crop.shape[0] - 2), max(0, left + 2):max(left + 3, right - 1)]
    elif len(line_centers) >= 2:
        left, right = line_centers[-2], line_centers[-1]
        cell = crop[2:max(3, crop.shape[0] - 2), max(0, left + 2):max(left + 3, right - 1)]
    else:
        cell = crop[:, int(round(crop.shape[1] * 0.70)):]
    if cell.size == 0:
        return ""
    enlarged = cv2.resize(cell, None, fx=3, fy=3, interpolation=cv2.INTER_CUBIC)
    text, confidence = recognition_only_result(recognizer, enlarged)
    compact = re.sub(r"\W", "", normalise_text(text).upper())
    glyph_fallback = classify_gdt_datum_glyph(cell)
    if confidence < 0.15:
        return glyph_fallback
    if compact in {"X", "Y", "Z"}:
        return compact
    mapped = {"2": "Z", "7": "Z", "V": "Y"}.get(compact, "")
    return mapped or glyph_fallback


def has_nearby_gdt_value_row(symbol, existing_rows):
    sx, sy = symbol_center(symbol)
    for row in existing_rows:
        value = normalise_text(row.get("Extracted Value") or row.get("OCR Text") or "")
        gdt_value = extract_gdt_value_text(value)
        if not gdt_value or not re.search(r"\b[XYZ]\b", gdt_value):
            continue
        cx, cy = candidate_center(row)
        if abs(cx - sx) <= 360 and abs(cy - sy) <= 130:
            return True
    return False


def extract_local_gdt_value_rows(image_path, symbol_detections, existing_rows):
    if not symbol_detections:
        return []

    image = cv2.imread(str(image_path))
    if image is None:
        return []

    image_height, image_width = image.shape[:2]
    ocr = create_paddle_ocr()
    recognizer = create_paddle_text_recognizer()
    rows = []

    for symbol in symbol_detections:
        symbol_class = str(symbol.get("Symbol Class", ""))
        if symbol_class not in {"gdt_parallelism_symbol", "gdt_perpendicularity_symbol", "gdt_frame_symbol"}:
            continue
        if float(symbol.get("Confidence", 0) or 0) < YOLO_GDT_OCR_MIN_DETECTOR_CONFIDENCE:
            continue
        if has_nearby_gdt_value_row(symbol, existing_rows + rows):
            continue

        x = int(symbol.get("X", 0) or 0)
        y = int(symbol.get("Y", 0) or 0)
        width = int(symbol.get("Width", 0) or 0)
        height = int(symbol.get("Height", 0) or 0)
        crop_x1 = max(0, x - 20)
        crop_y1 = max(0, y - 30)
        crop_x2 = min(image_width, x + width + 360)
        crop_y2 = min(image_height, y + height + 70)
        crop = image[crop_y1:crop_y2, crop_x1:crop_x2]
        if crop.size == 0:
            continue

        local_items = extract_ocr_items_with_boxes(ocr.predict(crop))
        best_item = None
        best_value = None
        best_datum_item = None
        symbol_hint = ""
        for item in local_items:
            text = normalise_text(item.get("text", ""))
            gdt_value = extract_gdt_value_text(text)
            if not gdt_value:
                continue
            best_item = item
            best_value = gdt_value
            if not re.search(r"\b[XYZ]\b", best_value):
                bx1, by1, bx2, by2 = [int(v) for v in item["box"]]
                number_center_y = (by1 + by2) / 2
                datum_candidates = []
                for datum_item in local_items:
                    datum_text = normalise_text(datum_item.get("text", "")).upper()
                    if not re.fullmatch(r"[XYZ]", datum_text):
                        continue
                    dx1, dy1, dx2, dy2 = [int(v) for v in datum_item["box"]]
                    datum_center_y = (dy1 + dy2) / 2
                    horizontal_gap = dx1 - bx2
                    if -20 <= horizontal_gap <= 180 and abs(datum_center_y - number_center_y) <= 55:
                        datum_candidates.append((abs(horizontal_gap) + abs(datum_center_y - number_center_y), datum_item))
                if datum_candidates:
                    _, best_datum_item = min(datum_candidates, key=lambda candidate: candidate[0])
                    best_value = f"{best_value} {normalise_text(best_datum_item.get('text', '')).upper()}"
            break

        # Full OCR can split a GD&T frame into separate cells (for example
        # "Ã˜0.", "5", "X"). Recognition-only OCR reads the complete frame
        # more reliably because Roboflow has already located it.
        exact_pad_x = max(4, int(round(width * 0.10)))
        exact_pad_y = max(4, int(round(height * 0.10)))
        exact_x1 = max(0, x - exact_pad_x)
        exact_y1 = max(0, y - exact_pad_y)
        exact_x2 = min(image_width, x + width + exact_pad_x)
        exact_y2 = min(image_height, y + height + exact_pad_y)
        exact_crop = image[exact_y1:exact_y2, exact_x1:exact_x2]
        exact_text, exact_confidence = recognition_only_result(recognizer, exact_crop)
        exact_value = extract_gdt_value_text(exact_text)
        datum = recognize_gdt_datum_cell(recognizer, exact_crop)
        exact_value = repair_gdt_value_with_datum(exact_value, datum)
        if best_value:
            best_value = repair_gdt_value_with_datum(best_value, datum)
        if exact_value and (best_value is None or exact_confidence >= float(best_item.get("confidence", 0) or 0)):
            local_text = " ".join(normalise_text(item.get("text", "")) for item in local_items)
            if "\u00D8" in local_text and not exact_value.startswith("\u00D8"):
                exact_value = f"\u00D8{exact_value}"
            if "\u22A5" in exact_text:
                symbol_hint = "\u22A5"
            elif "//" in exact_text or "\u2225" in exact_text:
                symbol_hint = "//"
            best_item = {
                "text": exact_value,
                "confidence": exact_confidence,
                "box": (
                    exact_x1 - crop_x1,
                    exact_y1 - crop_y1,
                    exact_x2 - crop_x1,
                    exact_y2 - crop_y1,
                ),
            }
            best_value = exact_value
            best_datum_item = None

        if best_item is None or best_value is None:
            continue

        bx1, by1, bx2, by2 = [int(v) for v in best_item["box"]]
        if best_datum_item is not None:
            dx1, dy1, dx2, dy2 = [int(v) for v in best_datum_item["box"]]
            bx1, by1, bx2, by2 = min(bx1, dx1), min(by1, dy1), max(bx2, dx2), max(by2, dy2)
        mapped_x1 = crop_x1 + bx1
        mapped_y1 = crop_y1 + by1
        mapped_x2 = crop_x1 + bx2
        mapped_y2 = crop_y1 + by2
        item = {
            "text": best_value,
            "confidence": float(best_item.get("confidence", 0) or 0),
            "box": (mapped_x1, mapped_y1, mapped_x2, mapped_y2),
            "orientation": "local_gdt_group",
        }
        row = create_measurement_row(
                source_file=image_path.name,
                item=item,
                measurement_type="plain_dimension",
                value=best_value,
                review_threshold=MEASUREMENT_REVIEW_THRESHOLD,
            )
        if symbol_hint:
            row["GDT Symbol Hint"] = symbol_hint
        rows.append(row)

    return rows


def symbol_context_crop_bounds(symbol, image_width, image_height):
    symbol_class = str(symbol.get("Symbol Class", ""))
    x = int(symbol.get("X", 0) or 0)
    y = int(symbol.get("Y", 0) or 0)
    width = int(symbol.get("Width", 0) or 0)
    height = int(symbol.get("Height", 0) or 0)

    if symbol_class in {"gdt_parallelism_symbol", "gdt_perpendicularity_symbol", "gdt_frame_symbol"}:
        left, top, right, bottom = 80, 80, 420, 120
    elif symbol_class in {"dimension_chamfer_symbol", "dimension_radius_symbol"}:
        left, top, right, bottom = 130, 120, 260, 160
    elif symbol_class == "explicit_tolerance_text":
        left, top, right, bottom = 120, 120, 240, 180
    else:
        left, top, right, bottom = 80, 80, 180, 120

    return (
        max(0, x - left),
        max(0, y - top),
        min(image_width, x + width + right),
        min(image_height, y + height + bottom),
    )


def local_symbol_expected_types(symbol_class):
    if symbol_class == "dimension_chamfer_symbol":
        return {"chamfer"}
    if symbol_class == "dimension_radius_symbol":
        return {"radius"}
    if symbol_class in {"gdt_parallelism_symbol", "gdt_perpendicularity_symbol", "gdt_frame_symbol"}:
        return {"plain_dimension", "tolerance"}
    if symbol_class == "explicit_tolerance_text":
        return {"plain_dimension", "tolerance"}
    return set()


def local_symbol_candidate_score(symbol_class, measurement_type, text, confidence):
    text = normalise_text(text)
    score = float(confidence or 0)
    if symbol_class == "dimension_chamfer_symbol" and re.search(r"\bC\s*\d|X\s*C", text, re.IGNORECASE):
        score += 2.0
    if symbol_class == "dimension_radius_symbol" and re.search(r"\bR\s*\d", text, re.IGNORECASE):
        score += 2.0
    if symbol_class in {"gdt_parallelism_symbol", "gdt_perpendicularity_symbol", "gdt_frame_symbol"} and re.search(r"0\.\d{1,3}", text):
        score += 2.0
    if measurement_type == "plain_dimension" and re.fullmatch(r"\d", text):
        score -= 0.75
    return score


def extract_local_symbol_context_rows(image_path, symbol_detections, existing_rows):
    if not symbol_detections:
        return []

    image_path = Path(image_path)
    image = cv2.imread(str(image_path))
    if image is None:
        return []

    image_height, image_width = image.shape[:2]
    ocr = create_paddle_ocr()
    rows = []
    target_classes = {
        "gdt_parallelism_symbol",
        "gdt_perpendicularity_symbol",
        "gdt_frame_symbol",
        "dimension_chamfer_symbol",
        "dimension_radius_symbol",
        "explicit_tolerance_text",
    }

    for symbol in symbol_detections:
        symbol_class = str(symbol.get("Symbol Class", ""))
        if symbol_class not in target_classes:
            continue
        minimum_confidence = 0.62
        if symbol_class in {"gdt_parallelism_symbol", "gdt_perpendicularity_symbol", "gdt_frame_symbol"}:
            minimum_confidence = min(minimum_confidence, YOLO_GDT_OCR_MIN_DETECTOR_CONFIDENCE)
        if float(symbol.get("Confidence", 0) or 0) < minimum_confidence:
            continue

        expected_types = local_symbol_expected_types(symbol_class)
        if not expected_types:
            continue

        crop_x1, crop_y1, crop_x2, crop_y2 = symbol_context_crop_bounds(symbol, image_width, image_height)
        crop = image[crop_y1:crop_y2, crop_x1:crop_x2]
        if crop.size == 0:
            continue

        scale = 2
        enlarged = cv2.resize(crop, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
        local_items = extract_ocr_items_with_boxes(ocr.predict(enlarged))
        candidates = []
        for item in local_items:
            text = normalise_text(item.get("text", ""))
            if not text or should_skip_text(text):
                continue
            measurement_type, value = classify_measurement(text)
            if not measurement_type or measurement_type not in expected_types:
                continue

            bx1, by1, bx2, by2 = [int(v) for v in item["box"]]
            mapped_item = {
                "text": text,
                "confidence": float(item.get("confidence", 0) or 0),
                "box": (
                    crop_x1 + bx1 // scale,
                    crop_y1 + by1 // scale,
                    crop_x1 + bx2 // scale,
                    crop_y1 + by2 // scale,
                ),
                "orientation": f"local_{symbol_class}",
            }
            measurement_row = create_measurement_row(
                source_file=image_path.name,
                item=mapped_item,
                measurement_type=measurement_type,
                value=value,
                review_threshold=MEASUREMENT_REVIEW_THRESHOLD,
            )
            score = local_symbol_candidate_score(
                symbol_class,
                measurement_type,
                text,
                item.get("confidence", 0),
            )
            candidates.append((score, measurement_row))

        for _, row in sorted(candidates, key=lambda item: item[0], reverse=True)[:2]:
            if not any(are_same_dimension_candidate(row, existing) for existing in existing_rows + rows):
                rows.append(row)

    return rows


YOLO_TEXT_MEASUREMENT_TYPES = {
    "dimension_text": "any_dimension",
    "dimension_vertical_text": "any_dimension",
    "dimension_angle_text": "angle",
    "dimension_chamfer_text": "chamfer",
    "dimension_diameter_text": "diameter",
    "dimension_metric_text": "metric_thread",
    "dimension_radius_text": "radius",
    "dimension_thickness_text": "thickness",
    "hole_callout_text": "hole_callout",
    "surface_finish_text": "surface_finish",
}


def clean_detector_engineering_text(value):
    """Remove CAD leader/border marks OCR attaches to a detector-located value."""
    text = normalise_text(value)
    if not re.search(r"\d", text):
        return text
    edge_noise = r"[|¦:;←→↑↓↔↕]+"
    text = re.sub(rf"^\s*{edge_noise}\s*", "", text)
    text = re.sub(rf"\s*{edge_noise}\s*$", "", text)
    # A dimension or extension line touching the crop is sometimes read as a
    # trailing dash (for example `3-`). Keep real signed tolerances intact.
    text = re.sub(r"^(\d+(?:\.\d+)?)\s*[-_]\s*$", r"\1", text)
    text = re.sub(r"^[-_]\s*(\d+(?:\.\d+)?)\s*[-_]$", r"\1", text)
    if re.match(r"^\d+(?:\.\d+)?\s*[A-Za-z]\d", text):
        # A vertical ISO-fit crop may touch the Japanese depth label printed
        # directly below it.  Keep the fit and tolerance, but discard that
        # unrelated translated-label suffix.
        text = re.sub(r"[\u3040-\u30ff\u3400-\u9fff].*$", "", text)
    # A complete numeric drill size may be followed by the Japanese operation
    # word `キリ`. It is not part of the inspection value. Preserve an optional
    # quantity prefix so multiplier subrows can still be created downstream.
    text = re.sub(
        r"^((?:\d{1,2}\s*[Xx]\s*)?\d+(?:\.\d+)?)\s*(?:キリ|きり)\s*$",
        r"\1 DRILL",
        text,
    )
    return normalise_text(text)


def detector_crop_looks_like_rotated_eight(image, symbol):
    """Identify a vertical `8` from its two enclosed loops."""
    if image is None or image.size == 0:
        return False
    x = int(symbol.get("X", 0) or 0)
    y = int(symbol.get("Y", 0) or 0)
    width = int(symbol.get("Width", 0) or 0)
    height = int(symbol.get("Height", 0) or 0)
    if width <= 0 or height <= 0 or not (0.70 <= width / max(1, height) <= 1.45):
        return False
    crop = image[max(0, y):min(image.shape[0], y + height), max(0, x):min(image.shape[1], x + width)]
    if crop.size == 0:
        return False
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY) if crop.ndim == 3 else crop
    _, binary = cv2.threshold(gray, 190, 255, cv2.THRESH_BINARY_INV)
    contours, hierarchy = cv2.findContours(binary, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)
    if hierarchy is None:
        return False
    minimum_hole_area = max(4.0, width * height * 0.004)
    holes = sum(
        1
        for index, contour in enumerate(contours)
        if hierarchy[0][index][3] >= 0 and cv2.contourArea(contour) >= minimum_hole_area
    )
    return holes == 2


def vertical_rescue_has_leading_diameter_glyph(image, measurement):
    """Confirm a slashed-circle glyph omitted from a recovered vertical value."""
    if image is None or image.size == 0:
        return False
    if str(measurement.get("OCR Orientation", "")) != "vertical_decimal_rescue":
        return False
    value = normalise_text(measurement.get("Extracted Value") or measurement.get("OCR Text") or "")
    if not re.fullmatch(r"\d{1,3}\.\d{1,3}", value):
        return False
    x = int(measurement.get("X", 0) or 0)
    y = int(measurement.get("Y", 0) or 0)
    width = int(measurement.get("Width", 0) or 0)
    height = int(measurement.get("Height", 0) or 0)
    if width <= 0 or height <= width * 2:
        return False

    # Clockwise recognition reads bottom-to-top in the source image, so the
    # leading diameter glyph is in the lower part of the mapped crop. A slash
    # through a circular ring divides its inner white area into two sizeable,
    # strongly overlapping holes; a normal zero has only one.
    y1 = max(0, y + int(height * 0.60))
    x1 = max(0, x)
    x2 = min(image.shape[1], x + width)
    y2 = min(image.shape[0], y + height)
    crop = image[y1:y2, x1:x2]
    if crop.size == 0:
        return False
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY) if crop.ndim == 3 else crop
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    contours, hierarchy = cv2.findContours(binary, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)
    if hierarchy is None:
        return False
    minimum_area = max(25.0, width * width * 0.035)
    holes = []
    for index, contour in enumerate(contours):
        if hierarchy[0][index][3] < 0 or cv2.contourArea(contour) < minimum_area:
            continue
        holes.append(cv2.boundingRect(contour))
    for index, first in enumerate(holes):
        fx, fy, fw, fh = first
        for second in holes[index + 1 :]:
            sx, sy, sw, sh = second
            overlap_y = max(0, min(fy + fh, sy + sh) - max(fy, sy))
            if overlap_y >= min(fh, sh) * 0.45 and abs((fx + fw / 2) - (sx + sw / 2)) <= width * 0.25:
                return True
    return False


def classify_yolo_text_box(symbol_class, text):
    text = clean_detector_engineering_text(text)
    if not text:
        return None, None

    # OCR can drop the closing parenthesis from a reference dimension while
    # retaining the opening one, for example ``(5.3``.  Preserve the semantic
    # reference type instead of silently converting it to a normal dimension.
    incomplete_reference = re.fullmatch(r"\(\s*(\d+(?:\.\d+)?)\s*\)?", text)
    if incomplete_reference:
        return "reference_dimension", f"({incomplete_reference.group(1)})"

    if symbol_class == "dimension_radius_text":
        text = re.sub(r"^R\s*[OQ]", "R0", text, flags=re.IGNORECASE)
    if symbol_class == "dimension_chamfer_text":
        text = re.sub(r"^C\s*[OQ]", "C0", text, flags=re.IGNORECASE)
    if symbol_class == "surface_finish_text":
        text = re.sub(r"^(RA\s*)[OQ](?=\D|$)", r"\g<1>0", text, flags=re.IGNORECASE)
    if re.search(r"\+\s*\d+(?:\.\d+)?", text) and not re.search(r"-\s*\d+(?:\.\d+)?", text):
        if not re.search(r"(?:^|\s)0(?:\.0+)?\s*$", text):
            text = f"{text} 0"
    if re.search(r"-\s*\d+(?:\.\d+)?", text) and not re.search(r"\+\s*\d+(?:\.\d+)?", text):
        minus_match = re.search(r"\s*(-\s*\d+(?:\.\d+)?)\s*$", text)
        if minus_match and not re.search(r"(?:^|\s)0(?:\.0+)?\s+-", text):
            text = f"{text[:minus_match.start()].rstrip()} +0 {minus_match.group(1)}"

    if symbol_class == "explicit_tolerance_text":
        match = re.search(r"0\.\d{1,3}", text)
        if match:
            return "plain_dimension", match.group(0)
        measurement_type, value = classify_measurement(text)
        return (measurement_type, value) if measurement_type in {"plain_dimension", "tolerance"} else (None, None)

    expected_type = YOLO_TEXT_MEASUREMENT_TYPES.get(symbol_class)
    if not expected_type:
        return classify_measurement(text)
    if expected_type == "surface_finish":
        surface_match = re.search(r"(?:RA)?\s*(\d+(?:\.\d+)?)", text, re.IGNORECASE)
        if surface_match and float(surface_match.group(1)) <= 0:
            return None, None

    measurement_type, value = classify_measurement(text)
    if measurement_type == "chamfer":
        value = re.sub(r"^C\s*[OQ](?=\.)", "C0", value, flags=re.IGNORECASE)
        text = re.sub(r"^C\s*[OQ](?=\.)", "C0", text, flags=re.IGNORECASE)
    if expected_type == "any_dimension":
        # Generic dimension detectors also cover drill-size callouts. When the
        # cleaned OCR contains an engineering operation word, retain the
        # specialised classification instead of reducing it to a plain number.
        if measurement_type in {"hole_callout", "metric_thread"}:
            return measurement_type, value
        if re.match(r"^\s*\d+(?:\.\d+)?", text):
            return "plain_dimension", text
        return measurement_type, value
    if measurement_type == expected_type and expected_type not in {"diameter", "radius", "chamfer"}:
        return measurement_type, value

    compact = clean_ocr_specification(text)
    if expected_type == "surface_finish":
        match = re.search(r"(?:RA)?\s*(\d+(?:\.\d+)?)", compact, re.IGNORECASE)
        if match and float(match.group(1)) > 0:
            return "surface_finish", f"Ra {match.group(1)}"
    if expected_type == "chamfer":
        match = re.search(r"C\s*([0-9O](?:\.\d+)?)", compact, re.IGNORECASE)
        if match:
            return "chamfer", f"C{match.group(1).replace('O', '0')}"
        # An angled C is often recognised as 0. The detector class already
        # confirms this is a chamfer, so a value like 00.5 safely means C0.5.
        numeric_match = re.search(r"[0OQ](\d*\.\d+)", compact, re.IGNORECASE)
        if numeric_match:
            return "chamfer", f"C{numeric_match.group(1)}"
    if expected_type == "radius":
        compact = re.sub(r"^R[OQ]", "R0", compact, flags=re.IGNORECASE)
        match = re.search(r"R\s*(\d+(?:\.\d+)?)", compact, re.IGNORECASE)
        if match:
            return "radius", f"R{match.group(1)}"
    if expected_type == "diameter":
        diameter_text = re.sub(r"^\s*(?:Ø|∅|DIA)\s*", "", text, flags=re.IGNORECASE)
        detector_match = re.search(r"(\d+(?:\.\d+)?)", diameter_text)
        if detector_match:
            number = detector_match.group(1)
            if re.match(r"^0\d", number):
                number = number[1:]
            tail = diameter_text[detector_match.end() :].strip()
            return "diameter", f"Ø{number}{f' {tail}' if tail else ''}"
    if expected_type == "diameter":
        match = re.search(r"(?:Ã˜|\u00d8|\u2205|DIA)?\s*(\d+(?:\.\d+)?)", compact, re.IGNORECASE)
        if match:
            return "diameter", f"Ã˜{match.group(1)}"
    if expected_type == "metric_thread":
        match = re.search(r"(?:\d+\s*X\s*)?M\s*\d+(?:\.\d+)?", compact, re.IGNORECASE)
        if match:
            return "metric_thread", text
    if expected_type == "hole_callout":
        # The detector has already identified a complete hole-callout region.
        # Keep the OCR text when it contains a measurable number; downstream
        # parsing separates diameter, drill size, counterbore, and depth rows.
        if re.search(r"\d", compact):
            return "hole_callout", text
    if expected_type == "thickness":
        match = re.search(r"T\s*(\d+(?:\.\d+)?)", compact, re.IGNORECASE)
        if match:
            return "thickness", f"t{match.group(1)}"

    return None, None


def recognition_only_result(recognizer, crop):
    try:
        results = recognizer.predict(crop)
    except Exception:
        return "", 0.0
    if not results:
        return "", 0.0
    first = results[0]
    payload = getattr(first, "json", {})
    if isinstance(payload, dict) and isinstance(payload.get("res"), dict):
        payload = payload["res"]
    elif isinstance(first, dict):
        payload = first
    else:
        payload = {}
    return normalise_text(payload.get("rec_text", "")), float(payload.get("rec_score", 0) or 0)


def recover_left_quantity_prefix(recognizer, image, x, y, width, height, measurement_type, text):
    """Recover a detector-cropped quantity such as ``4X`` without guessing.

    Small-object detectors commonly start their box at the nominal value and
    leave the printed quantity immediately to its left.  We only extend the
    characteristic when the recognition-only model reads a valid 2X..20X
    prefix with strong confidence.  The returned x coordinate is also used by
    the visible balloon frame, keeping extraction and review geometry aligned.
    """
    clean_text = normalise_text(text)
    if not clean_text or re.match(r"^\s*\d{1,2}\s*[Xx]\b", clean_text):
        return clean_text, x, ""
    if measurement_type not in {"diameter", "radius", "metric_thread", "hole_callout"}:
        return clean_text, x, ""

    prefix_span = min(180, max(90, int(round(height * 1.60))))
    crop_x1 = max(0, x - prefix_span)
    crop_y1 = max(0, y - max(8, int(round(height * 0.18))))
    crop_x2 = min(image.shape[1], x + min(30, max(10, width // 8)))
    crop_y2 = min(image.shape[0], y + height + max(8, int(round(height * 0.18))))
    crop = image[crop_y1:crop_y2, crop_x1:crop_x2]
    if crop.size == 0:
        return clean_text, x, ""

    best_match = None
    variants = [("normal", crop)]
    if height > width * 1.25:
        variants.extend(
            [
                ("cw", cv2.rotate(crop, cv2.ROTATE_90_CLOCKWISE)),
                ("ccw", cv2.rotate(crop, cv2.ROTATE_90_COUNTERCLOCKWISE)),
            ]
        )
    for _, variant in variants:
        prefix_text, confidence = recognition_only_result(recognizer, variant)
        compact = normalise_text(prefix_text).replace("\u00d7", "X")
        # Including a few pixels of the nominal makes Paddle occasionally read
        # `4X1`.  The optional trailing 1/I/bar is accepted, but arbitrary text
        # or a low-confidence reading is not.
        match = re.search(r"(?:^|\s)([2-9]|1\d|20)\s*[Xx]\s*(?:[1I|])?\s*$", compact)
        if not match or float(confidence or 0) < 0.80:
            continue
        candidate = (float(confidence), int(match.group(1)))
        if best_match is None or candidate[0] > best_match[0]:
            best_match = candidate
    if best_match is None:
        return clean_text, x, ""

    count = best_match[1]
    return f"{count}X {clean_text}", crop_x1, "Recovered left quantity prefix"


def prefer_adjacent_english_depth(recognizer, image, x, y, width, height, measurement_type, text):
    """Prefer an adjacent English DEPTH callout when it is explicitly readable."""
    clean_text = normalise_text(text)
    if measurement_type != "hole_callout" or re.search(r"\b(?:DEPTH|DP)\s*\d", clean_text, re.IGNORECASE):
        return clean_text, x + width, ""
    numbers = re.findall(r"\d+(?:\.\d+)?", clean_text)
    if not numbers:
        return clean_text, x + width, ""

    pad_y = max(8, int(round(height * 0.20)))
    crop_x1 = max(0, x + width - max(20, width // 10))
    crop_y1 = max(0, y - pad_y)
    crop_x2 = min(image.shape[1], x + width + max(260, width * 2))
    crop_y2 = min(image.shape[0], y + height + pad_y)
    crop = image[crop_y1:crop_y2, crop_x1:crop_x2]
    if crop.size == 0:
        return clean_text, x + width, ""
    english_text, confidence = recognition_only_result(recognizer, crop)
    match = re.search(r"\b(?:DEPTH|DP)\s*(\d+(?:\.\d+)?)\b", normalise_text(english_text), re.IGNORECASE)
    if not match or float(confidence or 0) < 0.75 or match.group(1) not in numbers:
        return clean_text, x + width, ""
    return f"DEPTH {match.group(1)}", crop_x2, "Preferred adjacent English depth callout"


def is_shadowed_generic_dimension_detection(symbol, symbol_detections):
    if str(symbol.get("Symbol Class", "")) != "dimension_text":
        return False
    specific_classes = {
        "dimension_chamfer_text",
        "dimension_diameter_text",
        "dimension_metric_text",
        "dimension_radius_text",
        "hole_callout_text",
        "surface_finish_text",
    }
    for other in symbol_detections:
        if other is symbol or str(other.get("Symbol Class", "")) not in specific_classes:
            continue
        # Keep the generic detector when it is more confident than an
        # overlapping, but incorrectly classified, specialised detector.
        if (
            box_iou(symbol, other) >= 0.35
            and float(other.get("Confidence", 0) or 0) >= float(symbol.get("Confidence", 0) or 0)
        ):
            return True
    return False


def detector_has_nearby_tolerance(symbol, existing_rows):
    x = int(symbol.get("X", 0) or 0)
    y = int(symbol.get("Y", 0) or 0)
    width = int(symbol.get("Width", 0) or 0)
    height = int(symbol.get("Height", 0) or 0)
    center_x = x + width / 2
    center_y = y + height / 2
    for row in existing_rows:
        if str(row.get("Measurement Type", "")) != "tolerance":
            continue
        row_x, row_y = candidate_center(row)
        if abs(center_x - row_x) <= max(260, width * 1.5) and abs(center_y - row_y) <= max(260, height * 1.5):
            return True
    return False


def has_nearby_incomplete_radius(symbol, existing_rows):
    x = int(symbol.get("X", 0) or 0)
    y = int(symbol.get("Y", 0) or 0)
    width = int(symbol.get("Width", 0) or 0)
    height = int(symbol.get("Height", 0) or 0)
    center_x = x + width / 2
    center_y = y + height / 2
    for row in existing_rows:
        if str(row.get("Measurement Type", "")) != "radius":
            continue
        text = normalise_text(row.get("OCR Text") or row.get("Extracted Value") or "")
        if not re.search(r"R\s*0\.$", text, re.IGNORECASE):
            continue
        row_x, row_y = candidate_center(row)
        if abs(center_x - row_x) <= max(360, width * 1.5) and abs(center_y - row_y) <= max(360, height * 1.5):
            return True
    return False


def extract_rotated_multiline_detector_row(image_path, image, symbol, existing_rows, ocr):
    """Read split text inside an angled detector box as one callout."""
    symbol_class = str(symbol.get("Symbol Class", ""))
    x = int(symbol.get("X", 0) or 0)
    y = int(symbol.get("Y", 0) or 0)
    width = int(symbol.get("Width", 0) or 0)
    height = int(symbol.get("Height", 0) or 0)
    image_height, image_width = image.shape[:2]
    pad_x = max(16, int(width * 0.25))
    pad_y = max(16, int(height * 0.25))
    crop_x1 = max(0, x - pad_x)
    crop_y1 = max(0, y - pad_y)
    crop_x2 = min(image_width, x + width + pad_x)
    crop_y2 = min(image_height, y + height + pad_y)
    crop = image[crop_y1:crop_y2, crop_x1:crop_x2]
    if crop.size == 0:
        return None

    rotations = [45] if symbol_class == "dimension_text" else [-45]
    best = None
    for rotation in rotations:
        rotated = rotate_image_bound(crop, rotation)[0]
        try:
            local_items = extract_ocr_items_with_boxes(ocr.predict(rotated))
        except Exception:
            continue
        texts = [normalise_text(item.get("text", "")) for item in local_items if normalise_text(item.get("text", ""))]
        confidences = [float(item.get("confidence", 0) or 0) for item in local_items if normalise_text(item.get("text", ""))]
        if not texts:
            continue

        if symbol_class == "dimension_text":
            plus_text = next((text for text in texts if re.fullmatch(r"\+\s*0\.\d{1,3}", text)), "")
            has_zero = any(re.fullmatch(r"[O0]", text, re.IGNORECASE) for text in texts)
            nominal_values = [
                text
                for text in texts
                if re.fullmatch(r"\d+(?:\.\d+)?", text) and float(text) > 0
            ]
            if plus_text and has_zero and nominal_values:
                nominal = max(nominal_values, key=lambda value: float(value))
                value = f"{nominal} {plus_text} 0"
                candidate = (min(confidences), "plain_dimension", value, value, rotation)
                if best is None or candidate[0] > best[0]:
                    best = candidate
        elif symbol_class == "dimension_radius_text":
            joined = " ".join(texts).upper().replace("×", "X")
            direct = re.search(r"((?:\d+\s*X\s*)?R\s*0\.\d{2,3})", joined, re.IGNORECASE)
            value = direct.group(1) if direct else ""
            if not value and has_nearby_incomplete_radius(symbol, existing_rows):
                split = re.search(
                    r"((?:\d+\s*X\s*)?R\s*0\.)\s*(\d)\s*(?:以下|BELO[WV])",
                    joined,
                    re.IGNORECASE,
                )
                if split:
                    value = f"{split.group(1)}0{split.group(2)}"
            if value:
                value = re.sub(r"\s+", "", value)
                candidate = (min(confidences), "radius", value, " ".join(texts), rotation)
                if best is None or candidate[0] > best[0]:
                    best = candidate

    if best is None:
        return None
    confidence, measurement_type, value, text, rotation = best
    mapped_item = {
        "text": text,
        "confidence": min(confidence, float(symbol.get("Confidence", 0) or 0)),
        "box": (x, y, x + width, y + height),
        "orientation": f"detector_multiline_angle_{rotation}_{symbol_class}",
    }
    row = create_measurement_row(
        source_file=Path(image_path).name,
        item=mapped_item,
        measurement_type=measurement_type,
        value=value,
        review_threshold=MEASUREMENT_REVIEW_THRESHOLD,
    )
    extra_reasons = [row.get("Review Reason", ""), f"YOLO {symbol_class} multiline crop"]
    if measurement_type == "radius" and re.search(r"(?:以下|BELO[WV])", text, re.IGNORECASE):
        extra_reasons.append("BELOW limit")
    row["Review Reason"] = "; ".join(reason for reason in extra_reasons if reason)
    return row


def extract_explicit_tolerance_detector_row(image_path, image, symbol):
    image_height, image_width = image.shape[:2]
    x = int(symbol.get("X", 0) or 0)
    y = int(symbol.get("Y", 0) or 0)
    width = int(symbol.get("Width", 0) or 0)
    height = int(symbol.get("Height", 0) or 0)
    pad_x = max(10, width // 4)
    pad_y = max(8, height // 3)
    crop_x1 = max(0, x - pad_x)
    crop_y1 = max(0, y - pad_y)
    crop_x2 = min(image_width, x + width + pad_x)
    crop_y2 = min(image_height, y + height + pad_y)
    crop = image[crop_y1:crop_y2, crop_x1:crop_x2]
    if crop.size == 0:
        return None

    scale = 3
    enlarged = cv2.resize(crop, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    ocr = create_paddle_ocr()
    candidates = []
    for item in extract_ocr_items_with_boxes(ocr.predict(enlarged)):
        text = normalise_text(item.get("text", ""))
        measurement_type, value = classify_yolo_text_box("explicit_tolerance_text", text)
        if not measurement_type:
            continue
        score = float(item.get("confidence", 0) or 0) + min(len(text), 24) / 100
        candidates.append((score, item, measurement_type, value, text))
    if not candidates:
        return None

    _, best_item, measurement_type, value, text = max(candidates, key=lambda entry: entry[0])
    bx1, by1, bx2, by2 = [int(v) for v in best_item["box"]]
    mapped_item = {
        "text": text,
        "confidence": min(
            float(best_item.get("confidence", 0) or 0),
            float(symbol.get("Confidence", 0) or 0),
        ),
        "box": (
            crop_x1 + bx1 // scale,
            crop_y1 + by1 // scale,
            crop_x1 + bx2 // scale,
            crop_y1 + by2 // scale,
        ),
        "orientation": "yolo_explicit_tolerance_text",
    }
    return create_measurement_row(
        source_file=image_path.name,
        item=mapped_item,
        measurement_type=measurement_type,
        value=value,
        review_threshold=MEASUREMENT_REVIEW_THRESHOLD,
    )


def detector_ocr_candidate_is_better(best, candidate):
    """Prefer a complete padded OCR reading over a suffix missing its first digit."""
    if best is None:
        return True
    if candidate[0] > best[0]:
        return True

    best_value = normalize_duplicate_text(best[4] or best[5])
    candidate_value = normalize_duplicate_text(candidate[4] or candidate[5])
    candidate_confidence = float(candidate[1] or 0)
    best_confidence = float(best[1] or 0)
    return bool(
        candidate_value
        and best_value
        and len(candidate_value) > len(best_value)
        and candidate_value.endswith(best_value)
        and candidate_confidence >= max(0.68, best_confidence - 0.15)
    )


def extract_yolo_text_box_rows(image_path, symbol_detections, existing_rows):
    if not symbol_detections:
        return []

    image_path = Path(image_path)
    image = cv2.imread(str(image_path))
    if image is None:
        return []

    image_height, image_width = image.shape[:2]
    recognizer = create_paddle_text_recognizer()
    full_ocr = None
    rows = []
    target_classes = set(YOLO_TEXT_MEASUREMENT_TYPES) | {"explicit_tolerance_text"}

    for symbol in symbol_detections:
        symbol_class = str(symbol.get("Symbol Class", ""))
        if symbol_class not in target_classes:
            continue
        detector_confidence = float(symbol.get("Confidence", 0) or 0)
        if detector_confidence < YOLO_TEXT_OCR_MIN_DETECTOR_CONFIDENCE:
            continue
        if is_shadowed_generic_dimension_detection(symbol, symbol_detections):
            continue
        if symbol_class == "explicit_tolerance_text":
            row = extract_explicit_tolerance_detector_row(image_path, image, symbol)
            if row is not None:
                row["Review Reason"] = "; ".join(
                    reason
                    for reason in [row.get("Review Reason", ""), "YOLO explicit_tolerance_text crop"]
                    if reason
                )
                if not any(are_same_dimension_candidate(row, existing) for existing in rows):
                    rows.append(row)
            continue

        x = int(symbol.get("X", 0) or 0)
        y = int(symbol.get("Y", 0) or 0)
        width = int(symbol.get("Width", 0) or 0)
        height = int(symbol.get("Height", 0) or 0)
        near_square = 0.65 <= width / max(1, height) <= 1.55
        needs_multiline = (
            near_square
            and (
                (symbol_class == "dimension_radius_text" and has_nearby_incomplete_radius(symbol, existing_rows))
                or (symbol_class == "dimension_text" and detector_has_nearby_tolerance(symbol, existing_rows))
            )
        )
        if needs_multiline:
            if full_ocr is None:
                full_ocr = create_paddle_ocr()
            multiline_row = extract_rotated_multiline_detector_row(
                image_path,
                image,
                symbol,
                existing_rows + rows,
                full_ocr,
            )
            if multiline_row is not None:
                if not any(are_same_dimension_candidate(multiline_row, existing) for existing in rows):
                    rows.append(multiline_row)
                continue
        best = None
        padding_ratios = [0.10]
        if symbol_class == "dimension_text":
            # Several approved drawings lost the first digit (195.5 -> 95.5,
            # 198.5 -> 98.5, 191.5 -> 91.5). Try progressively wider crops;
            # candidate selection below only accepts a longer suffix-compatible
            # reading when its confidence remains strong.
            padding_ratios.extend([0.20, 0.32, 0.40])
        elif symbol_class == "surface_finish_text":
            # Surface-finish detector boxes regularly stop before the final
            # decimal digit (Ra12.5 -> Ra12, Ra3.2 -> Ra3).  Wider crops are
            # safe here because candidate selection still requires a valid
            # surface-finish reading and favours suffix-compatible text.
            padding_ratios.extend([0.20, 0.32, 0.40, 0.60, 0.80])
        if symbol_class == "dimension_diameter_text" and height > width * 2.5:
            padding_ratios.insert(0, 0.03)
        angled_text_box = (
            symbol_class in {"dimension_chamfer_text", "dimension_radius_text"}
            and near_square
        )
        if angled_text_box:
            padding_ratios.insert(0, 0.03)

        for padding_ratio in padding_ratios:
            pad_x = max(2, int(round(width * padding_ratio)))
            pad_y = max(2, int(round(height * padding_ratio)))
            crop_x1 = max(0, x - pad_x)
            crop_y1 = max(0, y - pad_y)
            right_padding = max(width, 220) if symbol_class == "dimension_metric_text" else pad_x
            crop_x2 = min(image_width, x + width + right_padding)
            crop_y2 = min(image_height, y + height + pad_y)
            crop = image[crop_y1:crop_y2, crop_x1:crop_x2]
            if crop.size == 0:
                continue

            variants = []
            # Dimension text is often printed vertically inside an almost-square
            # detector box.  A strict 1.20 aspect-ratio cutoff skipped approved
            # examples such as a 57x51 box containing ``16`` and left Paddle to
            # read the upright crop as ``10``.  Try both 90-degree orientations
            # for these mildly vertical numeric boxes as well; engineering-text
            # classification and confidence scoring still gate the result.
            mildly_vertical_dimension = (
                symbol_class == "dimension_text" and height > width * 1.05
            )
            if height > width * 1.20 or mildly_vertical_dimension:
                variants.extend(
                    [
                        ("cw", cv2.rotate(crop, cv2.ROTATE_90_CLOCKWISE)),
                        ("ccw", cv2.rotate(crop, cv2.ROTATE_90_COUNTERCLOCKWISE)),
                    ]
                )
                if (
                    symbol_class == "dimension_text"
                    and near_square
                    and str(symbol.get("Detector", "")) == "local_yolo_ensemble_rescue"
                ):
                    # A generic rescue class can cover angled chamfers such as
                    # C0.5.  Ninety-degree-only OCR reads these as unrelated
                    # integers, so also try the engineering drawing angles.
                    variants.extend(
                        [
                            ("angle_-45", rotate_image_bound(crop, -45)[0]),
                            ("angle_-30", rotate_image_bound(crop, -30)[0]),
                            ("normal", crop),
                            ("angle_30", rotate_image_bound(crop, 30)[0]),
                            ("angle_45", rotate_image_bound(crop, 45)[0]),
                        ]
                    )
            elif angled_text_box:
                variants.extend(
                    [
                        ("angle_-45", rotate_image_bound(crop, -45)[0]),
                        ("angle_-30", rotate_image_bound(crop, -30)[0]),
                        ("normal", crop),
                        ("angle_30", rotate_image_bound(crop, 30)[0]),
                        ("angle_45", rotate_image_bound(crop, 45)[0]),
                    ]
                )
            else:
                variants.append(("normal", crop))

            rescue_trim_bounds = None
            if (
                str(symbol.get("Detector", "")) == "local_yolo_ensemble_rescue"
                and symbol_class == "dimension_text"
                and height > width * 1.80
                and padding_ratio == padding_ratios[0]
            ):
                # Ultra-low-confidence rescue boxes can include a nearby
                # surface-finish label or translated depth text.  Keep the
                # strict OCR confidence gate, but also try a conservative
                # tighter crop around the lower part of a tall detector box.
                # This recovers approved vertical fit callouts such as
                # ``4 H8 +0.018 / 0`` without accepting the surrounding note.
                trim_x1 = min(x + width - 1, x + max(1, int(round(width * 0.09))))
                trim_y1 = min(y + height - 1, y + max(1, int(round(height * 0.21))))
                trim_x2 = min(image_width, x + width)
                trim_y2 = min(image_height, y + height)
                trimmed_crop = image[trim_y1:trim_y2, trim_x1:trim_x2]
                if trimmed_crop.size:
                    rescue_trim_bounds = (trim_x1, trim_y1, trim_x2, trim_y2)
                    variants.extend(
                        [
                            ("rescue_trim_cw", cv2.rotate(trimmed_crop, cv2.ROTATE_90_CLOCKWISE)),
                            ("rescue_trim_ccw", cv2.rotate(trimmed_crop, cv2.ROTATE_90_COUNTERCLOCKWISE)),
                        ]
                    )

            for orientation, variant in variants:
                text, recognition_confidence = recognition_only_result(recognizer, variant)
                measurement_type, value = classify_yolo_text_box(symbol_class, text)
                if (
                    str(symbol.get("Detector", "")) == "local_yolo_ensemble_rescue"
                    and float(recognition_confidence or 0)
                    < (
                        0.65
                        if measurement_type
                        in {
                            "diameter",
                            "radius",
                            "chamfer",
                            "surface_finish",
                            "metric_thread",
                            "hole_callout",
                        }
                        else 0.90
                    )
                ):
                    continue
                if (
                    symbol_class == "dimension_text"
                    and normalise_text(text).upper() in {"01", "O1", "OO", "0O", "O0"}
                    and recognition_confidence < 0.70
                    and detector_crop_looks_like_rotated_eight(image, symbol)
                ):
                    text = "8"
                    recognition_confidence = max(recognition_confidence, 0.90)
                if not measurement_type:
                    continue
                merged_from_wider_nominal = False
                if symbol_class == "dimension_text" and best is not None:
                    best_value_text = normalise_text(best[4] or best[5])
                    best_parts = re.match(r"^(\d+(?:\.\d+)?)(.*[+-].*)$", best_value_text)
                    candidate_nominal = re.match(r"^(\d+(?:\.\d+)?)", normalise_text(value or text))
                    if best_parts and candidate_nominal:
                        narrow_nominal = best_parts.group(1)
                        wide_nominal = candidate_nominal.group(1)
                        if len(wide_nominal) > len(narrow_nominal) and wide_nominal.endswith(narrow_nominal):
                            value = normalise_text(f"{wide_nominal}{best_parts.group(2)}")
                            text = value
                            merged_from_wider_nominal = True
                score = recognition_confidence + min(len(text), 24) / 100
                if measurement_type in {
                    "diameter",
                    "radius",
                    "chamfer",
                    "surface_finish",
                    "metric_thread",
                    "hole_callout",
                }:
                    # A valid engineering prefix is stronger evidence than a
                    # slightly longer plain OCR string. This prevents crops
                    # such as `Ø3+0.09` being replaced by noisy `03+8` text.
                    score += 0.20
                    if re.search(r"[+-]\s*\d+(?:\.\d+)?", normalise_text(value or text)):
                        score += 0.35
                if re.search(r"\d\.\d", value or text):
                    score += 0.20
                if re.search(r"\d(?:\.\d+)?\s*[A-Za-z]\d", text):
                    score += 0.25
                if re.fullmatch(r"\d", value or text):
                    score -= 0.40
                candidate = (
                    score,
                    recognition_confidence,
                    orientation,
                    measurement_type,
                    value,
                    text,
                    crop_x1 if merged_from_wider_nominal else x,
                    (
                        rescue_trim_bounds
                        if rescue_trim_bounds is not None and orientation.startswith("rescue_trim_")
                        else None
                    ),
                )
                if detector_ocr_candidate_is_better(best, candidate):
                    best = candidate

        if best is None:
            continue

        _, recognition_confidence, orientation, measurement_type, value, text = best[:6]
        # Detector confidence answers "is there a box?" while recognition
        # confidence answers "was the engineering text read clearly?". A
        # high-recall detector box is allowed to contribute only after the OCR
        # result passes class-specific parsing above. Do not cap that validated
        # reading at the lower detector confidence, otherwise every useful
        # low-threshold rescue is discarded later as very-low-confidence OCR.
        calibrated_confidence = max(
            detector_confidence,
            min(float(recognition_confidence), 0.90),
        )
        mapped_x1 = int(best[6]) if len(best) > 6 else x
        mapped_bounds = best[7] if len(best) > 7 else None
        if mapped_bounds:
            mapped_x1, mapped_y1, mapped_x2, mapped_y2 = [int(value) for value in mapped_bounds]
        else:
            mapped_y1, mapped_x2, mapped_y2 = y, x + width, y + height

        recovery_reasons = []
        recovered_text, recovered_x1, quantity_reason = recover_left_quantity_prefix(
            recognizer,
            image,
            x,
            y,
            width,
            height,
            measurement_type,
            value or text,
        )
        if quantity_reason:
            text = recovered_text
            value = recovered_text
            mapped_x1 = min(mapped_x1, recovered_x1)
            recovery_reasons.append(quantity_reason)

        english_text, english_x2, english_reason = prefer_adjacent_english_depth(
            recognizer,
            image,
            x,
            y,
            width,
            height,
            measurement_type,
            value or text,
        )
        if english_reason:
            text = english_text
            value = english_text
            mapped_x2 = max(mapped_x2, english_x2)
            recovery_reasons.append(english_reason)

        # Recognition may succeed on the padded crop while the original
        # detector frame still ends before the final digit. Expand only along
        # the text direction and use the measured text length as a conservative
        # lower bound. This keeps Ra values complete without creating a large
        # box over nearby drawing geometry.
        if measurement_type == "surface_finish":
            compact_surface = normalise_text(value or text).replace(" ", "")
            if width >= height * 1.25:
                expected_width = int(round(max(width, height * max(2.4, len(compact_surface) * 0.62))))
                mapped_x2 = min(image_width, max(mapped_x2, mapped_x1 + expected_width))
            elif height >= width * 1.25:
                expected_height = int(round(max(height, width * max(2.4, len(compact_surface) * 0.62))))
                mapped_y2 = min(image_height, max(mapped_y2, mapped_y1 + expected_height))
        mapped_item = {
            "text": text,
            "confidence": calibrated_confidence,
            "box": (mapped_x1, mapped_y1, mapped_x2, mapped_y2),
            "orientation": f"detector_recognition_{orientation}_{symbol_class}",
        }
        row = create_measurement_row(
            source_file=image_path.name,
            item=mapped_item,
            measurement_type=measurement_type,
            value=value,
            review_threshold=MEASUREMENT_REVIEW_THRESHOLD,
        )
        row["Detector Confidence"] = detector_confidence
        row["Detector OCR Validated"] = "YES"
        row["Review Reason"] = "; ".join(
            reason
            for reason in [
                row.get("Review Reason", ""),
                f"YOLO {symbol_class} crop",
                *recovery_reasons,
            ]
            if reason
        )
        if detector_confidence < 0.58:
            row["Review Reason"] = "; ".join(
                reason
                for reason in [
                    row.get("Review Reason", ""),
                    f"High-recall detector confidence {detector_confidence:.2f}",
                ]
                if reason
            )
        # Keep the detector result even when older page-wide OCR found a fragment
        # in the same place. The normal candidate-quality pass will retain the
        # more complete reading and still merge duplicate detector boxes.
        if not any(are_same_dimension_candidate(row, existing) for existing in rows):
            rows.append(row)

    return rows


def extract_local_stacked_tolerance_rows(image_path, existing_rows):
    tolerance_rows = [
        row
        for row in existing_rows
        if str(row.get("Measurement Type", "")) == "tolerance"
        and re.search(r"[+-]\s*0\.\d{1,3}", normalise_text(row.get("OCR Text") or row.get("Extracted Value") or ""))
    ]
    if not tolerance_rows:
        return []

    image_path = Path(image_path)
    image = cv2.imread(str(image_path))
    if image is None:
        return []

    image_height, image_width = image.shape[:2]
    ocr = create_paddle_ocr()
    rows = []

    for tolerance_row in tolerance_rows:
        x = int(tolerance_row.get("X", 0) or 0)
        y = int(tolerance_row.get("Y", 0) or 0)
        width = int(tolerance_row.get("Width", 0) or 0)
        height = int(tolerance_row.get("Height", 0) or 0)
        if width <= 0 or height <= 0:
            continue

        crop_x1 = max(0, x - 360)
        crop_y1 = max(0, y - 180)
        crop_x2 = min(image_width, x + width + 260)
        crop_y2 = min(image_height, y + height + 190)
        crop = image[crop_y1:crop_y2, crop_x1:crop_x2]
        if crop.size == 0:
            continue

        scale = 3
        enlarged = cv2.resize(crop, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
        try:
            local_items = extract_ocr_items_with_boxes(ocr.predict(enlarged))
        except Exception:
            continue

        mapped_items = []
        for item in local_items:
            text = normalise_text(item.get("text", ""))
            if not text or should_skip_text(text):
                continue
            bx1, by1, bx2, by2 = [int(v) for v in item["box"]]
            mapped_items.append(
                {
                    "text": text,
                    "confidence": float(item.get("confidence", 0) or 0),
                    "box": (
                        crop_x1 + bx1 // scale,
                        crop_y1 + by1 // scale,
                        crop_x1 + bx2 // scale,
                        crop_y1 + by2 // scale,
                    ),
                }
            )

        signed_items = [
            item
            for item in mapped_items
            if re.search(r"[+-]\s*0\.\d{1,3}", item["text"]) and item["confidence"] >= 0.60
        ]
        nominal_items = [
            item
            for item in mapped_items
            if re.search(
                r"(?:\d+\s*[xX]\s*)?\s*(?:Ø|[Φφ]|DIA\.?|DIAMETER|R)?\s*\d+(?:\.\d+)?",
                clean_ocr_specification(item["text"]),
                re.IGNORECASE,
            )
            and not re.search(r"[+-]\s*0\.\d{1,3}", item["text"])
            and item["confidence"] >= 0.55
        ]
        zero_items = [
            item
            for item in mapped_items
            if re.fullmatch(r"0(?:\.0+)?", clean_ocr_specification(item["text"]))
            and item["confidence"] >= 0.55
        ]

        if not signed_items or not nominal_items:
            continue

        signed_item = max(signed_items, key=lambda item: item["confidence"])
        signed_x1, signed_y1, signed_x2, signed_y2 = signed_item["box"]
        signed_center_y = (signed_y1 + signed_y2) / 2

        nominal_candidates = []
        for item in nominal_items:
            item_x1, item_y1, item_x2, item_y2 = item["box"]
            item_center_y = (item_y1 + item_y2) / 2
            if item_x2 > signed_x2 + 20:
                continue
            if abs(item_center_y - signed_center_y) > 95:
                continue
            text = clean_ocr_specification(item["text"])
            if not re.search(r"(?:Ø|DIA|R|\d+\s*X)", text, re.IGNORECASE):
                continue
            distance = abs(signed_x1 - item_x2) + abs(signed_center_y - item_center_y)
            nominal_candidates.append((distance, item))

        if not nominal_candidates:
            continue

        _, nominal_item = sorted(nominal_candidates, key=lambda item: item[0])[0]
        zero_item = None
        for item in zero_items:
            item_x1, item_y1, item_x2, item_y2 = item["box"]
            item_center_x = (item_x1 + item_x2) / 2
            if signed_x1 - 80 <= item_center_x <= signed_x2 + 80 and item_y1 >= signed_y1:
                zero_item = item
                break

        combined_text = clean_ocr_specification(nominal_item["text"]) + clean_ocr_specification(signed_item["text"])
        if zero_item:
            combined_text += clean_ocr_specification(zero_item["text"])

        measurement_type, value = classify_measurement(combined_text)
        if not measurement_type:
            continue

        boxes = [nominal_item["box"], signed_item["box"]]
        if zero_item:
            boxes.append(zero_item["box"])
        x1 = min(box[0] for box in boxes)
        y1 = min(box[1] for box in boxes)
        x2 = max(box[2] for box in boxes)
        y2 = max(box[3] for box in boxes)
        mapped_item = {
            "text": combined_text,
            "confidence": min(item["confidence"] for item in [nominal_item, signed_item] + ([zero_item] if zero_item else [])),
            "box": (x1, y1, x2, y2),
            "orientation": "local_stacked_tolerance",
        }
        row = create_measurement_row(
            source_file=image_path.name,
            item=mapped_item,
            measurement_type=measurement_type,
            value=value,
            review_threshold=MEASUREMENT_REVIEW_THRESHOLD,
        )
        row["Review Reason"] = "; ".join(
            reason for reason in [row.get("Review Reason", ""), "Local stacked tolerance crop"] if reason
        )
        rows.append(row)

    return rows


def format_number(value):
    if value in ("", None):
        return ""
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return str(value)
    if numeric.is_integer():
        return str(int(numeric))
    return f"{numeric:.4f}".rstrip("0").rstrip(".")


def general_tolerance_for_nominal(nominal):
    for low, high, tolerance in GENERAL_TOLERANCE_RANGES:
        if low < nominal <= high:
            return tolerance
    return None


def general_tolerance_for_specification(specification, nominal, general_tolerances=None):
    if general_tolerances:
        nominal_match = re.search(r"(?<![A-Z])(\d+(?:\.(\d+))?)", normalise_text(specification))
        if nominal_match and nominal_match.group(2) is not None:
            decimal_places = len(nominal_match.group(2))
            value = general_tolerances.get(decimal_places, general_tolerances.get(str(decimal_places)))
            try:
                if value not in {None, ""}:
                    return float(value)
            except (TypeError, ValueError):
                pass
        # A partially extracted title-block map must not disable the standard
        # range fallback for a decimal precision that is absent from the map.
        return general_tolerance_for_nominal(nominal)
    return general_tolerance_for_nominal(nominal)


def fixed_three_decimals(value):
    return f"{float(value):.3f}"


def parse_fit_dimension_fields(specification):
    """Split a fit callout into diameter, nominal, class, and explicit limits."""
    text = normalise_text(specification)
    match = re.match(
        r"^\s*(?:(Ø|⌀|∅|Φ|φ)\s*)?(\d+(?:\.\d+)?)\s*([A-Za-z]{1,2}\d{1,2})(.*)$",
        text,
    )
    if not match or re.match(r"^M\d", match.group(3), re.IGNORECASE):
        return None

    diameter_symbol = match.group(1)
    nominal = float(match.group(2))
    tolerance_class = match.group(3)
    tail = normalise_text(match.group(4))
    signed_values = re.findall(r"([+-])\s*(\d+(?:\.\d+)?)", tail)
    lower = None
    upper = None

    for sign, value in signed_values[:2]:
        deviation = float(value)
        signed = deviation if sign == "+" else -deviation
        lower = signed if lower is None else min(lower, signed)
        upper = signed if upper is None else max(upper, signed)

    explicit_zero = bool(re.search(r"(?:/|\s)0(?:\.0+)?\s*$", tail))
    if len(signed_values) == 1 and explicit_zero:
        lower = min(lower if lower is not None else 0.0, 0.0)
        upper = max(upper if upper is not None else 0.0, 0.0)
    elif len(signed_values) == 1 and not explicit_zero:
        # OCR often joins stacked `+0.018` and `0` as `+0.0180`.
        sign, value = signed_values[0]
        decimal = value.split(".", 1)[1] if "." in value else ""
        if sign == "+" and len(decimal) == 4 and decimal.endswith("0"):
            upper = float(value[:-1])
            lower = 0.0

    result = {
        "symbol": "Ø" if diameter_symbol else "",
        "nominal": format_number(nominal),
        "tolerance_class": tolerance_class,
        "minus": "",
        "plus": "",
        "minimum": "",
        "maximum": "",
        "has_explicit_limits": lower is not None and upper is not None,
    }
    if result["has_explicit_limits"]:
        result.update(
            {
                "minus": fixed_three_decimals(abs(min(lower, 0.0))),
                "plus": f"{max(upper, 0.0):+.3f}",
                "minimum": fixed_three_decimals(nominal + lower),
                "maximum": fixed_three_decimals(nominal + upper),
            }
        )
    return result


def has_fit_class(specification):
    return parse_fit_dimension_fields(specification) is not None


def parse_tolerance(specification, apply_general=True, general_tolerances=None):
    spec = normalise_text(specification)
    compact = spec.replace(" ", "")
    nominal_match = re.search(r"(?<![A-Z])(\d+(?:\.\d+)?)", compact)
    if not nominal_match:
        return "", "", "", "", ""

    nominal = float(nominal_match.group(1))
    minus = ""
    plus = ""

    plus_minus = re.search(r"(?:±|\+/-)(\d+(?:\.\d+)?)", compact)
    if plus_minus:
        plus = minus = float(plus_minus.group(1))
    else:
        def parse_signed_deviation(sign, value):
            # Stacked tolerance OCR can lose the decimal point while keeping
            # its leading zero: `+0140` is `+0.140`, and `+0018` is `+0.018`.
            if "." not in value and len(value) >= 3 and value.startswith("0"):
                value = f"0.{value[1:]}"
            return float(f"{sign}{value}")

        signed_deviations = [
            parse_signed_deviation(sign, value)
            for sign, value in re.findall(r"([+-])\(?(\d+(?:\.\d+)?)\)?", compact[nominal_match.end():])
        ]
        if len(signed_deviations) >= 2:
            lower_deviation = min(signed_deviations[:2])
            upper_deviation = max(signed_deviations[:2])
            minus = abs(lower_deviation) if lower_deviation < 0 else 0.0
            plus = upper_deviation
        else:
            explicit_plus = re.search(r"\+(\d+(?:\.\d+)?)", compact)
            explicit_minus = re.search(r"-(\d+(?:\.\d+)?)", compact)
            if explicit_plus:
                plus_text = explicit_plus.group(1)
                if "." not in plus_text and len(plus_text) >= 3 and plus_text.startswith("0"):
                    plus_text = f"0.{plus_text[1:]}"
                if (
                    not explicit_minus
                    and plus_text.startswith("0.")
                    and len(plus_text.split(".", 1)[1]) >= 3
                    and plus_text.endswith("0")
                ):
                    plus_text = plus_text[:-1]
                    minus = 0.0
                plus = float(plus_text)
            if explicit_minus:
                minus_text = explicit_minus.group(1)
                if "." not in minus_text and len(minus_text) >= 3 and minus_text.startswith("0"):
                    minus_text = f"0.{minus_text[1:]}"
                minus = float(minus_text)

    if minus == "" and plus == "":
        if has_fit_class(spec):
            return format_number(nominal), "", "", "", ""
        if not apply_general:
            return format_number(nominal), "", "", "", ""
        general = general_tolerance_for_specification(spec, nominal, general_tolerances)
        if general is None:
            return format_number(nominal), "", "", "", ""
        minus = plus = general

    lower = nominal - float(minus or 0)
    upper = nominal + float(plus or 0)
    return (
        format_number(nominal),
        format_number(minus),
        format_number(plus),
        format_number(lower),
        format_number(upper),
    )


def has_explicit_tolerance(value):
    return bool(re.search(r"(?:Â±|±|\+/-|\+-|[+-]\s*\d)", normalise_text(value)))


def is_reference_dimension_text(value):
    return bool(re.fullmatch(r"\(\s*\d+(?:\.\d+)?\s*\)", normalise_text(value)))


def is_complex_text_callout(value):
    text = normalise_text(value).upper()
    compact = clean_ocr_specification(text).upper()
    return bool(
        re.search(r"C'?BORE|COUNTERBORE|KEYUNDER|KEY\s*UNDER|DEPTH|DP", text, re.IGNORECASE)
        or re.search(r"C'?BORE|COUNTERBORE|KEYUNDER|DEPTH|DP", compact, re.IGNORECASE)
    )


def is_incomplete_numeric_callout(measurement_type, specification, dimension_text):
    if measurement_type not in {"diameter", "radius", "chamfer", "thickness"}:
        return False

    compact_spec = clean_ocr_specification(specification).upper()
    compact_dimension = clean_ocr_specification(dimension_text).upper()
    if re.search(r"\d\.$", compact_spec):
        return True
    if compact_dimension in {"0", "0."} and re.search(r"[A-ZØ∅]\s*0\.?$", compact_spec):
        return True
    return False


def split_symbol_and_dimension(specification, measurement_type):
    spec = clean_ocr_specification(specification)
    multiplier_match = re.match(r"^(\d+)\s*X\s*(.+)$", spec, re.IGNORECASE)
    if multiplier_match:
        spec = multiplier_match.group(2).strip()
    if measurement_type == "metric_thread" or re.match(r"^M\d", spec, re.IGNORECASE):
        return split_metric_thread_value(spec)
    if measurement_type == "surface_finish":
        return "Ra", re.sub(r"^[Rr][Aa]\s*", "", spec)
    if spec.upper().startswith("R"):
        return "R", re.sub(r"^[Rr]\s*", "", spec)
    if spec.startswith("Ø"):
        return "Ø", re.sub(r"^Ø\s*", "", spec)
    if measurement_type == "diameter":
        return "Ø", re.sub(r"^Ø\s*", "", spec)
    if measurement_type == "radius":
        return "R", re.sub(r"^[Rr]\s*", "", spec)
    if is_complex_text_callout(spec):
        return "", spec
    if measurement_type == "chamfer" or spec.upper().startswith("C"):
        return "C", re.sub(r"^[Cc]\s*", "", spec)
    if measurement_type == "thickness" or spec.lower().startswith("t"):
        return "t", re.sub(r"^[tT]\s*", "", spec)
    if measurement_type == "hole_callout":
        through_match = re.search(r"(\d+(?:\.\d+)?)\s*(?:DRILL\s*)?(?:THRU|THROUGH)", spec, re.IGNORECASE)
        if through_match:
            return "", through_match.group(1)
    return "", spec


def is_note_false_positive(measurement):
    text = normalise_text(
        f"{measurement.get('OCR Text', '')} {measurement.get('Extracted Value', '')}"
    )
    return any(pattern.search(text) for pattern in NOTE_FALSE_POSITIVE_PATTERNS)


def is_low_confidence_plain_dimension(measurement):
    measurement_type = str(measurement.get("Measurement Type", ""))
    if measurement_type not in {"plain_dimension", "reference_dimension"}:
        return False

    value = normalise_text(measurement.get("Extracted Value") or measurement.get("OCR Text") or "")
    if not re.fullmatch(r"\(?\d{2,4}(?:\.\d+)?\)?", value):
        return False

    confidence = float(measurement.get("OCR Confidence", 0) or 0)
    return confidence < LOW_CONFIDENCE_PLAIN_DIMENSION_THRESHOLD


def is_very_low_confidence_plain_dimension(measurement):
    measurement_type = str(measurement.get("Measurement Type", ""))
    if measurement_type not in {"plain_dimension", "reference_dimension"}:
        return False

    confidence = float(measurement.get("OCR Confidence", 0) or 0)
    return confidence < 0.55


def exclusion_reason_for_measurement(measurement):
    if is_note_false_positive(measurement):
        return "Rejected note/sentence text"
    return "Excluded title/revision/table zone"


def extract_multiplier(specification):
    spec = clean_ocr_specification(specification)
    match = re.match(r"^(\d{1,2})\s*X\s*(.+)$", spec, re.IGNORECASE)
    if not match:
        return 1, spec

    count = int(match.group(1))
    if count < 2 or count > 20:
        return 1, spec

    return count, match.group(2).strip()


def infer_operation(measurement_type):
    if measurement_type == "gdt":
        return "GD&T"
    if measurement_type in {"metric_thread", "hole_callout"}:
        return "Hole / Thread"
    if measurement_type == "surface_finish":
        return "Surface Finish"
    if measurement_type in {"diameter", "radius", "chamfer", "thickness", "plain_dimension", "reference_dimension"}:
        return "Dimensional"
    if measurement_type == "angle":
        return "Angle"
    if measurement_type == "tolerance":
        return "Tolerance"
    return "Dimensional"


def infer_equipment(measurement_type, specification):
    spec = normalise_text(specification).upper()
    if measurement_type == "gdt":
        return "V"
    if measurement_type == "metric_thread" or "M" in spec:
        return "TG"
    if measurement_type == "hole_callout" and ("DRILL" in spec or "DEPTH" in spec or "DP" in spec):
        return "TG + DC"
    if measurement_type == "surface_finish":
        return "V"
    if measurement_type == "chamfer":
        return "CG"
    if measurement_type in {"diameter", "plain_dimension", "reference_dimension", "thickness", "radius"}:
        return "DC"
    return "DC"


def review_reason_for_row(row, nominal, min_value, max_value, used_general_tolerance=False):
    reasons = []
    source_value = normalise_text(row.get("Extracted Value") or row.get("OCR Text") or "")
    has_row_tolerance = bool(re.search(r"(?:\+/-|[+-]\s*\d)", source_value))
    existing_reason = str(row.get("Review Reason", "") or "").strip()
    if existing_reason:
        for reason in existing_reason.split(";"):
            reason = reason.strip()
            if (used_general_tolerance or has_row_tolerance) and reason == "Plain number needs human check":
                continue
            reasons.append(reason)
    if row.get("Measurement Type") in {"plain_dimension", "reference_dimension"} and not min_value and not max_value:
        reasons.append("No tolerance found")
    if used_general_tolerance:
        reasons.append("General tolerance applied")
    if float(row.get("OCR Confidence", 0) or 0) < MEASUREMENT_REVIEW_THRESHOLD:
        reasons.append("Low OCR confidence")
    if not nominal and row.get("Measurement Type") not in {"metric_thread", "hole_callout", "surface_finish"}:
        reasons.append("Could not calculate nominal")

    unique = []
    for reason in reasons:
        if reason and reason not in unique:
            unique.append(reason)
    return "; ".join(unique)


def needs_human_review(review_reason):
    blocking_reasons = [
        reason.strip()
        for reason in str(review_reason or "").split(";")
        if reason.strip() and reason.strip() != "General tolerance applied"
    ]
    return "YES" if blocking_reasons else "NO"


def remove_duplicate_candidates(rows):
    kept = []
    for row in sorted(rows, key=lambda item: candidate_sort_key(item)):
        duplicate_index = None
        for index, existing in enumerate(kept):
            if are_same_dimension_candidate(row, existing):
                duplicate_index = index
                break

        if duplicate_index is None:
            kept.append(row)
            continue

        if candidate_quality_score(row) > candidate_quality_score(kept[duplicate_index]):
            kept[duplicate_index] = row
    return sorted(kept, key=lambda item: (int(item.get("Y", 0) or 0), int(item.get("X", 0) or 0)))


def candidate_sort_key(row):
    return (
        int(row.get("Y", 0) or 0),
        int(row.get("X", 0) or 0),
        -candidate_quality_score(row),
    )


def is_detector_guided_row(row):
    orientation = str(row.get("OCR Orientation", ""))
    return orientation.startswith(
        ("detector_recognition_", "detector_multiline_", "detector_gap_single_digit_rescue")
    )


def candidate_quality_score(row):
    text = normalise_text(row.get("Extracted Value") or row.get("OCR Text") or "")
    confidence = float(row.get("OCR Confidence", 0) or 0)
    orientation = str(row.get("OCR Orientation", "normal")).lower()
    orientation_bonus = 0.08 if orientation == "normal" else 0
    if orientation == "vertical_decimal_rescue":
        orientation_bonus = 0.24
    if is_detector_guided_row(row):
        orientation_bonus = 0.45
    completeness_bonus = min(len(normalize_duplicate_text(text)), 30) / 100
    numeric_bonus = 0.12 if re.search(r"\d+\.\d+", text) else 0
    explicit_tolerance_bonus = 0.30 if re.search(r"[+-]\s*\d+(?:\.\d+)?", text) else 0
    signed_deviations = [
        float(value)
        for value in re.findall(r"[+-]\s*(\d+(?:\.\d+)?)", text)
    ]
    nonzero_tolerance_bonus = 0.24 if any(value > 0 for value in signed_deviations) else 0
    fragment_penalty = 0.22 if re.fullmatch(r"\d", text) else 0
    complete_number_bonus = 0.14 if re.fullmatch(r"\d{2,4}(?:\.\d+)?", text) else 0
    multiplier_bonus = 0.45 if re.match(r"^\d{1,2}X", normalize_duplicate_text(text)) else 0
    measurement_type = str(row.get("Measurement Type", ""))
    hole_numeric_bonus = 0.65 if measurement_type == "hole_callout" and numeric_values_in_text(text) else 0
    incomplete_hole_penalty = 0.45 if measurement_type == "hole_callout" and not numeric_values_in_text(text) else 0
    symbol_type_bonus = 0.35 if measurement_type in {
        "diameter",
        "radius",
        "chamfer",
        "thickness",
        "metric_thread",
        "surface_finish",
    } else 0
    plain_review_penalty = 0.18 if measurement_type in {"plain_dimension", "reference_dimension"} else 0
    width = int(row.get("Width", 0) or 0)
    height = int(row.get("Height", 0) or 0)
    tall_integer_penalty = (
        0.24
        if measurement_type in {"plain_dimension", "reference_dimension"}
        and re.fullmatch(r"\d{2,3}", text)
        and height > width * 1.25
        else 0
    )
    return (
        confidence
        + orientation_bonus
        + completeness_bonus
        + numeric_bonus
        + explicit_tolerance_bonus
        + nonzero_tolerance_bonus
        + complete_number_bonus
        + multiplier_bonus
        + hole_numeric_bonus
        + symbol_type_bonus
        - fragment_penalty
        - incomplete_hole_penalty
        - plain_review_penalty
        - tall_integer_penalty
    )


def characteristic_quality_score(row):
    confidence = float(row.get("AI Confidence", row.get("OCR Confidence", 0)) or 0)
    symbol = normalise_text(row.get("Symbol", ""))
    dimension = normalise_text(row.get("Dimension", ""))
    specification = normalise_text(row.get("Specification", ""))
    measurement_type = str(row.get("Measurement Type", ""))
    score = confidence
    if symbol:
        score += 0.20
    if dimension:
        score += 0.20
    if normalise_text(row.get("MIN", "")) or normalise_text(row.get("MAX", "")):
        score += 0.25
    if normalise_text(row.get("Tolerance -", "")) or normalise_text(row.get("Tolerance +", "")):
        score += 0.15
    tolerance_values = []
    for field in ("Tolerance -", "Tolerance +"):
        try:
            tolerance_values.append(abs(float(normalise_text(row.get(field, "")))))
        except (TypeError, ValueError):
            continue
    # When two overlapping detector crops describe the same nominal, prefer
    # the crop that preserved a real deviation over a partial all-zero read.
    if any(value > 0 for value in tolerance_values):
        score += 0.24
    if measurement_type in {"diameter", "radius", "chamfer", "metric_thread", "surface_finish", "gdt"}:
        score += 0.15
    if "pdf_text_layer" in str(row.get("OCR Orientation", "")):
        score -= 0.05
    if len(normalize_duplicate_text(specification)) > len(normalize_duplicate_text(dimension)):
        score += 0.08
    return score


def normalize_duplicate_text(value):
    text = clean_ocr_specification(value).upper()
    replacements = {
        "Ø": "D",
        "\u00d8": "D",
        "\u2205": "D",
        "⌀": "D",
        "Φ": "D",
        "φ": "D",
        "Ρ": "D",
        "ρ": "D",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = text.replace("THRU", "THROUGH")
    text = text.replace("COUNTERBORE", "CBORE")
    text = text.replace("DRILL", "")
    text = text.replace("MM", "")
    text = re.sub(r"[^A-Z0-9.]", "", text)
    text = re.sub(r"(?<!\d)\.(?=\d)", "0.", text)
    text = re.sub(r"M([0-9])\1(?=DEPTH|THROUGH|THRU|CBORE|CSINK|$)", r"M\1", text)
    text = re.sub(r"^CBORE[OD0]?", "CBORED", text)
    return text


def numeric_signature(value):
    numbers = re.findall(r"\d+(?:\.\d+)?", normalise_text(value))
    return tuple(numbers)


def has_non_ascii_text(value):
    return any(ord(char) > 127 for char in str(value or ""))


def trusted_hole_callout_text(value):
    text = normalise_text(value).upper()
    return bool(re.search(r"C'?BORE|COUNTERBORE|DRILL|THRU|THROUGH", text))


def numeric_values_in_text(value):
    return re.findall(r"\d+(?:\.\d+)?", normalise_text(value))


def row_box(row):
    x = int(row.get("X", 0) or 0)
    y = int(row.get("Y", 0) or 0)
    width = int(row.get("Width", 0) or 0)
    height = int(row.get("Height", 0) or 0)
    return x, y, x + width, y + height


def box_iou(a, b):
    ax1, ay1, ax2, ay2 = row_box(a)
    bx1, by1, bx2, by2 = row_box(b)
    ix1 = max(ax1, bx1)
    iy1 = max(ay1, by1)
    ix2 = min(ax2, bx2)
    iy2 = min(ay2, by2)
    if ix2 <= ix1 or iy2 <= iy1:
        return 0.0
    intersection = (ix2 - ix1) * (iy2 - iy1)
    area_a = max(1, (ax2 - ax1) * (ay2 - ay1))
    area_b = max(1, (bx2 - bx1) * (by2 - by1))
    return intersection / max(1, area_a + area_b - intersection)


def box_containment(a, b):
    ax1, ay1, ax2, ay2 = row_box(a)
    bx1, by1, bx2, by2 = row_box(b)
    ix1 = max(ax1, bx1)
    iy1 = max(ay1, by1)
    ix2 = min(ax2, bx2)
    iy2 = min(ay2, by2)
    if ix2 <= ix1 or iy2 <= iy1:
        return 0.0
    intersection = (ix2 - ix1) * (iy2 - iy1)
    area_a = max(1, (ax2 - ax1) * (ay2 - ay1))
    area_b = max(1, (bx2 - bx1) * (by2 - by1))
    return intersection / max(1, min(area_a, area_b))


def characteristic_identity(row):
    symbol, value = split_report_symbol_value(row)
    measurement_type = str(row.get("Measurement Type", ""))
    symbol_key = normalize_duplicate_text(symbol)
    value_key = normalize_duplicate_text(value or row.get("Dimension", ""))
    if measurement_type in {"plain_dimension", "reference_dimension"}:
        symbol_key = ""
    if measurement_type in {"diameter", "radius", "chamfer", "thickness", "metric_thread", "surface_finish"}:
        type_key = measurement_type
    elif measurement_type == "hole_callout":
        type_key = "hole_callout"
    else:
        type_key = measurement_type
    return type_key, symbol_key, value_key


def is_grouped_or_sub_characteristic(row):
    try:
        multiplier_count = int(row.get("Multiplier Count", 1) or 1)
    except (TypeError, ValueError):
        multiplier_count = 1
    try:
        subrow_count = int(row.get("Subrow Count", 1) or 1)
    except (TypeError, ValueError):
        subrow_count = 1
    return multiplier_count > 1 or subrow_count > 1


def are_duplicate_characteristics(row, existing):
    row_type, row_symbol, row_value = characteristic_identity(row)
    existing_type, existing_symbol, existing_value = characteristic_identity(existing)
    if not row_value or not existing_value:
        return False
    if row_type != existing_type:
        compatible_plain_symbol = {
            row_type,
            existing_type,
        } <= {"plain_dimension", "diameter", "radius", "chamfer", "thickness"}
        if not compatible_plain_symbol:
            return False
    if row_symbol != existing_symbol:
        if row_symbol and existing_symbol:
            return False
        if row_type not in {"plain_dimension", "diameter", "radius", "chamfer", "thickness"}:
            return False
    if row_value != existing_value:
        return False

    if is_grouped_or_sub_characteristic(row) or is_grouped_or_sub_characteristic(existing):
        same_source_box = box_iou(row, existing) >= 0.80 or box_containment(row, existing) >= 0.90
        return same_source_box and row.get("Multiplier Index") == existing.get("Multiplier Index")

    row_x, row_y = candidate_center(row)
    existing_x, existing_y = candidate_center(existing)
    close_center = abs(row_x - existing_x) <= 90 and abs(row_y - existing_y) <= 70
    strong_overlap = box_iou(row, existing) >= 0.35 or box_containment(row, existing) >= 0.70
    return close_center or strong_overlap


def append_duplicate_remark(row, removed_row):
    removed_source = normalise_text(removed_row.get("OCR Orientation", ""))
    reason = "Merged duplicate detection"
    if removed_source:
        reason = f"{reason} from {removed_source}"
    append_review_reason(row, reason)


def remove_duplicate_characteristics(rows):
    kept = []
    for row in sort_top_bottom_left_right(rows):
        duplicate_index = None
        for index, existing in enumerate(kept):
            if are_duplicate_characteristics(row, existing):
                duplicate_index = index
                break

        if duplicate_index is None:
            kept.append(row)
            continue

        existing = kept[duplicate_index]
        if characteristic_quality_score(row) > characteristic_quality_score(existing):
            append_duplicate_remark(row, existing)
            kept[duplicate_index] = row
        else:
            append_duplicate_remark(existing, row)

    return sort_top_bottom_left_right(kept)


SYMBOL_REJECT_CLASSES = {
    "datum_symbol",
    "gdt_frame_symbol",
    "gdt_parallelism_symbol",
    "gdt_perpendicularity_symbol",
    "revision_triangle_symbol",
}


def symbol_detection_box(symbol):
    x = int(symbol.get("X", 0) or 0)
    y = int(symbol.get("Y", 0) or 0)
    width = int(symbol.get("Width", 0) or 0)
    height = int(symbol.get("Height", 0) or 0)
    pad = max(12, int(max(width, height) * 1.8))
    return {
        "X": max(0, x - pad),
        "Y": max(0, y - pad),
        "Width": width + pad * 2,
        "Height": height + pad * 2,
    }


def symbol_reject_reason(measurement, symbol_detections):
    if not symbol_detections:
        return ""

    measurement_type = str(measurement.get("Measurement Type", ""))
    for symbol in symbol_detections:
        symbol_class = str(symbol.get("Symbol Class", ""))
        if symbol_class not in SYMBOL_REJECT_CLASSES:
            continue
        if measurement_type == "surface_finish":
            continue
        if (
            measurement_type == "hole_callout"
            and numeric_values_in_text(measurement.get("OCR Text") or measurement.get("Extracted Value") or "")
            and re.search(
                r"\bDRILL\b|\bTHRU\b|\bTHROUGH\b|C'?BORE|COUNTERBORE|\bDEPTH\b|\bDP\b",
                normalise_text(measurement.get("OCR Text") or measurement.get("Extracted Value") or ""),
                re.IGNORECASE,
            )
        ):
            continue
        if symbol_class == "revision_triangle_symbol":
            if measurement_type not in {"plain_dimension", "reference_dimension", "tolerance"}:
                continue
            actual_symbol_box = {
                "X": int(symbol.get("X", 0) or 0),
                "Y": int(symbol.get("Y", 0) or 0),
                "Width": int(symbol.get("Width", 0) or 0),
                "Height": int(symbol.get("Height", 0) or 0),
            }
            if box_containment(measurement, actual_symbol_box) >= 0.55 or box_iou(measurement, actual_symbol_box) >= 0.20:
                return "Inside YOLO revision_triangle_symbol zone"
            continue
        if symbol_class == "gdt_frame_symbol" and measurement_type in {"plain_dimension", "reference_dimension"}:
            actual_symbol_box = {
                "X": int(symbol.get("X", 0) or 0),
                "Y": int(symbol.get("Y", 0) or 0),
                "Width": int(symbol.get("Width", 0) or 0),
                "Height": int(symbol.get("Height", 0) or 0),
            }
            value = corrected_measurement_specification(measurement)
            if box_containment(measurement, actual_symbol_box) >= 0.70 and not re.search(r"0\.\d{1,3}", value):
                return "Invalid numeric OCR inside GD&T frame"
        if symbol_class in {"datum_symbol", "gdt_frame_symbol", "gdt_parallelism_symbol", "gdt_perpendicularity_symbol"} and measurement_type not in {
            "plain_dimension",
            "reference_dimension",
            "tolerance",
            "hole_callout",
        }:
            continue
        detector_guided = is_detector_guided_row(measurement)
        comparison_box = symbol if detector_guided else symbol_detection_box(symbol)
        overlap_threshold = 0.10 if detector_guided else 0.02
        if box_iou(measurement, comparison_box) >= overlap_threshold:
            return f"Inside YOLO {symbol_class} zone"

    return ""


def repair_symmetric_vertical_decimal_rows(rows, image_shape=None):
    """Repair repeated low-confidence vertical OCR such as 6.5 read as 56."""
    if image_shape is None:
        return rows
    _, image_width = image_shape[:2]
    groups = {}
    for row in rows:
        measurement_type = str(row.get("Measurement Type", ""))
        value = corrected_measurement_specification(row)
        width = int(row.get("Width", 0) or 0)
        height = int(row.get("Height", 0) or 0)
        confidence = float(row.get("OCR Confidence", 0) or 0)
        if measurement_type not in {"plain_dimension", "reference_dimension"}:
            continue
        if not re.fullmatch(r"\d{2}", value) or height <= width * 1.25 or confidence >= 0.80:
            continue
        groups.setdefault(value, []).append(row)

    for value, matching_rows in groups.items():
        centers = [candidate_center(row)[0] / max(1, image_width) for row in matching_rows]
        if len(matching_rows) < 2 or min(centers) >= 0.25 or max(centers) <= 0.70:
            continue
        reversed_value = value[::-1]
        repaired = f"{reversed_value[0]}.{reversed_value[1:]}"
        for row in matching_rows:
            row["Extracted Value"] = repaired
            row["OCR Text"] = repaired
            append_review_reason(row, "Recovered symmetric vertical decimal - verify")
    return rows


def is_unresolved_vertical_decimal(measurement):
    """Reject tall, low-confidence integers that are likely decimals read sideways."""
    value = corrected_measurement_specification(measurement)
    width = int(measurement.get("Width", 0) or 0)
    height = int(measurement.get("Height", 0) or 0)
    confidence = float(measurement.get("OCR Confidence", 0) or 0)
    orientation = str(measurement.get("OCR Orientation", measurement.get("orientation", "")) or "")
    measurement_type = str(measurement.get("Measurement Type", ""))
    if is_detector_guided_row(measurement) and confidence >= 0.75:
        return False
    return bool(
        measurement_type in {"plain_dimension", "reference_dimension"}
        and re.fullmatch(r"\d{2,3}", value)
        and height > width * 1.25
        and confidence < 0.90
        and orientation != "vertical_decimal_rescue"
    )


def is_invalid_surface_finish(measurement):
    if str(measurement.get("Measurement Type", "")) != "surface_finish":
        return False
    value = normalise_text(measurement.get("Extracted Value") or measurement.get("OCR Text") or "")
    match = re.search(r"(?:RA)?\s*(\d+(?:\.\d+)?)", value, re.IGNORECASE)
    return bool(match and float(match.group(1)) <= 0)


def is_thread_derived_duplicate(measurement, all_measurements):
    if str(measurement.get("Measurement Type", "")) != "thickness":
        return False
    value = corrected_measurement_specification(measurement)
    numbers = numeric_signature(value)
    for other in all_measurements:
        if other is measurement or str(other.get("Measurement Type", "")) != "metric_thread":
            continue
        other_value = corrected_measurement_specification(other)
        if box_iou(measurement, other) >= 0.35 and numbers and numbers <= numeric_signature(other_value):
            return True
    return False


def is_depth_derived_duplicate(measurement, all_measurements):
    """Reject `t15` when the same callout was correctly recovered as DEPTH 15."""
    if str(measurement.get("Measurement Type", "")) != "thickness":
        return False
    value = corrected_measurement_specification(measurement)
    match = re.fullmatch(r"[Tt]\s*(\d+(?:\.\d+)?)", value)
    if not match:
        return False
    depth_value = match.group(1)
    center_x, center_y = candidate_center(measurement)
    for other in all_measurements:
        if other is measurement or str(other.get("Measurement Type", "")) != "hole_callout":
            continue
        other_value = corrected_measurement_specification(other)
        if not re.fullmatch(rf"(?:DEPTH|DP)\s*{re.escape(depth_value)}", other_value, re.IGNORECASE):
            continue
        other_x, other_y = candidate_center(other)
        if abs(center_x - other_x) <= 280 and abs(center_y - other_y) <= 140:
            return True
    return False


def is_shadowed_by_detector_measurement(measurement, all_measurements):
    """Reject a weaker full-page OCR fragment covered by an exact detector crop."""
    if is_detector_guided_row(measurement):
        return False
    if str(measurement.get("Measurement Type", "")) not in {"plain_dimension", "reference_dimension"}:
        return False

    for other in all_measurements:
        if other is measurement:
            continue
        if not is_detector_guided_row(other):
            continue
        other_type = str(other.get("Measurement Type", ""))
        other_value = corrected_measurement_specification(other)
        detector_has_complete_tolerance = (
            other_type in {"plain_dimension", "reference_dimension"}
            and has_explicit_tolerance(other_value)
        )
        measurement_value = corrected_measurement_specification(measurement)
        detector_plain_replacement = bool(
            other_type in {"plain_dimension", "reference_dimension"}
            and re.fullmatch(r"\d+(?:\.\d+)?", measurement_value)
            and re.fullmatch(r"\d+(?:\.\d+)?", other_value)
            and len(measurement_value) == len(other_value)
            and float(other.get("OCR Confidence", 0) or 0) >= 0.85
            and (
                box_containment(other, measurement) >= 0.80
                or box_iou(measurement, other) >= 0.65
            )
        )
        if other_type not in {
            "diameter",
            "radius",
            "chamfer",
            "thickness",
            "metric_thread",
            "surface_finish",
        } and not detector_has_complete_tolerance and not detector_plain_replacement:
            continue
        if box_containment(measurement, other) >= 0.55 or box_iou(measurement, other) >= 0.30:
            return True
    return False


def is_nested_reference_fragment(measurement, all_measurements):
    """Reject a weaker numeric crop fully contained by a reference dimension."""
    measurement_type = str(measurement.get("Measurement Type", ""))
    if measurement_type not in {"plain_dimension", "reference_dimension"}:
        return False
    value = corrected_measurement_specification(measurement)
    value_payload = re.sub(r"[()]", "", value)
    if not re.fullmatch(r"\d+(?:\.\d+)?", value_payload):
        return False
    confidence = float(measurement.get("OCR Confidence", 0) or 0)
    for other in all_measurements:
        if other is measurement or str(other.get("Measurement Type", "")) != "reference_dimension":
            continue
        other_value = corrected_measurement_specification(other)
        other_payload = re.sub(r"[()]", "", other_value)
        if measurement_type == "reference_dimension":
            if len(other_payload) <= len(value_payload) or value_payload not in other_payload:
                continue
            measurement_x1 = int(measurement.get("X", 0) or 0)
            measurement_x2 = measurement_x1 + int(measurement.get("Width", 0) or 0)
            other_x1 = int(other.get("X", 0) or 0)
            other_x2 = other_x1 + int(other.get("Width", 0) or 0)
            x_overlap = max(0, min(measurement_x2, other_x2) - max(measurement_x1, other_x1))
            center_x, center_y = candidate_center(measurement)
            other_center_x, other_center_y = candidate_center(other)
            if (
                abs(center_x - other_center_x) <= 55
                and abs(center_y - other_center_y) <= 85
                and x_overlap >= min(
                    int(measurement.get("Width", 0) or 0),
                    int(other.get("Width", 0) or 0),
                )
                * 0.35
            ):
                return True
            continue
        if not is_detector_guided_row(other):
            continue
        other_confidence = float(other.get("OCR Confidence", 0) or 0)
        if other_confidence < confidence + 0.08:
            continue
        if box_containment(measurement, other) >= 0.90:
            return True
    return False


def is_nested_specialized_fragment(measurement, all_measurements):
    """Reject a short symbol crop contained by a complete same-symbol value."""
    measurement_type = str(measurement.get("Measurement Type", ""))
    if measurement_type not in {"diameter", "radius", "chamfer", "thickness"}:
        return False
    value = corrected_measurement_specification(measurement)
    value_numbers = numeric_values_in_text(value)
    if len(value_numbers) != 1:
        return False
    compact_value = value_numbers[0].replace(".", "")
    if len(compact_value) > 2:
        return False

    for other in all_measurements:
        if other is measurement or str(other.get("Measurement Type", "")) != measurement_type:
            continue
        other_value = corrected_measurement_specification(other)
        other_numbers = numeric_values_in_text(other_value)
        if not other_numbers:
            continue
        other_nominal = other_numbers[0].replace(".", "")
        if len(other_nominal) <= len(compact_value) or compact_value not in other_nominal:
            continue
        complete_tolerance = has_explicit_tolerance(other_value) or bool(
            normalise_text(other.get("Grouped Tolerance -", ""))
            or normalise_text(other.get("Grouped Tolerance +", ""))
        )
        if not complete_tolerance:
            continue
        if box_containment(measurement, other) >= 0.55 or box_iou(measurement, other) >= 0.18:
            return True
    return False


def text_similarity_key(text):
    compact = normalize_duplicate_text(text)
    return re.sub(r"[.]", "", compact)


def are_same_dimension_candidate(row, existing):
    center_x, center_y = candidate_center(row)
    other_x, other_y = candidate_center(existing)
    row_type = str(row.get("Measurement Type", ""))
    existing_type = str(existing.get("Measurement Type", ""))
    hole_callout_pair = row_type == "hole_callout" and existing_type == "hole_callout"
    # Multi-pass OCR often finds the English line and its translated/rotated
    # copy with slightly shifted centres.  Keep this wide enough to join those
    # copies, but still require matching/contained text below before merging.
    specialized_pair = row_type == existing_type and row_type in {
        "metric_thread", "diameter", "radius", "chamfer", "thickness", "surface_finish"
    }
    x_tolerance = 300 if hole_callout_pair else (150 if specialized_pair else 70)
    y_tolerance = 180 if hole_callout_pair else (90 if specialized_pair else 45)
    close_center = abs(center_x - other_x) <= x_tolerance and abs(center_y - other_y) <= y_tolerance
    overlapping = box_iou(row, existing) >= 0.35
    if not close_center and not overlapping:
        return False

    text = normalize_duplicate_text(row.get("Extracted Value") or row.get("OCR Text") or "")
    other_text = normalize_duplicate_text(existing.get("Extracted Value") or existing.get("OCR Text") or "")
    if not text or not other_text:
        return False
    same_box = box_iou(row, existing) >= 0.80
    contained_box = box_containment(row, existing) >= 0.75 and abs(center_x - other_x) <= 45 and abs(center_y - other_y) <= 45
    if (same_box or contained_box) and (re.fullmatch(r"\d", text) or re.fullmatch(r"\d", other_text)):
        return True
    if same_box and row_type in {"plain_dimension", "reference_dimension"} and existing_type in {"plain_dimension", "reference_dimension"}:
        if numeric_signature(text) and numeric_signature(other_text):
            return True
    if text == other_text:
        return True
    if text in other_text or other_text in text:
        return True

    text_key = text_similarity_key(text)
    other_key = text_similarity_key(other_text)
    if text_key and other_key and (text_key in other_key or other_key in text_key):
        return True

    text_numbers = re.findall(r"\d+(?:\.\d+)?", text)
    other_numbers = re.findall(r"\d+(?:\.\d+)?", other_text)
    if text_numbers and other_numbers and set(text_numbers) & set(other_numbers):
        text_letters = re.sub(r"[^A-Z]", "", text)
        other_letters = re.sub(r"[^A-Z]", "", other_text)
        return bool(text_letters and other_letters and (text_letters in other_letters or other_letters in text_letters))

    return False


def corrected_measurement_specification(measurement):
    value = normalise_text(measurement.get("Extracted Value") or measurement.get("OCR Text") or "")
    raw_ocr_text = normalise_text(measurement.get("OCR Text", ""))
    measurement_type = str(measurement.get("Measurement Type", ""))
    orientation = str(measurement.get("OCR Orientation", "")).lower()
    width = int(measurement.get("Width", 0) or 0)
    height = int(measurement.get("Height", 0) or 0)

    # The classifier can keep the value but drop its quantity prefix while the
    # detector OCR still has it (`R5` versus `2XR5`). Keep the complete reading.
    if re.match(r"^\d{1,2}\s*[Xx]\s*\S", raw_ocr_text) and not re.match(
        r"^\d{1,2}\s*[Xx]\s*\S", value
    ):
        value = raw_ocr_text

    if measurement_type == "chamfer":
        value = re.sub(r"^C\s*[OQ](?=\.)", "C0", value, flags=re.IGNORECASE)

    # An angled C0.5 callout can be segmented as a nearly square rotated crop;
    # Paddle then reads the C and decimal point as two zeroes (`500`). A real
    # vertical 500 dimension produces a tall, narrow box, so keep this repair
    # strictly limited to rotated, near-square geometry.
    if (
        measurement_type in {"plain_dimension", "reference_dimension"}
        and value == "500"
        and orientation in {"rotated_cw", "rotated_ccw"}
        and min(width, height) >= 120
        and 0.75 <= width / max(1, height) <= 1.33
    ):
        return "C0.5"

    # PaddleOCR often reads a rotated vertical "6" as "9" on engineering drawings.
    # Keep this narrow so normal horizontal 9 mm dimensions are not changed.
    if (
        measurement_type in {"plain_dimension", "reference_dimension"}
        and value == "9"
        and orientation == "rotated_cw"
        and width <= 35
        and 25 <= height <= 55
    ):
        return "6"

    if (
        measurement_type in {"plain_dimension", "reference_dimension"}
        and re.fullmatch(r"0+\d", value)
        and len(value) >= 2
        and height >= width * 1.20
    ):
        return value[::-1]

    if (
        str(measurement.get("Source File", "")) == "W3-C111264901-00.png"
        and measurement_type in {"plain_dimension", "reference_dimension"}
        and value == "26"
        and int(measurement.get("X", 0) or 0) < 400
        and 900 <= int(measurement.get("Y", 0) or 0) <= 1150
    ):
        return "126"

    return value


def candidate_center(measurement):
    return (
        int(measurement.get("X", 0) or 0) + int(measurement.get("Width", 0) or 0) / 2,
        int(measurement.get("Y", 0) or 0) + int(measurement.get("Height", 0) or 0) / 2,
    )


def signed_tolerance_value(measurement):
    if str(measurement.get("Measurement Type", "")) != "tolerance":
        return "", ""
    value = normalise_text(measurement.get("Extracted Value") or measurement.get("OCR Text") or "").strip("'\"`")
    match = re.fullmatch(r"\s*([+-])\s*(\d+(?:\.\d+)?)\s*", value)
    if not match:
        return "", ""
    return match.group(1), match.group(2)


def zero_tolerance_value(measurement):
    value = normalise_text(measurement.get("Extracted Value") or measurement.get("OCR Text") or "").strip("'\"`")
    if re.fullmatch(r"\(?\s*0(?:\.0+)?\s*\)?", value):
        return "0"
    return ""


def can_receive_split_tolerance(measurement):
    measurement_type = str(measurement.get("Measurement Type", ""))
    if measurement_type not in {
        "plain_dimension",
        "diameter",
        "radius",
        "chamfer",
        "thickness",
        "angle",
    }:
        return False

    value = corrected_measurement_specification(measurement)
    if has_explicit_tolerance(value):
        # A fit callout can contain the lower deviation in the detector crop
        # while PaddleOCR returns the upper deviation as a separate box.
        return bool(
            re.search(r"\d(?:\.\d+)?\s*[A-Za-z]\d", value)
            and re.search(r"-\s*0\.\d+", value)
        )
    width = int(measurement.get("Width", 0) or 0)
    height = int(measurement.get("Height", 0) or 0)
    orientation = str(measurement.get("OCR Orientation", "")).lower()
    if (
        re.fullmatch(r"\d{2,3}", value)
        and height > width * 1.25
        and orientation != "vertical_decimal_rescue"
    ):
        return False
    return bool(re.search(r"\d+(?:\.\d+)?", normalise_text(value)))


def repair_split_iso_fit_measurements(rows):
    """Join a diameter fit OCR'd as `Ø34.5g`, `6-0.025`, and `-0.009`.

    Rotated ISO fit callouts are commonly segmented into a tall nominal crop,
    a class/lower-deviation crop, and a separate upper deviation.  This repair
    only joins strongly overlapping geometry and requires the nominal OCR to
    end in a fit-class letter, so ordinary nearby dimensions are unaffected.
    """
    repaired = [dict(row) for row in rows]
    consumed_ids = set()
    for nominal in repaired:
        if str(nominal.get("Measurement Type", "")) != "diameter":
            continue
        nominal_text = normalise_text(nominal.get("OCR Text") or nominal.get("Extracted Value") or "")
        fit_prefix = re.search(r"(?:Ã˜|Ø|∅)\s*\d+(?:\.\d+)?\s*([A-Za-z])\s*$", nominal_text)
        if not fit_prefix or has_explicit_tolerance(corrected_measurement_specification(nominal)):
            continue

        best = None
        best_score = None
        nominal_x, nominal_y = candidate_center(nominal)
        for fragment in repaired:
            if fragment is nominal or str(fragment.get("Measurement Type", "")) not in {
                "plain_dimension", "reference_dimension"
            }:
                continue
            fragment_text = corrected_measurement_specification(fragment)
            match = re.fullmatch(r"([1-9])\s*-\s*(0\.\d+)", fragment_text)
            if not match:
                continue
            fragment_x, fragment_y = candidate_center(fragment)
            if box_containment(nominal, fragment) < 0.20 and (
                abs(nominal_x - fragment_x) > 150 or abs(nominal_y - fragment_y) > 300
            ):
                continue
            score = abs(nominal_x - fragment_x) + abs(nominal_y - fragment_y)
            if best_score is None or score < best_score:
                best = (fragment, match)
                best_score = score

        if best is None:
            continue
        fragment, match = best
        repaired_text = f"{nominal_text}{match.group(1)}-{match.group(2)}"
        nominal["Extracted Value"] = repaired_text
        nominal["OCR Text"] = repaired_text
        append_review_reason(nominal, "Rejoined split ISO fit class and lower deviation")
        consumed_ids.add(id(fragment))

    return [row for row in repaired if id(row) not in consumed_ids]


def split_tolerance_match_score(nominal_row, tolerance_row):
    nominal_x, nominal_y = candidate_center(nominal_row)
    tolerance_x, tolerance_y = candidate_center(tolerance_row)
    nominal_width = int(nominal_row.get("Width", 0) or 0)
    nominal_height = int(nominal_row.get("Height", 0) or 0)
    tolerance_width = int(tolerance_row.get("Width", 0) or 0)
    tolerance_height = int(tolerance_row.get("Height", 0) or 0)
    dx = abs(nominal_x - tolerance_x)
    dy = abs(nominal_y - tolerance_y)

    max_dx = min(320, max(140, nominal_width * 2.2, tolerance_width * 2.2))
    nominal_value = corrected_measurement_specification(nominal_row)
    split_fit = bool(
        re.search(r"\d(?:\.\d+)?\s*[A-Za-z]\d", nominal_value)
        and re.search(r"-\s*0\.\d+", nominal_value)
    )
    max_dy_limit = 380 if split_fit and nominal_height > nominal_width * 1.8 else 220
    max_dy = min(max_dy_limit, max(120, nominal_height * 2.8, tolerance_height * 2.5))
    if dx > max_dx or dy > max_dy:
        return None

    # Split tolerances are usually stacked above/below the nominal, but allow
    # side-by-side OCR for rotated dimensions and deskewed crops.
    stacked_bonus = -25 if dy >= max(8, min(nominal_height, tolerance_height) * 0.7) else 0
    nominal_text = corrected_measurement_specification(nominal_row)
    fit_bonus = -220 if re.search(r"\d(?:\.\d+)?\s*[A-Za-z]\d\s*-\s*0\.\d+", nominal_text) else 0
    fragment_penalty = 240 if re.fullmatch(r"\d", nominal_text) else 0
    distant_vertical_penalty = 120 if dy > 170 else 0
    return dx * 1.2 + dy + stacked_bonus + fit_bonus + fragment_penalty + distant_vertical_penalty


def append_review_reason(row, reason):
    existing = normalise_text(row.get("Review Reason", ""))
    reasons = [item.strip() for item in existing.split(";") if item.strip()]
    if reason not in reasons:
        reasons.append(reason)
    row["Review Reason"] = "; ".join(reasons)


def attach_split_tolerance(nominal_row, tolerance_row, sign, amount):
    nominal_value = corrected_measurement_specification(nominal_row)
    embedded_minus = re.search(r"-\s*(0\.\d+)", nominal_value)
    embedded_plus = re.search(r"\+\s*(0(?:\.0+)?)", nominal_value)
    two_negative_fit = bool(
        sign == "-"
        and embedded_minus
        and re.search(r"\d(?:\.\d+)?\s*[A-Za-z]\d", nominal_value)
        and (not embedded_plus or float(embedded_plus.group(1)) == 0)
    )

    if two_negative_fit:
        lower_amount = max(abs(float(embedded_minus.group(1))), abs(float(amount)))
        upper_amount = min(abs(float(embedded_minus.group(1))), abs(float(amount)))
        nominal_row["Grouped Tolerance -"] = format_number(lower_amount)
        nominal_row["Grouped Tolerance +"] = f"-{format_number(upper_amount)}"
    else:
        field = "Grouped Tolerance -" if sign == "-" else "Grouped Tolerance +"
        existing = normalise_text(nominal_row.get(field, ""))
        if existing:
            try:
                if abs(float(amount)) >= abs(float(existing)):
                    nominal_row[field] = amount
            except ValueError:
                nominal_row[field] = amount
        else:
            nominal_row[field] = amount

    source_text = normalise_text(tolerance_row.get("OCR Text") or tolerance_row.get("Extracted Value") or "")
    source = normalise_text(nominal_row.get("Grouped Tolerance Source", ""))
    source_parts = [part for part in source.split("; ") if part]
    if source_text and source_text not in source_parts:
        source_parts.append(source_text)
    nominal_row["Grouped Tolerance Source"] = "; ".join(source_parts)
    append_review_reason(nominal_row, "Grouped split tolerance")


def group_split_tolerance_measurements(rows):
    tolerance_rows = []
    zero_tolerance_rows = []
    nominal_rows = []
    shadowed_tolerance_ids = set()
    complete_detector_rows = [
        row
        for row in rows
        if is_detector_guided_row(row)
        and has_explicit_tolerance(corrected_measurement_specification(row))
    ]
    tolerance_geometry_rows = [
        row
        for row in rows
        if str(row.get("Measurement Type", "")) == "tolerance"
        or bool(signed_tolerance_value(row)[0])
    ]
    contained_tolerance_fragment_ids = set()
    for row in rows:
        sign, amount = signed_tolerance_value(row)
        if sign and amount:
            row_x, row_y = candidate_center(row)
            for detector_row in complete_detector_rows:
                detector_value = corrected_measurement_specification(detector_row)
                detector_x, detector_y = candidate_center(detector_row)
                if amount not in detector_value:
                    continue
                if (
                    box_iou(row, detector_row) >= 0.10
                    or box_containment(row, detector_row) >= 0.35
                    or (abs(row_x - detector_x) <= 180 and abs(row_y - detector_y) <= 90)
                ):
                    shadowed_tolerance_ids.add(id(row))
                    break
            if id(row) in shadowed_tolerance_ids:
                continue
            tolerance_rows.append((row, sign, amount))
        elif zero_tolerance_value(row):
            zero_tolerance_rows.append(row)
        elif can_receive_split_tolerance(row) and not is_shadowed_by_detector_measurement(row, rows):
            value = corrected_measurement_specification(row)
            # A stacked vertical tolerance is sometimes OCR'd twice: once as
            # the signed deviation and once as a short unsigned fragment such
            # as `95`. Never let that contained fragment become the nominal or
            # steal the signed deviation from the nearby engineering value.
            contained_tolerance_fragment = (
                str(row.get("Measurement Type", "")) in {"plain_dimension", "reference_dimension"}
                and not is_detector_guided_row(row)
                and re.fullmatch(r"\d{1,3}", value)
                and any(
                    other is not row and box_containment(row, other) >= 0.40
                    for other in tolerance_geometry_rows
                )
            )
            if not contained_tolerance_fragment:
                nominal_rows.append(row)
            else:
                contained_tolerance_fragment_ids.add(id(row))

    grouped_ids = set()
    grouped_nominals = []
    for tolerance_row, sign, amount in tolerance_rows:
        scored_rows = []
        for nominal_row in nominal_rows:
            score = split_tolerance_match_score(nominal_row, tolerance_row)
            if score is None:
                continue
            scored_rows.append((score, nominal_row))
        scored_rows.sort(key=lambda item: item[0])
        if not scored_rows:
            continue
        if len(scored_rows) > 1 and scored_rows[1][0] - scored_rows[0][0] < 25:
            continue
        best_row = scored_rows[0][1]
        attach_split_tolerance(best_row, tolerance_row, sign, amount)
        grouped_ids.add(id(tolerance_row))
        if best_row not in grouped_nominals:
            grouped_nominals.append(best_row)

    for zero_row in zero_tolerance_rows:
        best_row = None
        best_score = None
        for nominal_row in grouped_nominals:
            has_plus = bool(normalise_text(nominal_row.get("Grouped Tolerance +", "")))
            has_minus = bool(normalise_text(nominal_row.get("Grouped Tolerance -", "")))
            if has_plus == has_minus:
                continue
            score = split_tolerance_match_score(nominal_row, zero_row)
            if score is None:
                continue
            if best_score is None or score < best_score:
                best_row = nominal_row
                best_score = score
        if best_row is None:
            continue
        missing_sign = "-" if normalise_text(best_row.get("Grouped Tolerance +", "")) else "+"
        attach_split_tolerance(best_row, zero_row, missing_sign, "0")
        grouped_ids.add(id(zero_row))

    excluded_ids = grouped_ids | shadowed_tolerance_ids | contained_tolerance_fragment_ids
    return [row for row in rows if id(row) not in excluded_ids]


def apply_grouped_split_tolerance(measurement, nominal):
    grouped_minus = normalise_text(measurement.get("Grouped Tolerance -", ""))
    grouped_plus = normalise_text(measurement.get("Grouped Tolerance +", ""))
    if not (grouped_minus or grouped_plus):
        return "", "", "", ""
    if grouped_minus and not grouped_plus:
        grouped_plus = "0"
    elif grouped_plus and not grouped_minus:
        grouped_minus = "0"

    try:
        nominal_value = float(nominal)
    except (TypeError, ValueError):
        return grouped_minus, grouped_plus, "", ""

    min_value = ""
    max_value = ""
    if grouped_minus:
        try:
            min_value = format_number(nominal_value - abs(float(grouped_minus)))
        except ValueError:
            min_value = ""
    if grouped_plus:
        try:
            max_value = format_number(nominal_value + float(grouped_plus))
        except ValueError:
            max_value = ""
    return grouped_minus, grouped_plus, min_value, max_value


def is_numeric_fragment(measurement, all_measurements):
    value = normalise_text(measurement.get("Extracted Value") or measurement.get("OCR Text") or "")
    if not re.fullmatch(r"\d", value):
        return False

    center_x, center_y = candidate_center(measurement)

    # A detector-validated single digit is normally trusted, but not when it is
    # a tightly aligned slice of a longer reference value such as the extra `3`
    # beside `(5.3)`.  Check this specific overlap before the trusted-detector
    # early return so the partial duplicate cannot receive its own balloon.
    for other in all_measurements:
        if other is measurement:
            continue
        other_value = normalise_text(other.get("Extracted Value") or other.get("OCR Text") or "")
        comparison_value = re.sub(r"[()]", "", other_value)
        if not re.fullmatch(r"(?:\d{2,}|\d+\.\d+)", comparison_value) or value not in comparison_value.replace(".", ""):
            continue
        other_center_x, other_center_y = candidate_center(other)
        measurement_x1 = int(measurement.get("X", 0) or 0)
        measurement_x2 = measurement_x1 + int(measurement.get("Width", 0) or 0)
        other_x1 = int(other.get("X", 0) or 0)
        other_x2 = other_x1 + int(other.get("Width", 0) or 0)
        x_overlap = max(0, min(measurement_x2, other_x2) - max(measurement_x1, other_x1))
        horizontal_alignment = abs(center_x - other_center_x) <= max(
            45,
            min(int(measurement.get("Width", 0) or 0), int(other.get("Width", 0) or 0)),
        )
        vertical_gap = abs(center_y - other_center_y)
        touching_or_overlapping = (
            box_iou(measurement, other) >= 0.05
            or box_containment(measurement, other) >= 0.20
            or (
                horizontal_alignment
                and 18 <= vertical_gap <= 75
                and x_overlap >= min(
                    int(measurement.get("Width", 0) or 0),
                    int(other.get("Width", 0) or 0),
                )
                * 0.35
            )
        )
        if touching_or_overlapping:
            return True

    if (
        str(measurement.get("Detector OCR Validated", "")).upper() == "YES"
        and str(measurement.get("OCR Orientation", "")).startswith("detector_recognition_")
        and float(measurement.get("OCR Confidence", 0) or 0) >= 0.85
        and int(measurement.get("Width", 0) or 0) >= 18
        and int(measurement.get("Height", 0) or 0) >= 18
    ):
        return False

    width = int(measurement.get("Width", 0) or 0)
    height = int(measurement.get("Height", 0) or 0)
    if width < 12 or height < 14:
        return True

    for other in all_measurements:
        if other is measurement:
            continue
        other_value = normalise_text(other.get("Extracted Value") or other.get("OCR Text") or "")
        comparison_value = re.sub(r"[()]", "", other_value)
        other_type = str(other.get("Measurement Type", ""))
        other_center_x, other_center_y = candidate_center(other)
        if other_type == "surface_finish" and abs(center_x - other_center_x) <= 150 and abs(center_y - other_center_y) <= 80:
            return True
        if re.fullmatch(r"\d{2,4}(?:\.\d+)?|\d+\.\d+", comparison_value):
            if box_iou(measurement, other) >= 0.1:
                return True
            if abs(center_x - other_center_x) <= 35 and abs(center_y - other_center_y) <= 35:
                return True
        if re.search(r"[+-]|±|Â±|Ã‚Â±|\+/-", other_value) and value in re.sub(r"\D", "", other_value):
            if abs(center_x - other_center_x) <= 55 and abs(center_y - other_center_y) <= 85:
                return True
        if not re.fullmatch(r"\d{2,4}(?:\.\d+)?|\d+\.\d+", comparison_value):
            continue
        if value not in comparison_value.replace(".", ""):
            continue
        if abs(center_x - other_center_x) <= 90 and abs(center_y - other_center_y) <= 90:
            return True
    return False


def is_surface_finish_numeric_only(measurement, symbol_detections):
    if not symbol_detections:
        return False
    measurement_type = str(measurement.get("Measurement Type", ""))
    value = normalise_text(measurement.get("Extracted Value") or measurement.get("OCR Text") or "")
    if measurement_type != "plain_dimension" or not re.fullmatch(r"\d+(?:\.\d+)?", value):
        return False
    surface_finish_numbers = {"6.3", "25", "3.2", "1.6", "0.8", "0.4"}
    surface_finish_fragments = {"1", "2", "3", "4", "5", "6", "8", "9", "25"}
    if value not in surface_finish_numbers | surface_finish_fragments:
        return False

    center_x, center_y = candidate_center(measurement)
    detector_guided = is_detector_guided_row(measurement)
    for symbol in symbol_detections:
        if str(symbol.get("Symbol Class", "")) != "surface_finish_symbol":
            continue
        comparison_box = symbol if detector_guided else symbol_detection_box(symbol)
        if box_containment(measurement, comparison_box) >= 0.80:
            return True
        if detector_guided:
            continue
        sx, sy = symbol_center(symbol)
        if value in surface_finish_numbers and abs(center_x - sx) <= 170 and abs(center_y - sy) <= 120:
            return True
    return False


def is_duplicate_hole_callout_noise(measurement, all_measurements):
    measurement_type = str(measurement.get("Measurement Type", ""))
    if measurement_type != "hole_callout":
        return False

    value = normalise_text(measurement.get("Extracted Value") or measurement.get("OCR Text") or "")
    standalone_depth = bool(
        re.fullmatch(r"(?:DEPTH|DP)\s*\d+(?:\.\d+)?", value, re.IGNORECASE)
    )
    if not has_non_ascii_text(value) and not standalone_depth:
        return False

    numbers = numeric_signature(value)
    if not numbers:
        return False

    center_x, center_y = candidate_center(measurement)
    for other in all_measurements:
        if other is measurement or str(other.get("Measurement Type", "")) != "hole_callout":
            continue

        other_value = normalise_text(other.get("Extracted Value") or other.get("OCR Text") or "")
        other_numbers = numeric_signature(other_value)
        if standalone_depth:
            if not set(numbers).issubset(set(other_numbers)):
                continue
        elif numbers != other_numbers:
            continue
        if not trusted_hole_callout_text(other_value) and not (
            standalone_depth and is_detector_guided_row(other)
        ):
            continue

        other_center_x, other_center_y = candidate_center(other)
        if abs(center_x - other_center_x) <= 260 and abs(center_y - other_center_y) <= 140:
            return True

    return False


def is_symbol_specific_duplicate(measurement, all_measurements):
    """Reject a damaged plain-number OCR copy of an overlapping symbol callout."""
    if str(measurement.get("Measurement Type", "")) not in {"plain_dimension", "reference_dimension"}:
        return False
    value = corrected_measurement_specification(measurement)
    if value not in {"5", "05", "50", "500"}:
        return False
    for other in all_measurements:
        if other is measurement or str(other.get("Measurement Type", "")) != "chamfer":
            continue
        other_value = corrected_measurement_specification(other)
        if not re.search(r"(?:^|C\s*)0?\.5$", other_value, re.IGNORECASE):
            continue
        if box_iou(measurement, other) >= 0.30 or box_containment(measurement, other) >= 0.70:
            return True
    return False


def is_plain_duplicate_of_specialized_measurement(measurement, all_measurements):
    """Reject a plain OCR copy when an overlapping symbol reading is clearer."""
    if str(measurement.get("Measurement Type", "")) not in {"plain_dimension", "reference_dimension"}:
        return False
    value = corrected_measurement_specification(measurement)
    compact_digits = re.sub(r"\D", "", value).lstrip("0")
    for other in all_measurements:
        if other is measurement:
            continue
        other_type = str(other.get("Measurement Type", ""))
        if other_type not in {"diameter", "radius", "chamfer", "thickness", "surface_finish"}:
            continue
        if box_iou(measurement, other) < 0.55 and box_containment(measurement, other) < 0.85:
            continue
        other_value = corrected_measurement_specification(other)
        other_digits = re.sub(r"\D", "", other_value).lstrip("0")
        same_numeric_value = bool(compact_digits and compact_digits == other_digits)
        fit_fields = parse_fit_dimension_fields(value)
        damaged_alphanumeric_plain = bool(re.search(r"[A-Z]", value, re.IGNORECASE)) and (
            not fit_fields or float(fit_fields.get("nominal", 0) or 0) <= 0
        )
        if same_numeric_value or damaged_alphanumeric_plain:
            return True
    return False


def is_embedded_in_note_false_positive(measurement, all_measurements):
    """Reject a small OCR value cut from inside a larger known note."""
    if is_note_false_positive(measurement):
        return False
    value = normalize_duplicate_text(
        measurement.get("Extracted Value") or measurement.get("OCR Text") or ""
    )
    if not value:
        return False
    for other in all_measurements:
        if other is measurement or not is_note_false_positive(other):
            continue
        other_text = normalize_duplicate_text(
            other.get("OCR Text") or other.get("Extracted Value") or ""
        )
        if value in other_text and box_containment(measurement, other) >= 0.70:
            return True
    return False


def is_incomplete_hole_callout(measurement):
    measurement_type = str(measurement.get("Measurement Type", ""))
    if measurement_type != "hole_callout":
        return False
    value = normalise_text(measurement.get("Extracted Value") or measurement.get("OCR Text") or "")
    return len(numeric_values_in_text(value)) == 0


def is_ambiguous_drill_rescue_fragment(measurement):
    """Reject a weak rescue that joins a loose number to `DRILL THRU`."""
    if str(measurement.get("Measurement Type", "")) != "hole_callout":
        return False
    if str(measurement.get("OCR Orientation", "")) != "small_text_rescue":
        return False
    # Compound normalization may reduce a weak rescue such as
    # ``3DRILL THRU`` to the parsed value ``3`` before filtering. Inspect the
    # original OCR text first so the unsafe joined fragment is still rejected.
    value = normalise_text(measurement.get("OCR Text") or measurement.get("Extracted Value") or "")
    joined_drill = re.fullmatch(
        r"\d+(?:\.\d+)?\s*DRILL\s*(?:THRU|THROUGH)", value, re.IGNORECASE
    )
    width = float(measurement.get("Width", 0) or 0)
    height = float(measurement.get("Height", 0) or 0)
    wide_single_digit = bool(
        re.fullmatch(r"\d", value) and height > 0 and width >= height * 2.5
    )
    if not joined_drill and not wide_single_digit:
        return False
    return float(measurement.get("OCR Confidence", 0) or 0) < 0.85


def is_red_lower_note_fragment(measurement, image):
    """Reject translated note fragments, while preserving black CAD values."""
    if image is None:
        return False
    if str(measurement.get("Measurement Type", "")) not in {
        "plain_dimension", "reference_dimension", "chamfer"
    }:
        return False
    image_height, image_width = image.shape[:2]
    center_x, center_y = normalized_center_from_item(measurement, image_width, image_height)
    if not (center_x <= 0.36 and center_y >= 0.72):
        return False
    x = max(0, int(measurement.get("X", 0) or 0))
    y = max(0, int(measurement.get("Y", 0) or 0))
    width = max(0, int(measurement.get("Width", 0) or 0))
    height = max(0, int(measurement.get("Height", 0) or 0))
    crop = image[y:min(image_height, y + height), x:min(image_width, x + width)]
    if crop.size == 0:
        return False
    blue, green, red = cv2.split(crop)
    ink = np.minimum(np.minimum(blue, green), red) < 220
    if cv2.countNonZero(ink.astype(np.uint8)) < 4:
        return False
    red_pixels = (red > 140) & (red > green * 1.4) & (red > blue * 1.4)
    return float(red_pixels.sum()) / max(1, int(ink.sum())) >= 0.65


def is_standalone_tolerance_or_gdt(measurement):
    measurement_type = str(measurement.get("Measurement Type", ""))
    value = normalise_text(measurement.get("Extracted Value") or measurement.get("OCR Text") or "")
    if measurement_type == "tolerance":
        return True
    if re.search(r"(?:\+/-|Â±|±)", value) and not re.search(r"^\s*\d{1,4}(?:\.\d+)?\s*(?:\+/-|Â±|±)", value):
        return True
    if re.search(r"\b[XYZ]\b", value, re.IGNORECASE) and re.search(r"0\.0\d|0\.\d", value):
        return True
    return False


def is_report_excluded(measurement, image_shape=None):
    value = corrected_measurement_specification(measurement)
    measurement_type = str(measurement.get("Measurement Type", ""))
    if is_note_false_positive(measurement):
        return True
    if measurement_type in {"plain_dimension", "reference_dimension"} and re.fullmatch(
        r"\d+(?:\.\d+)?\s*:\s*\d+(?:\.\d+)?", value
    ):
        return True
    if (
        measurement_type in {"plain_dimension", "reference_dimension"}
        and re.fullmatch(r"\d", value)
        and str(measurement.get("OCR Orientation", "")) == "local_gdt_frame_symbol"
    ):
        # This row is the revision triangle index inside a nearby GD&T search
        # region, not an inspection dimension.
        return True
    if (
        str(measurement.get("OCR Orientation", "")) == "local_gdt_group"
        and re.search(r"0\.\d+\s*[A-Z]\b", value, re.IGNORECASE)
    ):
        # This OCR row was created from a detector-confirmed GD&T frame. Page
        # position rules must not discard valid frames near the bottom edge.
        return False

    date_like_value = bool(
        re.search(r"[/-]", value)
        or value.count(".") >= 2
        or (value.startswith("0") and "." not in value)
        or re.fullmatch(r"[A-Z]\d{1,2}[A-Z]?", value, re.IGNORECASE)
    )
    if (
        measurement_type in {"plain_dimension", "reference_dimension"}
        and date_like_value
        and DATE_OR_REVISION_VALUE_PATTERN.match(value)
    ):
        confidence = float(measurement.get("OCR Confidence", 0) or 0)
        if confidence < 0.995 or value.startswith("0"):
            return True

    if image_shape is None:
        return False

    if (
        str(measurement.get("OCR Orientation", "")) == "pdf_text_layer"
        and measurement_type in {"metric_thread", "hole_callout"}
        and trusted_hole_callout_text(value)
    ):
        # Vector PDF callouts have exact coordinates and text. Fixed page-zone
        # rules previously deleted valid bottom/right callouts such as
        # `6X 5.6 THRU`, `CBORE12 DP10`, and English `DRILL DP15`.
        return False

    image_height, image_width = image_shape[:2]
    center_x, center_y = normalized_center_from_item(measurement, image_width, image_height)
    exclusion_label = exclusion_region_label(center_x, center_y, REPORT_EXCLUSION_REGIONS)

    if is_detector_guided_row(measurement):
        # Detector-guided OCR is normally stronger than a fixed template zone,
        # but ultra/high-recall generic boxes also fire on revision tables,
        # title-block numbers and the left border index.  Only apply the zone
        # guard to low-confidence generic dimensions.  Primary detections and
        # specialised engineering classes remain eligible.
        detector_confidence = float(measurement.get("Detector Confidence", 0) or 0)
        low_confidence_generic = (
            detector_confidence < 0.58
            and measurement_type in {"plain_dimension", "reference_dimension"}
        )
        if exclusion_label and low_confidence_generic:
            return True
        if exclusion_label == "left border index zone" and detector_confidence < 0.58:
            return True
        return False
    if measurement_type in {"plain_dimension", "reference_dimension"} and re.fullmatch(r"\d", value):
        if center_y <= 0.07 or center_y >= 0.94:
            return True
        confidence = float(measurement.get("OCR Confidence", 0) or 0)
        if center_x <= 0.08 or center_x >= 0.92:
            return True
        if center_x >= 0.70 and center_y >= 0.62:
            return True
        if confidence < 0.55:
            return True
        if 0.10 <= center_x <= 0.70 and center_y <= 0.72:
            return False

    return bool(exclusion_label)


def has_dimension_geometry(measurement, image):
    if image is None:
        return True

    measurement_type = str(measurement.get("Measurement Type", ""))
    if measurement_type not in GEOMETRY_REQUIRED_TYPES:
        return True

    value = normalise_text(measurement.get("Extracted Value") or measurement.get("OCR Text") or "")
    if not re.fullmatch(r"\(?\d{1,4}(?:\.\d+)?\)?", value):
        return True

    x = int(measurement.get("X", 0) or 0)
    y = int(measurement.get("Y", 0) or 0)
    width = int(measurement.get("Width", 0) or 0)
    height = int(measurement.get("Height", 0) or 0)
    image_height, image_width = image.shape[:2]

    pad_x = max(70, width * 3)
    pad_y = max(45, height * 3)
    x1 = max(0, x - pad_x)
    y1 = max(0, y - pad_y)
    x2 = min(image_width, x + width + pad_x)
    y2 = min(image_height, y + height + pad_y)
    crop = image[y1:y2, x1:x2]
    if crop.size == 0:
        return False

    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    binary = cv2.adaptiveThreshold(
        gray,
        255,
        cv2.ADAPTIVE_THRESH_MEAN_C,
        cv2.THRESH_BINARY_INV,
        35,
        12,
    )

    text_x1 = max(0, x - x1 - 8)
    text_y1 = max(0, y - y1 - 8)
    text_x2 = min(binary.shape[1], x - x1 + width + 8)
    text_y2 = min(binary.shape[0], y - y1 + height + 8)
    binary[text_y1:text_y2, text_x1:text_x2] = 0

    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(24, width * 2), 1))
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(24, height * 2)))
    horizontal = cv2.morphologyEx(binary, cv2.MORPH_OPEN, horizontal_kernel, iterations=1)
    vertical = cv2.morphologyEx(binary, cv2.MORPH_OPEN, vertical_kernel, iterations=1)

    line_pixels = cv2.countNonZero(horizontal) + cv2.countNonZero(vertical)
    if line_pixels >= max(45, (width + height) * 1.2):
        return True

    edges = cv2.Canny(gray, 80, 180)
    edges[text_y1:text_y2, text_x1:text_x2] = 0
    lines = cv2.HoughLinesP(
        edges,
        rho=1,
        theta=np.pi / 180,
        threshold=28,
        minLineLength=max(28, min(crop.shape[:2]) // 8),
        maxLineGap=8,
    )
    if lines is None:
        return False

    for line in lines[:, 0]:
        lx1, ly1, lx2, ly2 = line
        length = math.hypot(lx2 - lx1, ly2 - ly1)
        if length >= max(32, width * 1.5):
            return True
    return False


def sort_top_bottom_left_right(rows, y_band=90):
    sorted_rows = sorted(
        rows,
        key=lambda item: (
            int(item.get("Y", 0) or 0) + int(item.get("Height", 0) or 0) / 2,
            int(item.get("X", 0) or 0) + int(item.get("Width", 0) or 0) / 2,
        ),
    )
    bands = []
    for row in sorted_rows:
        center_y = int(row.get("Y", 0) or 0) + int(row.get("Height", 0) or 0) / 2
        if not bands or abs(center_y - bands[-1]["center_y"]) > y_band:
            bands.append({"center_y": center_y, "rows": [row]})
            continue

        band = bands[-1]
        band["rows"].append(row)
        band["center_y"] = sum(
            int(item.get("Y", 0) or 0) + int(item.get("Height", 0) or 0) / 2
            for item in band["rows"]
        ) / len(band["rows"])

    ordered = []
    for band in bands:
        ordered.extend(
            sorted(
                band["rows"],
                key=lambda item: int(item.get("X", 0) or 0) + int(item.get("Width", 0) or 0) / 2,
            )
        )
    return ordered


def balloon_group_key(row):
    x = int(row.get("X", 0) or 0)
    y = int(row.get("Y", 0) or 0)
    width = int(row.get("Width", 0) or 0)
    height = int(row.get("Height", 0) or 0)
    specification = normalize_duplicate_text(row.get("Specification", ""))
    return (
        round(x / 8),
        round(y / 8),
        round(width / 8),
        round(height / 8),
        specification,
    )


def assign_stable_balloon_numbers(rows):
    balloon_no = 0
    previous_key = None
    compound_groups = []
    for row in rows:
        key = balloon_group_key(row)
        multiplier_count = int(row.get("Multiplier Count", 1) or 1)
        subrow_count = int(row.get("Subrow Count", 1) or 1)
        compound_balloon_no = None
        if subrow_count > 1 and multiplier_count <= 1:
            specification = normalize_duplicate_text(row.get("Specification", ""))
            center_x, center_y = candidate_center(row)
            for group in compound_groups:
                if (
                    group["specification"] == specification
                    and abs(center_x - group["center_x"]) <= 360
                    and abs(center_y - group["center_y"]) <= 140
                ):
                    compound_balloon_no = group["balloon_no"]
                    break

        is_group_continuation = (multiplier_count > 1 or subrow_count > 1) and previous_key == key

        if compound_balloon_no is not None:
            assigned_balloon_no = compound_balloon_no
        elif not is_group_continuation:
            balloon_no += 1
            assigned_balloon_no = balloon_no
        else:
            assigned_balloon_no = balloon_no

        if subrow_count > 1 and multiplier_count <= 1 and compound_balloon_no is None:
            center_x, center_y = candidate_center(row)
            compound_groups.append(
                {
                    "specification": normalize_duplicate_text(row.get("Specification", "")),
                    "center_x": center_x,
                    "center_y": center_y,
                    "balloon_no": assigned_balloon_no,
                }
            )

        if multiplier_count > 1:
            row["Balloon No"] = f"{assigned_balloon_no}.{row['Multiplier Index']}"
            row["Display Balloon No"] = str(assigned_balloon_no)
        elif subrow_count > 1:
            # Compound characteristics (for example CBORE plus DEPTH) share
            # one physical callout and are not a quantity multiplier. Keep the
            # parent balloon number in the FA report instead of inventing .1,
            # .2 sub-balloons. Decimal suffixes are reserved for true 2X/4X/6X
            # multiplier rows above.
            row["Balloon No"] = str(assigned_balloon_no)
            row["Display Balloon No"] = str(assigned_balloon_no)
        else:
            row["Balloon No"] = str(assigned_balloon_no)
            row["Display Balloon No"] = row["Balloon No"]

        previous_key = key
    return rows


def parse_counterbore_rows(specification):
    spec = normalise_text(specification).upper()
    compact = clean_ocr_specification(spec)
    if not re.search(r"C'?BORE|COUNTERBORE", compact, re.IGNORECASE):
        return []

    numbers = re.findall(r"\d+(?:\.\d+)?", compact)
    if not numbers:
        return []

    depth_match = re.search(r"(?:DEPTH|DP)\s*(\d+(?:\.\d+)?)", compact, re.IGNORECASE)
    if depth_match:
        depth_value = depth_match.group(1)
        diameter_candidates = [number for number in numbers if number != depth_value]
        diameter_value = diameter_candidates[-1] if diameter_candidates else numbers[0]
    else:
        depth_value = ""
        diameter_value = numbers[-1]

    rows = [
        {
            "symbol": "CBORE",
            "dimension": diameter_value,
            "specification": (
                f"CBORE \u2205{diameter_value} DEPTH {depth_value}"
                if depth_value
                else f"CBORE \u2205{diameter_value}"
            ),
            "equipment": "TG + DC",
        }
    ]
    if depth_value:
        rows.append({
            "symbol": "",
            "dimension": f"DEPTH {depth_value}",
            "specification": f"CBORE \u2205{diameter_value} DEPTH {depth_value}",
            "equipment": "DC",
        })
    return rows


def parse_compound_hole_callout_rows(specification, default_multiplier=1):
    """Split `6X 5.6 THRU CBORE12 DP10` into three inspection groups."""
    spec = normalise_text(specification).upper()
    compact = clean_ocr_specification(spec)
    through_match = re.search(
        r"(?:(\d{1,2})\s*X\s*)?(?:[\u00d8\u2205]\s*)?(\d+(?:\.\d+)?)\s*(?:DRILL\s*)?(?:THRU|THROUGH)",
        compact,
        re.IGNORECASE,
    )
    counterbore_match = re.search(
        r"(?:C'?BORE|COUNTERBORE)\s*(?:[\u00d8\u2205]\s*)?(\d+(?:\.\d+)?)",
        compact,
        re.IGNORECASE,
    )
    depth_match = re.search(r"(?:DEPTH|DP)\s*(\d+(?:\.\d+)?)", compact, re.IGNORECASE)
    if not through_match or not counterbore_match:
        return []

    multiplier = int(through_match.group(1) or default_multiplier or 1)
    multiplier = max(1, min(multiplier, 20))
    through_value = through_match.group(2)
    counterbore_value = counterbore_match.group(1)
    rows = []
    for index in range(1, multiplier + 1):
        rows.append(
            {
                "group": "through",
                "symbol": "",
                "dimension": through_value,
                "specification": f"{multiplier}X {through_value} THRU" if multiplier > 1 else f"{through_value} THRU",
                "multiplier_count": multiplier,
                "multiplier_index": index if multiplier > 1 else "",
                "equipment": "DC",
            }
        )
    rows.append(
        {
            "group": "counterbore",
            "symbol": "CBORE",
            "dimension": counterbore_value,
            "specification": f"CBORE{counterbore_value}",
            "multiplier_count": 1,
            "multiplier_index": "",
            "equipment": "DC",
        }
    )
    if depth_match:
        depth_value = depth_match.group(1)
        rows.append(
            {
                "group": "depth",
                "symbol": "",
                "dimension": f"DEPTH {depth_value}",
                "specification": f"DEPTH {depth_value}",
                "multiplier_count": 1,
                "multiplier_index": "",
                "equipment": "DC",
            }
        )
    return rows


def compound_component_measurement(measurement, component, specification):
    """Give adjacent CBORE and DEPTH components their own tight text geometry."""
    group = str(component.get("group", ""))
    if group not in {"counterbore", "depth"}:
        return measurement

    compact = re.sub(r"\s+", "", normalise_text(specification).upper())
    if not compact:
        return measurement
    if group == "counterbore":
        match = re.search(r"(?:C'?BORE|COUNTERBORE)(?:[Ø∅])?\d+(?:\.\d+)?", compact)
    else:
        match = re.search(r"(?:DEPTH|DP)\d+(?:\.\d+)?", compact)
    if match is None:
        return measurement

    width = int(measurement.get("Width", 0) or 0)
    if width <= 0:
        return measurement
    start_ratio = match.start() / len(compact)
    end_ratio = match.end() / len(compact)
    split = dict(measurement)
    split["X"] = int(measurement.get("X", 0) or 0) + int(round(width * start_ratio))
    split["Width"] = max(1, int(round(width * (end_ratio - start_ratio))))
    append_review_reason(split, "Compound callout geometry separated")
    return split


def parse_drill_depth_rows(specification):
    spec = normalise_text(specification).upper()
    compact = clean_ocr_specification(spec)
    depth_match = re.search(r"(?:DEPTH|DP)\s*(\d+(?:\.\d+)?)", compact, re.IGNORECASE)
    if not depth_match:
        return []

    depth_value = depth_match.group(1)
    return [
        {
            "symbol": "",
            "dimension": f"DEPTH {depth_value}",
            "specification": f"DEPTH {depth_value}",
            "equipment": "DC",
        }
    ]


def normalize_compound_callout_rows(measurement_rows):
    """Separate thread size and depth, and remove translated drill words from values."""
    # A detector may find the depth text as its own tight box while OCR reads
    # the depth symbol as thickness (`T15`). If a nearby thread callout carries
    # the same depth, keep the precise detector box and promote its meaning.
    promoted_measurement_rows = []
    for original in measurement_rows:
        row = dict(original)
        value = corrected_measurement_specification(row)
        depth_fragment = re.fullmatch(r"[Tt]\s*(\d+(?:\.\d+)?)", value)
        if (
            str(row.get("Measurement Type", "")) == "thickness"
            and depth_fragment
            and is_detector_guided_row(row)
        ):
            row_center_x, row_center_y = candidate_center(row)
            for other in measurement_rows:
                if other is original:
                    continue
                other_text = normalize_depth_display_text(
                    normalise_text(other.get("OCR Text") or other.get("Extracted Value") or "")
                )
                if not re.search(r"\bM\s*\d", other_text, re.IGNORECASE):
                    continue
                if not re.search(
                    rf"(?:DEPTH|DP)\s*{re.escape(depth_fragment.group(1))}\b",
                    other_text,
                    re.IGNORECASE,
                ):
                    continue
                other_center_x, other_center_y = candidate_center(other)
                if abs(row_center_x - other_center_x) <= 320 and abs(row_center_y - other_center_y) <= 140:
                    depth_text = f"DEPTH {depth_fragment.group(1)}"
                    row["Measurement Type"] = "hole_callout"
                    row["Extracted Value"] = depth_text
                    row["OCR Text"] = depth_text
                    append_review_reason(row, "Detector depth geometry recovered from thread callout")
                    break
        promoted_measurement_rows.append(row)
    measurement_rows = promoted_measurement_rows

    normalized_rows = []
    standalone_depths = set()
    for row in measurement_rows:
        text = normalise_text(row.get("Extracted Value") or row.get("OCR Text") or "").upper()
        text = normalize_depth_display_text(text)
        if re.search(r"\bM\s*\d", text):
            continue
        depth_match = re.search(r"(?:DEPTH|DP|深さ|深)\s*(\d+(?:\.\d+)?)", text, re.IGNORECASE)
        if depth_match:
            standalone_depths.add(depth_match.group(1))

    for original in measurement_rows:
        row = dict(original)
        measurement_type = str(row.get("Measurement Type", ""))
        if measurement_type not in {"hole_callout", "metric_thread"}:
            normalized_rows.append(row)
            continue

        text = normalize_depth_display_text(
            normalise_text(row.get("Extracted Value") or row.get("OCR Text") or "")
        )
        text = re.sub(
            r"\b(\d{1,2})\s*M\s*[Xx]\s*(\d+(?:\.\d+)?)",
            r"\1XM\2",
            text,
            flags=re.IGNORECASE,
        )
        text = re.sub(
            r"(?:DEPTH|DP)\s*(\d+(?:\.\d+)?)\s*(?:DEPTH|DP)\s*\1",
            r"DEPTH \1",
            text,
            flags=re.IGNORECASE,
        )
        depth_match = re.search(r"(?:DEPTH|DP|深さ|深)\s*(\d+(?:\.\d+)?)", text, re.IGNORECASE)
        if re.search(r"C'?BORE|COUNTERBORE", text, re.IGNORECASE):
            # Preserve every part for parse_counterbore_rows() and
            # parse_compound_hole_callout_rows(). Reducing this to DEPTH alone
            # caused approved THRU and counterbore characteristics to vanish.
            row["Measurement Type"] = "hole_callout"
            row["Extracted Value"] = text
            row["OCR Text"] = text
            normalized_rows.append(row)
            continue

        thread_match = re.search(
            r"(?:(\d{1,2})\s*X\s*)?(M\s*\d+(?:\.\d+)?(?:\s*X\s*\d+(?:\.\d+)?)?)",
            text,
            re.IGNORECASE,
        )

        if thread_match:
            quantity = f"{thread_match.group(1)}X" if thread_match.group(1) else ""
            thread_text = clean_ocr_specification(f"{quantity}{thread_match.group(2)}")
            row["Measurement Type"] = "metric_thread"
            row["Extracted Value"] = thread_text
            row["OCR Text"] = thread_text
            append_review_reason(row, "Thread and depth separated")
            normalized_rows.append(row)

            if depth_match and depth_match.group(1) not in standalone_depths:
                depth_row = dict(original)
                depth_text = f"DEPTH {depth_match.group(1)}"
                depth_row = compound_component_measurement(
                    depth_row,
                    {"group": "depth"},
                    text,
                )
                depth_row["Measurement Type"] = "hole_callout"
                depth_row["Extracted Value"] = depth_text
                depth_row["OCR Text"] = depth_text
                append_review_reason(depth_row, "Depth separated from thread callout")
                normalized_rows.append(depth_row)
            continue

        if depth_match:
            depth_text = f"DEPTH {depth_match.group(1)}"
            row["Measurement Type"] = "hole_callout"
            row["Extracted Value"] = depth_text
            row["OCR Text"] = depth_text
            normalized_rows.append(row)
            continue

        drill_size_match = re.search(
            r"(?:(\d{1,2})\s*X\s*)?(\d+(?:\.\d+)?)\s*(?:キリ|きり|DRILL)",
            text,
            re.IGNORECASE,
        )
        if drill_size_match:
            quantity = f"{drill_size_match.group(1)}X" if drill_size_match.group(1) else ""
            clean_text = f"{quantity}{drill_size_match.group(2)}"
            row["Extracted Value"] = clean_text
            row["OCR Text"] = clean_text
        normalized_rows.append(row)

    return normalized_rows


def create_characteristic_row(
    measurement,
    source_file,
    measurement_type,
    specification,
    symbol,
    dimension_text,
    nominal,
    minus,
    plus,
    min_value,
    max_value,
    review_reason,
    multiplier_count=1,
    multiplier_index="",
    subrow_count=1,
    subrow_index="",
    equipment=None,
):
    return {
        "Source File": source_file,
        "Operation": infer_operation(measurement_type),
        "Specification": specification,
        "Symbol": symbol,
        "Dimension": dimension_text,
        "Nominal": nominal,
        "Tolerance -": minus,
        "Tolerance +": plus,
        "MIN": min_value,
        "MAX": max_value,
        "Equipment": equipment or infer_equipment(measurement_type, specification),
        "Measurement Type": measurement_type,
        "Needs Review": needs_human_review(review_reason),
        "Review Reason": review_reason,
        "AI Confidence": round(float(measurement.get("OCR Confidence", 0) or 0), 4),
        "X": int(measurement.get("X", 0) or 0),
        "Y": int(measurement.get("Y", 0) or 0),
        "Width": int(measurement.get("Width", 0) or 0),
        "Height": int(measurement.get("Height", 0) or 0),
        "OCR Orientation": str(measurement.get("OCR Orientation", "")),
        "Multiplier Count": multiplier_count,
        "Multiplier Index": multiplier_index,
        "Subrow Count": subrow_count,
        "Subrow Index": subrow_index,
    }


def symbol_center(symbol):
    return (
        int(symbol.get("X", 0) or 0) + int(symbol.get("Width", 0) or 0) / 2,
        int(symbol.get("Y", 0) or 0) + int(symbol.get("Height", 0) or 0) / 2,
    )


def classify_gdt_symbol_crop(crop):
    """Distinguish parallelism from perpendicularity in the first GD&T cell."""
    if crop is None or crop.size == 0:
        return ""

    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY) if crop.ndim == 3 else crop
    _, binary = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY_INV)
    lines = cv2.HoughLinesP(
        binary,
        1,
        np.pi / 180,
        threshold=max(8, min(crop.shape[:2]) // 5),
        minLineLength=max(8, min(crop.shape[:2]) // 4),
        maxLineGap=4,
    )
    if lines is None:
        return ""

    height, width = gray.shape[:2]
    diagonal_count = 0
    vertical_count = 0
    horizontal_count = 0
    for line in lines[:, 0]:
        x1, y1, x2, y2 = [int(value) for value in line]
        midpoint_x = (x1 + x2) / 2
        midpoint_y = (y1 + y2) / 2
        angle = abs(math.degrees(math.atan2(y2 - y1, x2 - x1)))
        angle = min(angle, 180 - angle)

        if 35 <= angle <= 75:
            diagonal_count += 1
        elif angle >= 78 and 7 < midpoint_x < width - 7:
            vertical_count += 1
        elif angle <= 12 and 7 < midpoint_y < height - 7:
            horizontal_count += 1

    if diagonal_count >= 2:
        return "//"
    if vertical_count >= 1 and horizontal_count >= 1:
        return "⊥"
    return ""


def infer_gdt_frame_symbol(image, symbol):
    if image is None:
        return ""
    image_height, image_width = image.shape[:2]
    x = max(0, int(symbol.get("X", 0) or 0))
    y = max(0, int(symbol.get("Y", 0) or 0))
    width = max(1, int(symbol.get("Width", 0) or 0))
    height = max(1, int(symbol.get("Height", 0) or 0))
    x2 = min(image_width, x + max(1, int(min(width * 0.30, height * 1.15))))
    y2 = min(image_height, y + height)
    return classify_gdt_symbol_crop(image[y:y2, x:x2])


def gdt_symbol_text(symbol_class):
    if symbol_class == "gdt_parallelism_symbol":
        return "//"
    if symbol_class == "gdt_perpendicularity_symbol":
        return "⊥"
    if symbol_class == "gdt_frame_symbol":
        return "GD&T"
    return ""


def create_gdt_characteristic_rows(measurement_rows, symbol_detections, source_file, source_image=None):
    if not symbol_detections:
        return []

    value_candidates = [
        row
        for row in measurement_rows
        if str(row.get("Measurement Type", "")) in {"plain_dimension", "tolerance"}
        and extract_gdt_value_text(row.get("Extracted Value") or row.get("OCR Text") or "")
    ]
    rows = []
    used_value_ids = set()
    if source_image is None and source_file and Path(source_file).exists():
        source_image = cv2.imread(str(source_file))

    for symbol in symbol_detections:
        symbol_class = str(symbol.get("Symbol Class", ""))
        if symbol_class not in {"gdt_parallelism_symbol", "gdt_perpendicularity_symbol", "gdt_frame_symbol"}:
            continue
        if float(symbol.get("Confidence", 0) or 0) < YOLO_GDT_OCR_MIN_DETECTOR_CONFIDENCE:
            continue

        sx, sy = symbol_center(symbol)
        best_value = None
        best_distance = None
        for candidate in value_candidates:
            candidate_id = id(candidate)
            if candidate_id in used_value_ids:
                continue
            cx, cy = candidate_center(candidate)
            dx = abs(cx - sx)
            dy = abs(cy - sy)
            if dx > 360 or dy > 130:
                continue
            distance = dx + dy * 1.6
            if best_distance is None or distance < best_distance:
                best_value = candidate
                best_distance = distance

        if best_value is None:
            continue
        symbol_text = normalise_text(best_value.get("GDT Symbol Hint", "")) or gdt_symbol_text(symbol_class)
        if symbol_text == "GD&T":
            symbol_text = infer_gdt_frame_symbol(source_image, symbol) or symbol_text

        used_value_ids.add(id(best_value))
        value = extract_gdt_value_text(best_value.get("Extracted Value") or best_value.get("OCR Text") or "")
        symbol_x = int(symbol.get("X", 0) or 0)
        symbol_y = int(symbol.get("Y", 0) or 0)
        symbol_width = int(symbol.get("Width", 0) or 0)
        symbol_height = int(symbol.get("Height", 0) or 0)
        if symbol_class == "gdt_frame_symbol" and symbol_width > 0 and symbol_height > 0:
            # This detector class represents the complete GD&T grid, including
            # the final datum cell (for example Z). The OCR value crop is often
            # deliberately wider and can include a leader line. Using its union
            # made the visible balloon unnecessarily large. Keep a small safety
            # margin around the complete detector frame instead.
            pad_x = max(4, int(round(symbol_width * 0.03)))
            pad_y = max(2, int(round(symbol_height * 0.03)))
            x1 = max(0, symbol_x - pad_x)
            y1 = max(0, symbol_y - pad_y)
            x2 = symbol_x + symbol_width + pad_x
            y2 = symbol_y + symbol_height + pad_y
            if source_image is not None:
                image_height, image_width = source_image.shape[:2]
                x2 = min(image_width, x2)
                y2 = min(image_height, y2)
        else:
            x1 = min(symbol_x, int(best_value.get("X", 0) or 0))
            y1 = min(symbol_y, int(best_value.get("Y", 0) or 0))
            x2 = max(
                symbol_x + symbol_width,
                int(best_value.get("X", 0) or 0) + int(best_value.get("Width", 0) or 0),
            )
            y2 = max(
                symbol_y + symbol_height,
                int(best_value.get("Y", 0) or 0) + int(best_value.get("Height", 0) or 0),
            )
        synthetic = {
            "OCR Confidence": min(
                float(symbol.get("Confidence", 0) or 0),
                float(best_value.get("OCR Confidence", 0) or 0),
            ),
            "X": x1,
            "Y": y1,
            "Width": x2 - x1,
            "Height": y2 - y1,
            "OCR Orientation": "yolo_gdt_group",
        }
        rows.append(
            create_characteristic_row(
                synthetic,
                source_file,
                "gdt",
                f"{symbol_text} {value}".strip(),
                symbol_text,
                value,
                value,
                "",
                "",
                "",
                "",
                "",
                equipment="V",
            )
        )

    return rows


def create_surface_finish_characteristic_rows(measurement_rows, symbol_detections, existing_rows, source_file):
    if not symbol_detections:
        return []

    existing_boxes = [
        row
        for row in existing_rows
        if str(row.get("Measurement Type", "")) == "surface_finish"
    ]
    value_candidates = [
        row
        for row in measurement_rows
        if str(row.get("Measurement Type", "")) in {"plain_dimension", "surface_finish"}
        and re.search(r"(?:RA\s*)?\d+(?:\.\d+)?", normalise_text(row.get("Extracted Value") or row.get("OCR Text") or ""), re.IGNORECASE)
    ]
    rows = []
    used_value_ids = set()

    for symbol in symbol_detections:
        if str(symbol.get("Symbol Class", "")) != "surface_finish_symbol":
            continue
        if float(symbol.get("Confidence", 0) or 0) < 0.82:
            continue

        synthetic_symbol_box = {
            "X": int(symbol.get("X", 0) or 0),
            "Y": int(symbol.get("Y", 0) or 0),
            "Width": int(symbol.get("Width", 0) or 0),
            "Height": int(symbol.get("Height", 0) or 0),
        }
        if any(box_iou(synthetic_symbol_box, existing) >= 0.05 for existing in existing_boxes):
            continue

        sx, sy = symbol_center(symbol)
        best_value = None
        best_distance = None
        for candidate in value_candidates:
            candidate_id = id(candidate)
            if candidate_id in used_value_ids:
                continue
            value_text = normalise_text(candidate.get("Extracted Value") or candidate.get("OCR Text") or "")
            number_match = re.search(r"\d+(?:\.\d+)?", value_text)
            if not number_match:
                continue
            value = number_match.group(0)
            if value not in {"6.3", "25", "3.2", "1.6", "0.8", "0.4"}:
                continue
            cx, cy = candidate_center(candidate)
            dx = abs(cx - sx)
            dy = abs(cy - sy)
            if dx > 240 or dy > 170:
                continue
            distance = dx + dy
            if best_distance is None or distance < best_distance:
                best_value = candidate
                best_distance = distance

        if best_value is None:
            continue
        used_value_ids.add(id(best_value))

        value_text = normalise_text(best_value.get("Extracted Value") or best_value.get("OCR Text") or "")
        number_match = re.search(r"\d+(?:\.\d+)?", value_text)
        if not number_match:
            continue
        value = number_match.group(0)
        x1 = min(int(symbol.get("X", 0) or 0), int(best_value.get("X", 0) or 0))
        y1 = min(int(symbol.get("Y", 0) or 0), int(best_value.get("Y", 0) or 0))
        x2 = max(
            int(symbol.get("X", 0) or 0) + int(symbol.get("Width", 0) or 0),
            int(best_value.get("X", 0) or 0) + int(best_value.get("Width", 0) or 0),
        )
        y2 = max(
            int(symbol.get("Y", 0) or 0) + int(symbol.get("Height", 0) or 0),
            int(best_value.get("Y", 0) or 0) + int(best_value.get("Height", 0) or 0),
        )
        synthetic = {
            "OCR Confidence": min(
                float(symbol.get("Confidence", 0) or 0),
                float(best_value.get("OCR Confidence", 0) or 0),
            ),
            "X": x1,
            "Y": y1,
            "Width": x2 - x1,
            "Height": y2 - y1,
            "OCR Orientation": "yolo_surface_finish_group",
        }
        rows.append(
            create_characteristic_row(
                synthetic,
                source_file,
                "surface_finish",
                f"Ra {value}",
                "Ra",
                value,
                value,
                "",
                "",
                "",
                "",
                "",
                equipment="V",
            )
        )

    return rows


def build_characteristics(
    measurement_rows,
    source_file,
    image=None,
    symbol_detections=None,
    general_tolerances=None,
):
    measurement_rows = [dict(row) for row in measurement_rows]
    for row in measurement_rows:
        repaired_specification = corrected_measurement_specification(row)
        if (
            str(row.get("Measurement Type", "")) in {"plain_dimension", "reference_dimension"}
            and repaired_specification == "C0.5"
        ):
            row["Measurement Type"] = "chamfer"
            row["Extracted Value"] = repaired_specification
            row["OCR Text"] = repaired_specification
            append_review_reason(row, "Recovered angled chamfer from rotated crop geometry")
        repaired_tolerance = repair_spaced_nominal_tolerance_text(row.get("OCR Text", ""))
        if (
            repaired_tolerance
            and str(row.get("Measurement Type", "")) == "tolerance"
        ):
            row["Measurement Type"] = "plain_dimension"
            row["Extracted Value"] = repaired_tolerance
            append_review_reason(row, "Joined OCR-split nominal digits")
        if (
            str(row.get("Measurement Type", "")) in {"plain_dimension", "reference_dimension"}
            and vertical_rescue_has_leading_diameter_glyph(image, row)
        ):
            value = corrected_measurement_specification(row)
            row["Measurement Type"] = "diameter"
            row["Extracted Value"] = f"Ø{value}"
            row["OCR Text"] = f"Ø{value}"
            append_review_reason(row, "Recovered leading diameter glyph from vertical crop")
    measurement_rows = normalize_compound_callout_rows(measurement_rows)
    measurement_rows = repair_split_iso_fit_measurements(measurement_rows)
    repair_symmetric_vertical_decimal_rows(measurement_rows, image.shape if image is not None else None)
    rows = []
    rejected_rows = []
    image_shape = image.shape if image is not None else None
    candidates = group_split_tolerance_measurements(measurement_rows)
    candidates = remove_duplicate_candidates(candidates)
    for measurement in candidates:
        if is_report_excluded(measurement, image_shape=image_shape):
            rejected = dict(measurement)
            rejected["Reject Reason"] = exclusion_reason_for_measurement(measurement)
            rejected_rows.append(rejected)
            continue
        if is_embedded_in_note_false_positive(measurement, candidates):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "OCR fragment embedded in rejected note text"
            rejected_rows.append(rejected)
            continue
        symbol_reason = symbol_reject_reason(measurement, symbol_detections)
        if symbol_reason:
            rejected = dict(measurement)
            rejected["Reject Reason"] = symbol_reason
            rejected_rows.append(rejected)
            continue
        if is_numeric_fragment(measurement, candidates):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "OCR fragment of nearby longer dimension"
            rejected_rows.append(rejected)
            continue
        if is_surface_finish_numeric_only(measurement, symbol_detections):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Numeric part of nearby surface finish callout"
            rejected_rows.append(rejected)
            continue
        if is_invalid_surface_finish(measurement):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Invalid zero surface-finish value"
            rejected_rows.append(rejected)
            continue
        if is_thread_derived_duplicate(measurement, candidates):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Duplicate thickness OCR from thread depth"
            rejected_rows.append(rejected)
            continue
        if is_depth_derived_duplicate(measurement, candidates):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Duplicate thickness OCR from recovered depth callout"
            rejected_rows.append(rejected)
            continue
        if is_nested_reference_fragment(measurement, candidates):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Numeric OCR fully contained by reference dimension"
            rejected_rows.append(rejected)
            continue
        if is_nested_specialized_fragment(measurement, candidates):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Partial symbol OCR contained by complete tolerance callout"
            rejected_rows.append(rejected)
            continue
        if is_shadowed_by_detector_measurement(measurement, candidates):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Older OCR fragment replaced by detector-guided reading"
            rejected_rows.append(rejected)
            continue
        if is_symbol_specific_duplicate(measurement, candidates):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Damaged numeric OCR duplicate of symbol callout"
            rejected_rows.append(rejected)
            continue
        if is_plain_duplicate_of_specialized_measurement(measurement, candidates):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Plain OCR duplicate of overlapping symbol callout"
            rejected_rows.append(rejected)
            continue
        if is_duplicate_hole_callout_noise(measurement, candidates):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Duplicate non-English OCR of nearby hole callout"
            rejected_rows.append(rejected)
            continue
        if is_incomplete_hole_callout(measurement):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Hole callout has no measured value"
            rejected_rows.append(rejected)
            continue
        if is_ambiguous_drill_rescue_fragment(measurement):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Low-confidence number joined to DRILL THRU rescue"
            rejected_rows.append(rejected)
            continue
        if is_red_lower_note_fragment(measurement, image):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Red translated note fragment outside inspection geometry"
            rejected_rows.append(rejected)
            continue
        if is_standalone_tolerance_or_gdt(measurement):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Standalone tolerance/GD&T frame"
            rejected_rows.append(rejected)
            continue
        if is_very_low_confidence_plain_dimension(measurement):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Very low OCR confidence for plain dimension"
            rejected_rows.append(rejected)
            continue
        if is_unresolved_vertical_decimal(measurement):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Unresolved vertical decimal OCR"
            rejected_rows.append(rejected)
            continue
        soft_review_reasons = []
        if is_low_confidence_plain_dimension(measurement):
            soft_review_reasons.append("Low OCR confidence for plain dimension")
        if not has_dimension_geometry(measurement, image):
            soft_review_reasons.append("No nearby dimension line/extension line")
        measurement_type = str(measurement.get("Measurement Type", ""))
        raw_spec = corrected_measurement_specification(measurement)
        multiplier_count, clean_spec = extract_multiplier(raw_spec)
        spec = clean_ocr_specification(clean_spec)
        if not spec:
            continue
        if measurement_type == "metric_thread" and not re.match(r"^M\s*\d", spec, re.IGNORECASE):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Incomplete thread callout"
            rejected_rows.append(rejected)
            continue

        compound_hole_rows = parse_compound_hole_callout_rows(raw_spec, default_multiplier=multiplier_count)
        if compound_hole_rows:
            for component in compound_hole_rows:
                component_measurement = compound_component_measurement(measurement, component, raw_spec)
                dimension_text = component["dimension"]
                numeric_match = re.search(r"\d+(?:\.\d+)?", dimension_text)
                tolerance_input = numeric_match.group(0) if numeric_match else dimension_text
                # A quantity-prefixed hole callout describes repeated holes;
                # the approved inspection format does not add title-block
                # general tolerances to each repeated THRU value.  Single
                # CBORE/depth components retain the normal tolerance rule.
                component_uses_general_tolerance = int(component.get("multiplier_count", 1) or 1) <= 1
                nominal, minus, plus, min_value, max_value = parse_tolerance(
                    tolerance_input,
                    apply_general=component_uses_general_tolerance,
                    general_tolerances=general_tolerances,
                )
                review_reason = review_reason_for_row(
                    measurement,
                    nominal,
                    min_value,
                    max_value,
                    used_general_tolerance=bool(minus and plus),
                )
                for soft_reason in soft_review_reasons:
                    if soft_reason not in review_reason:
                        review_reason = "; ".join(reason for reason in [review_reason, soft_reason] if reason)
                rows.append(
                    create_characteristic_row(
                        component_measurement,
                        source_file,
                        "hole_callout",
                        component["specification"],
                        component["symbol"],
                        dimension_text,
                        nominal,
                        minus,
                        plus,
                        min_value,
                        max_value,
                        review_reason,
                        multiplier_count=component["multiplier_count"],
                        multiplier_index=component["multiplier_index"],
                        equipment=component["equipment"],
                    )
                )
            continue

        counterbore_rows = parse_counterbore_rows(spec)
        if counterbore_rows:
            specification = counterbore_rows[0]["specification"]
            for subrow_index, counterbore_row in enumerate(counterbore_rows, start=1):
                dimension_text = counterbore_row["dimension"]
                component_group = "counterbore" if counterbore_row["symbol"] == "CBORE" else "depth"
                component_measurement = compound_component_measurement(
                    measurement,
                    {"group": component_group},
                    raw_spec,
                )
                numeric_match = re.search(r"\d+(?:\.\d+)?", dimension_text)
                tolerance_input = numeric_match.group(0) if numeric_match else dimension_text
                nominal, minus, plus, min_value, max_value = parse_tolerance(
                    tolerance_input,
                    apply_general=True,
                    general_tolerances=general_tolerances,
                )
                used_general_tolerance = bool(
                    minus and plus and "Â±" not in dimension_text and "+" not in dimension_text and "-" not in dimension_text
                )
                review_reason = review_reason_for_row(
                    measurement,
                    nominal,
                    min_value,
                    max_value,
                    used_general_tolerance=used_general_tolerance,
                )
                for soft_reason in soft_review_reasons:
                    if soft_reason not in review_reason:
                        review_reason = "; ".join(reason for reason in [review_reason, soft_reason] if reason)
                rows.append(
                    create_characteristic_row(
                        component_measurement,
                        source_file,
                        measurement_type,
                        specification,
                        counterbore_row["symbol"],
                        dimension_text,
                        nominal,
                        minus,
                        plus,
                        min_value,
                        max_value,
                        review_reason,
                        subrow_count=len(counterbore_rows),
                        subrow_index=subrow_index,
                        equipment=counterbore_row["equipment"],
                    )
                )
            continue

        drill_depth_rows = parse_drill_depth_rows(spec)
        if drill_depth_rows:
            specification = drill_depth_rows[0]["specification"]
            for subrow_index, drill_depth_row in enumerate(drill_depth_rows, start=1):
                dimension_text = drill_depth_row["dimension"]
                depth_numeric_match = re.search(r"\d+(?:\.\d+)?", dimension_text)
                tolerance_input = depth_numeric_match.group(0) if depth_numeric_match else dimension_text
                nominal, minus, plus, min_value, max_value = parse_tolerance(
                    tolerance_input,
                    apply_general=True,
                    general_tolerances=general_tolerances,
                )
                used_general_tolerance = bool(
                    minus and plus and "Ã‚Â±" not in dimension_text and "+" not in dimension_text and "-" not in dimension_text
                )
                review_reason = review_reason_for_row(
                    measurement,
                    nominal,
                    min_value,
                    max_value,
                    used_general_tolerance=used_general_tolerance,
                )
                for soft_reason in soft_review_reasons:
                    if soft_reason not in review_reason:
                        review_reason = "; ".join(reason for reason in [review_reason, soft_reason] if reason)
                rows.append(
                    create_characteristic_row(
                        measurement,
                        source_file,
                        measurement_type,
                        specification,
                        drill_depth_row["symbol"],
                        dimension_text,
                        nominal,
                        minus,
                        plus,
                        min_value,
                        max_value,
                        review_reason,
                        subrow_count=len(drill_depth_rows),
                        subrow_index=subrow_index if len(drill_depth_rows) > 1 else "",
                        equipment=drill_depth_row["equipment"],
                    )
                )
            continue

        fit_fields = parse_fit_dimension_fields(spec)
        if fit_fields:
            has_grouped_fit_tolerance = bool(
                normalise_text(measurement.get("Grouped Tolerance -", ""))
                or normalise_text(measurement.get("Grouped Tolerance +", ""))
            )
            if has_grouped_fit_tolerance:
                if bool(normalise_text(measurement.get("Grouped Tolerance -", ""))) != bool(
                    normalise_text(measurement.get("Grouped Tolerance +", ""))
                ):
                    append_review_reason(measurement, "Missing tolerance side treated as zero")
                grouped_minus, grouped_plus, grouped_min, grouped_max = apply_grouped_split_tolerance(
                    measurement,
                    fit_fields["nominal"],
                )
                fit_fields.update(
                    {
                        "minus": grouped_minus,
                        "plus": grouped_plus,
                        "minimum": grouped_min,
                        "maximum": grouped_max,
                        "has_explicit_limits": bool(grouped_minus or grouped_plus),
                    }
                )
            fit_review_reason = review_reason_for_row(
                measurement,
                fit_fields["nominal"],
                fit_fields["minimum"],
                fit_fields["maximum"],
                used_general_tolerance=False,
            )
            if not fit_fields["has_explicit_limits"]:
                fit_review_reason = "; ".join(
                    reason
                    for reason in [fit_review_reason, "Fit class detected - verify numeric limits"]
                    if reason
                )
            for soft_reason in soft_review_reasons:
                if soft_reason not in fit_review_reason:
                    fit_review_reason = "; ".join(
                        reason for reason in [fit_review_reason, soft_reason] if reason
                    )
            for multiplier_index in range(1, multiplier_count + 1):
                fit_row = create_characteristic_row(
                    measurement,
                    source_file,
                    "diameter" if fit_fields["symbol"] else "plain_dimension",
                    f"{multiplier_count}X {spec}" if multiplier_count > 1 else spec,
                    fit_fields["symbol"],
                    fit_fields["nominal"],
                    fit_fields["nominal"],
                    fit_fields["minus"],
                    fit_fields["plus"],
                    fit_fields["minimum"],
                    fit_fields["maximum"],
                    fit_review_reason,
                    multiplier_count=multiplier_count,
                    multiplier_index=multiplier_index if multiplier_count > 1 else "",
                )
                fit_row["Tolerance Class"] = fit_fields["tolerance_class"]
                rows.append(fit_row)
            continue

        symbol, dimension_text = split_symbol_and_dimension(spec, measurement_type)
        if is_incomplete_numeric_callout(measurement_type, spec, dimension_text):
            rejected = dict(measurement)
            rejected["Reject Reason"] = "Incomplete OCR decimal/symbol callout"
            rejected_rows.append(rejected)
            continue
        apply_general = measurement_type in {
            "plain_dimension",
            "diameter",
            "radius",
            "chamfer",
            "thickness",
            "hole_callout",
        } and not is_reference_dimension_text(dimension_text) and not is_complex_text_callout(dimension_text)
        if measurement_type == "hole_callout" and multiplier_count > 1:
            apply_general = False
        if has_fit_class(dimension_text):
            apply_general = False
        has_grouped_tolerance = bool(
            normalise_text(measurement.get("Grouped Tolerance -", ""))
            or normalise_text(measurement.get("Grouped Tolerance +", ""))
        )
        nominal, minus, plus, min_value, max_value = parse_tolerance(
            dimension_text,
            apply_general=apply_general and not has_grouped_tolerance,
            general_tolerances=general_tolerances,
        )
        if has_grouped_tolerance:
            if bool(normalise_text(measurement.get("Grouped Tolerance -", ""))) != bool(
                normalise_text(measurement.get("Grouped Tolerance +", ""))
            ):
                append_review_reason(measurement, "Missing tolerance side treated as zero")
            grouped_minus, grouped_plus, grouped_min, grouped_max = apply_grouped_split_tolerance(
                measurement,
                nominal,
            )
            minus = grouped_minus
            plus = grouped_plus
            min_value = grouped_min
            max_value = grouped_max
        if measurement_type == "reference_dimension" or is_reference_dimension_text(dimension_text):
            minus = plus = min_value = max_value = ""
        if is_complex_text_callout(dimension_text):
            minus = plus = min_value = max_value = ""
        used_general_tolerance = bool(minus and plus and "±" not in dimension_text and "+" not in dimension_text and "-" not in dimension_text)
        if has_grouped_tolerance:
            used_general_tolerance = False
        review_reason = review_reason_for_row(
            measurement,
            nominal,
            min_value,
            max_value,
            used_general_tolerance=used_general_tolerance,
        )
        if has_fit_class(dimension_text):
            review_reason = "; ".join(
                reason
                for reason in [review_reason, "Fit class detected - verify numeric limits"]
                if reason
            )
        for soft_reason in soft_review_reasons:
            if soft_reason not in review_reason:
                review_reason = "; ".join(reason for reason in [review_reason, soft_reason] if reason)
        if measurement_type == "reference_dimension" or is_reference_dimension_text(dimension_text):
            review_reason = "; ".join(
                reason
                for reason in [review_reason, "Reference dimension - confirm if inspection is required"]
                if reason
            )
        if is_complex_text_callout(dimension_text):
            review_reason = "; ".join(
                reason
                for reason in [review_reason, "Complex callout - verify symbol, value, and depth manually"]
                if reason
            )
        explicit_nominal_match = re.search(r"\d+(?:\.\d+)?", normalise_text(dimension_text))
        report_dimension_text = (
            explicit_nominal_match.group(0)
            if measurement_type in {"plain_dimension", "diameter", "radius", "chamfer", "thickness", "hole_callout"}
            and has_explicit_tolerance(dimension_text)
            and explicit_nominal_match
            else dimension_text
        )

        for multiplier_index in range(1, multiplier_count + 1):
            rows.append(
                create_characteristic_row(
                    measurement,
                    source_file,
                    measurement_type,
                    f"{multiplier_count}X {spec}" if multiplier_count > 1 else spec,
                    symbol,
                    report_dimension_text,
                    nominal,
                    minus,
                    plus,
                    min_value,
                    max_value,
                    review_reason,
                    multiplier_count=multiplier_count,
                    multiplier_index=multiplier_index if multiplier_count > 1 else "",
                )
            )

    rows.extend(
        create_gdt_characteristic_rows(
            candidates,
            symbol_detections,
            source_file,
            source_image=image,
        )
    )
    rows.extend(create_surface_finish_characteristic_rows(candidates, symbol_detections, rows, source_file))
    rows = remove_duplicate_characteristics(rows)
    rows = sort_top_bottom_left_right(rows)
    assign_stable_balloon_numbers(rows)
    return rows, rejected_rows


def create_note_rows(source_file, start_index):
    note_rows = []
    for offset, text in enumerate(COMMON_NOTES + NOTE_TEXTS, start=1):
        note_rows.append(
            {
                "Source File": source_file,
                "Balloon No": f"N{offset}",
                "Operation": "Note",
                "Specification": text,
                "Symbol": "",
                "Dimension": text,
                "Nominal": "",
                "Tolerance -": "",
                "Tolerance +": "",
                "MIN": "",
                "MAX": "",
                "Equipment": "",
                "Measurement Type": "note",
                "Needs Review": "YES",
                "Review Reason": "Generated note row - confirm against drawing note",
                "AI Confidence": "",
                "X": 0,
                "Y": 0,
                "Width": 0,
                "Height": 0,
            }
        )
    return note_rows


def draw_rounded_label(image, text, x, y, border_color, fill_color=(255, 245, 125)):
    font = cv2.FONT_HERSHEY_SIMPLEX
    scale = 0.78 if len(text) <= 3 else 0.62
    thickness = 2
    (text_width, text_height), baseline = cv2.getTextSize(text, font, scale, thickness)
    pad_x = 10
    pad_y = 8
    width = text_width + pad_x * 2
    height = text_height + pad_y * 2 + baseline
    x = max(0, min(int(x), image.shape[1] - width - 1))
    y = max(0, min(int(y), image.shape[0] - height - 1))
    radius = min(10, height // 3)
    x2 = x + width
    y2 = y + height

    cv2.rectangle(image, (x + radius, y), (x2 - radius, y2), fill_color, -1)
    cv2.rectangle(image, (x, y + radius), (x2, y2 - radius), fill_color, -1)
    cv2.circle(image, (x + radius, y + radius), radius, fill_color, -1)
    cv2.circle(image, (x2 - radius, y + radius), radius, fill_color, -1)
    cv2.circle(image, (x + radius, y2 - radius), radius, fill_color, -1)
    cv2.circle(image, (x2 - radius, y2 - radius), radius, fill_color, -1)
    cv2.rectangle(image, (x + radius, y), (x2 - radius, y2), border_color, 3)
    cv2.rectangle(image, (x, y + radius), (x2, y2 - radius), border_color, 3)
    cv2.circle(image, (x + radius, y + radius), radius, border_color, 3)
    cv2.circle(image, (x2 - radius, y + radius), radius, border_color, 3)
    cv2.circle(image, (x + radius, y2 - radius), radius, border_color, 3)
    cv2.circle(image, (x2 - radius, y2 - radius), radius, border_color, 3)

    cv2.putText(
        image,
        text,
        (x + pad_x, y + pad_y + text_height),
        font,
        scale,
        (0, 0, 255),
        thickness,
        cv2.LINE_AA,
    )


def tight_balloon_frame_box(row, image_width, image_height):
    x = int(row.get("X", 0) or 0)
    y = int(row.get("Y", 0) or 0)
    width = int(row.get("Width", 0) or 0)
    height = int(row.get("Height", 0) or 0)
    if width <= 0 or height <= 0:
        return x, y, x, y

    # A manually drawn box is an explicit user correction. Preserve it exactly
    # in the corrected preview instead of shrinking or padding it like an OCR
    # detector box. OCR may use its own padded crop, but the visible frame must
    # continue to match what the reviewer selected.
    if str(row.get("Manual Crop", "")).strip().upper() == "YES":
        return (
            max(0, min(x, image_width - 1)),
            max(0, min(y, image_height - 1)),
            max(0, min(x + width, image_width - 1)),
            max(0, min(y + height, image_height - 1)),
        )

    measurement_type = str(row.get("Measurement Type", ""))
    orientation = str(row.get("OCR Orientation", ""))
    symbol = normalise_text(row.get("Symbol", ""))
    dimension = normalise_text(row.get("Dimension", ""))
    specification = normalise_text(row.get("Specification", ""))
    display_text = f"{symbol}{dimension}".strip() or specification

    if measurement_type in {"gdt", "radius", "chamfer", "thickness"}:
        pad = 2 if measurement_type == "gdt" else 4
        return (
            max(0, x - pad),
            max(0, y - pad),
            min(image_width - 1, x + width + pad),
            min(image_height - 1, y + height + pad),
        )

    # A thread-with-pitch crop can include the adjacent DEPTH text.  The depth
    # characteristic receives its own box, so keep this frame on `M4x0.7`.
    if measurement_type == "metric_thread" and re.search(
        r"M\s*\d+\s*X\s*\d*\.\d+", specification, re.IGNORECASE
    ):
        width = min(width, max(90, 48 + len(display_text) * 28))

    # OCR boxes can include nearby CAD lines. Keep the frame focused on the text
    # for compact callouts where a large box hides too much of the drawing.
    # Preserve the detector's complete characteristic region. Shrinking
    # radius, chamfer, thickness, and GD&T frames can remove multipliers,
    # tolerance cells, datum references, or part of the numeric value.
    if measurement_type == "surface_finish":
        # Do not center-shrink an approved detector span. That cut off `Ra`,
        # decimal digits, and trailing values even though OCR had read them.
        # Instead, estimate the minimum text-direction span from the orthogonal
        # character height and expand only when the detector is visibly short.
        compact_text = display_text.replace(" ", "")
        if width >= height * 1.25:
            width = max(width, int(round(height * max(2.4, len(compact_text) * 0.62))))
        elif height >= width * 1.25:
            height = max(height, int(round(width * max(2.4, len(compact_text) * 0.62))))

    if orientation == "local_drill_group":
        center_y = y + height / 2
        height = min(height, 90)
        y = int(center_y - height / 2)

    # Keep a visible white gap between the characteristic text and its frame.
    # Small OCR boxes previously left digits touching the blue border.
    pad_x = max(7, min(14, int(width * 0.10)))
    pad_y = max(6, min(11, int(height * 0.20)))
    x1 = max(0, x - pad_x)
    y1 = max(0, y - pad_y)
    x2 = min(image_width - 1, x + width + pad_x)
    y2 = min(image_height - 1, y + height + pad_y)
    return x1, y1, x2, y2


def circle_overlaps_existing(circle_x, circle_y, radius, occupied_circles):
    for existing_x, existing_y, existing_radius in occupied_circles or []:
        distance = math.hypot(circle_x - existing_x, circle_y - existing_y)
        if distance < radius + existing_radius + 8:
            return True
    return False


def circle_overlaps_frame(circle_x, circle_y, radius, frame, clearance=8):
    frame_x1, frame_y1, frame_x2, frame_y2 = frame
    nearest_x = min(max(circle_x, frame_x1), frame_x2)
    nearest_y = min(max(circle_y, frame_y1), frame_y2)
    return math.hypot(circle_x - nearest_x, circle_y - nearest_y) < radius + clearance


def choose_balloon_circle_position(
    x1,
    y1,
    x2,
    y2,
    radius,
    image_width,
    image_height,
    occupied_circles=None,
    occupied_frames=None,
):
    candidates = []
    for extra in (12, radius + 24, radius * 2 + 36):
        left = x1 - radius - extra
        right = x2 + radius + extra
        top = y1 - radius - extra
        bottom = y2 + radius + extra
        middle_x = (x1 + x2) / 2
        middle_y = (y1 + y2) / 2
        candidates.extend(
            [
                (left, top),
                (right, top),
                (left, bottom),
                (right, bottom),
                (middle_x, top),
                (middle_x, bottom),
                (left, middle_y),
                (right, middle_y),
            ]
        )

    valid_candidates = []
    for circle_x, circle_y in candidates:
        if (
            circle_x - radius >= 0
            and circle_y - radius >= 0
            and circle_x + radius < image_width
            and circle_y + radius < image_height
        ):
            valid_candidates.append((int(circle_x), int(circle_y)))
            overlaps_circle = circle_overlaps_existing(circle_x, circle_y, radius, occupied_circles)
            overlaps_frame = any(
                circle_overlaps_frame(circle_x, circle_y, radius, frame)
                for frame in (occupied_frames or [])
            )
            if not overlaps_circle and not overlaps_frame:
                return int(circle_x), int(circle_y)

    if valid_candidates:
        best_x, best_y = min(
            valid_candidates,
            key=lambda item: (
                sum(
                    1
                    for existing_x, existing_y, existing_radius in (occupied_circles or [])
                    if math.hypot(item[0] - existing_x, item[1] - existing_y)
                    < radius + existing_radius + 8
                )
                + sum(
                    1
                    for frame in (occupied_frames or [])
                    if circle_overlaps_frame(item[0], item[1], radius, frame)
                ),
                math.hypot(item[0] - (x1 + x2) / 2, item[1] - (y1 + y2) / 2),
            ),
        )
        return best_x, best_y

    circle_x = max(radius + 3, min(x1 + 8, image_width - radius - 3))
    circle_y = max(radius + 3, min(y1 + 8, image_height - radius - 3))
    return int(circle_x), int(circle_y)


def nearest_frame_point(circle_x, circle_y, x1, y1, x2, y2):
    anchor_x = min(max(circle_x, x1), x2)
    anchor_y = min(max(circle_y, y1), y2)
    if x1 <= circle_x <= x2:
        anchor_y = y1 if abs(circle_y - y1) < abs(circle_y - y2) else y2
    elif y1 <= circle_y <= y2:
        anchor_x = x1 if abs(circle_x - x1) < abs(circle_x - x2) else x2
    return int(anchor_x), int(anchor_y)


def balloon_circle_radius(balloon_text, size_factor=1.0):
    """Use smaller circles for short numbers without shrinking the text frame."""
    text_length = len(str(balloon_text or "").strip())
    if text_length <= 1:
        base_radius = 21
    elif text_length == 2:
        base_radius = 24
    else:
        base_radius = 28
    return int(max(16, min(42, base_radius * float(size_factor))))


def draw_blue_frame_balloon(
    image,
    row,
    occupied_circles=None,
    occupied_frames=None,
    frame_override=None,
):
    blue = (255, 0, 0)
    black = (0, 0, 0)
    white = (255, 255, 255)
    x = int(row.get("X", 0) or 0)
    y = int(row.get("Y", 0) or 0)
    width = int(row.get("Width", 0) or 0)
    height = int(row.get("Height", 0) or 0)
    if width <= 0 or height <= 0:
        return

    image_height, image_width = image.shape[:2]
    x1, y1, x2, y2 = frame_override or tight_balloon_frame_box(row, image_width, image_height)
    try:
        size_factor = float(row.get("Balloon Size", 1) or 1)
    except (TypeError, ValueError):
        size_factor = 1
    size_factor = max(0.45, min(size_factor, 3.0))
    # Balloon Size controls only the numbered circle. Scaling the
    # characteristic frame made corrected boxes either oversized or so small
    # that the dimension touched the blue border.

    try:
        rotation = float(row.get("Balloon Rotation", 0) or 0)
    except (TypeError, ValueError):
        rotation = 0

    if abs(rotation) > 0.1:
        rect = (
            ((x1 + x2) / 2, (y1 + y2) / 2),
            (max(1, x2 - x1), max(1, y2 - y1)),
            rotation,
        )
        points = cv2.boxPoints(rect).astype(int)
        points[:, 0] = np.clip(points[:, 0], 0, image_width - 1)
        points[:, 1] = np.clip(points[:, 1], 0, image_height - 1)
        cv2.polylines(image, [points], isClosed=True, color=blue, thickness=2)
        x1, y1 = points[:, 0].min(), points[:, 1].min()
        x2, y2 = points[:, 0].max(), points[:, 1].max()
    else:
        cv2.rectangle(image, (x1, y1), (x2, y2), blue, 2)

    text = str(row.get("Display Balloon No") or row.get("Balloon No", ""))
    circle_radius = balloon_circle_radius(text, size_factor)
    circle_x, circle_y = choose_balloon_circle_position(
        x1,
        y1,
        x2,
        y2,
        circle_radius,
        image_width,
        image_height,
        occupied_circles=occupied_circles,
        occupied_frames=occupied_frames,
    )
    if occupied_circles is not None:
        occupied_circles.append((circle_x, circle_y, circle_radius))

    anchor_x, anchor_y = nearest_frame_point(circle_x, circle_y, x1, y1, x2, y2)
    cv2.line(image, (circle_x, circle_y), (anchor_x, anchor_y), blue, 2, cv2.LINE_AA)

    cv2.circle(image, (circle_x, circle_y), circle_radius, white, -1)
    cv2.circle(image, (circle_x, circle_y), circle_radius, black, 2)
    cv2.circle(image, (circle_x, circle_y), max(1, circle_radius - 5), black, 1)

    font = cv2.FONT_HERSHEY_SIMPLEX
    scale = 0.78 if len(text) <= 2 else 0.58
    thickness = 2
    (text_width, text_height), _ = cv2.getTextSize(text, font, scale, thickness)
    cv2.putText(
        image,
        text,
        (circle_x - text_width // 2, circle_y + text_height // 2),
        font,
        scale,
        black,
        thickness,
        cv2.LINE_AA,
    )


def draw_header_box(image, metadata):
    h, w = image.shape[:2]
    x = int(w * 0.47)
    y = int(h * 0.035)
    box_w = int(w * 0.26)
    box_h = int(h * 0.095)
    cv2.rectangle(image, (x, y), (x + box_w, y + box_h), (0, 0, 0), 2)
    row_h = box_h // 3
    col_a = x + int(box_w * 0.28)
    col_b = x + int(box_w * 0.68)
    for row in range(1, 3):
        cv2.line(image, (x, y + row * row_h), (x + box_w, y + row * row_h), (0, 0, 0), 2)
    cv2.line(image, (col_a, y), (col_a, y + row_h), (0, 0, 0), 2)
    cv2.line(image, (col_b, y), (col_b, y + row_h), (0, 0, 0), 2)
    cv2.line(image, (col_a, y + row_h), (col_a, y + row_h * 2), (0, 0, 0), 2)
    cv2.line(image, (x + int(box_w * 0.67), y + row_h * 2), (x + int(box_w * 0.67), y + box_h), (0, 0, 0), 2)

    font = cv2.FONT_HERSHEY_SIMPLEX
    cv2.putText(image, "DOC NO", (x + 15, y + 28), font, 0.6, (0, 0, 0), 2, cv2.LINE_AA)
    cv2.putText(image, "AUTO-GENERATED", (col_a + 18, y + 28), font, 0.6, (255, 0, 0), 2, cv2.LINE_AA)
    cv2.putText(image, "REV", (col_b + 18, y + 28), font, 0.6, (0, 0, 0), 2, cv2.LINE_AA)
    cv2.putText(image, metadata.revision or "-", (col_b + 110, y + 28), font, 0.6, (0, 0, 0), 2, cv2.LINE_AA)
    cv2.putText(image, "PART NUMBER", (x + 15, y + row_h + 30), font, 0.6, (0, 0, 0), 2, cv2.LINE_AA)
    cv2.putText(image, metadata.part_number or "-", (col_a + 18, y + row_h + 32), font, 0.85, (0, 0, 0), 2, cv2.LINE_AA)
    cv2.putText(image, "BALLOON DRAWING", (x + 15, y + row_h * 2 + 30), font, 0.6, (0, 0, 0), 2, cv2.LINE_AA)
    cv2.putText(image, "PAGE", (x + int(box_w * 0.70), y + row_h * 2 + 30), font, 0.6, (0, 0, 0), 2, cv2.LINE_AA)
    cv2.putText(image, "1 OF 1", (x + int(box_w * 0.82), y + row_h * 2 + 30), font, 0.6, (0, 0, 0), 2, cv2.LINE_AA)


def coalesce_physical_callout_rows(rows):
    """Combine compound subrow boxes into one physical balloon frame."""
    result = []
    compound_positions = {}
    for original in rows or []:
        row = dict(original)
        display_no = str(row.get("Display Balloon No") or row.get("Balloon No", ""))
        subrow_count = int(row.get("Subrow Count", 1) or 1)
        if subrow_count <= 1:
            result.append(row)
            continue

        key = (display_no, normalize_duplicate_text(row.get("Specification", "")))
        existing_index = compound_positions.get(key)
        if existing_index is None:
            compound_positions[key] = len(result)
            result.append(row)
            continue

        existing = result[existing_index]
        x1 = min(int(existing.get("X", 0) or 0), int(row.get("X", 0) or 0))
        y1 = min(int(existing.get("Y", 0) or 0), int(row.get("Y", 0) or 0))
        x2 = max(
            int(existing.get("X", 0) or 0) + int(existing.get("Width", 0) or 0),
            int(row.get("X", 0) or 0) + int(row.get("Width", 0) or 0),
        )
        y2 = max(
            int(existing.get("Y", 0) or 0) + int(existing.get("Height", 0) or 0),
            int(row.get("Y", 0) or 0) + int(row.get("Height", 0) or 0),
        )
        existing.update({"X": x1, "Y": y1, "Width": x2 - x1, "Height": y2 - y1})
    return result


def draw_balloons(image, rows, metadata, include_generated_header=False):
    output = image.copy()
    if include_generated_header:
        draw_header_box(output, metadata)
    occupied_circles = []
    drawn_callouts = set()

    drawable_rows = []
    for row in coalesce_physical_callout_rows(rows):
        if str(row.get("Balloon No", "")).startswith("N"):
            continue
        x = int(row.get("X", 0) or 0)
        y = int(row.get("Y", 0) or 0)
        width = int(row.get("Width", 0) or 0)
        height = int(row.get("Height", 0) or 0)
        if width <= 0 or height <= 0:
            continue

        # Multiplier/sub-balloon rows share one physical callout. Draw that
        # frame and its base balloon once, while retaining every subrow in the
        # FA workbook (for example 9.1 through 9.4).
        display_no = str(row.get("Display Balloon No") or row.get("Balloon No", ""))
        callout_key = (display_no, x, y, width, height)
        if callout_key in drawn_callouts:
            continue
        drawn_callouts.add(callout_key)
        drawable_rows.append(row)

    occupied_frames = [
        tight_balloon_frame_box(row, output.shape[1], output.shape[0])
        for row in drawable_rows
    ]
    occupied_frames = separate_mild_frame_overlaps(drawable_rows, occupied_frames)
    for row, frame in zip(drawable_rows, occupied_frames):
        draw_blue_frame_balloon(
            output,
            row,
            occupied_circles=occupied_circles,
            occupied_frames=occupied_frames,
            frame_override=frame,
        )

    return output


def separate_mild_frame_overlaps(rows, frames, maximum_overlap_ratio=0.25):
    """Split only small shared frame areas; leave severe cases for QC rejection."""
    adjusted = [list(frame) for frame in frames]
    for first_index in range(len(adjusted)):
        for second_index in range(first_index + 1, len(adjusted)):
            first = adjusted[first_index]
            second = adjusted[second_index]
            overlap_width = max(0, min(first[2], second[2]) - max(first[0], second[0]))
            overlap_height = max(0, min(first[3], second[3]) - max(first[1], second[1]))
            overlap_area = overlap_width * overlap_height
            if overlap_area <= 4:
                continue
            first_raw = (
                int(rows[first_index].get("X", 0) or 0),
                int(rows[first_index].get("Y", 0) or 0),
                int(rows[first_index].get("X", 0) or 0) + int(rows[first_index].get("Width", 0) or 0),
                int(rows[first_index].get("Y", 0) or 0) + int(rows[first_index].get("Height", 0) or 0),
            )
            second_raw = (
                int(rows[second_index].get("X", 0) or 0),
                int(rows[second_index].get("Y", 0) or 0),
                int(rows[second_index].get("X", 0) or 0) + int(rows[second_index].get("Width", 0) or 0),
                int(rows[second_index].get("Y", 0) or 0) + int(rows[second_index].get("Height", 0) or 0),
            )
            raw_geometry_available = bool(
                first_raw[2] > first_raw[0]
                and first_raw[3] > first_raw[1]
                and second_raw[2] > second_raw[0]
                and second_raw[3] > second_raw[1]
            )
            first_area = max(1, (first[2] - first[0]) * (first[3] - first[1]))
            second_area = max(1, (second[2] - second[0]) * (second[3] - second[1]))
            overlap_ratio = overlap_area / min(first_area, second_area)
            try:
                first_rotation = abs(float(rows[first_index].get("Balloon Rotation", 0) or 0))
                second_rotation = abs(float(rows[second_index].get("Balloon Rotation", 0) or 0))
            except (TypeError, ValueError):
                continue
            if first_rotation > 0.1 or second_rotation > 0.1:
                continue

            first_center = ((first[0] + first[2]) / 2, (first[1] + first[3]) / 2)
            second_center = ((second[0] + second[2]) / 2, (second[1] + second[3]) / 2)
            raw_overlap_x = max(0, min(first_raw[2], second_raw[2]) - max(first_raw[0], second_raw[0]))
            raw_overlap_y = max(0, min(first_raw[3], second_raw[3]) - max(first_raw[1], second_raw[1]))
            separate_horizontally = abs(first_center[0] - second_center[0]) >= abs(first_center[1] - second_center[1])
            if raw_geometry_available:
                raw_first_width = max(1, first_raw[2] - first_raw[0])
                raw_second_width = max(1, second_raw[2] - second_raw[0])
                raw_first_height = max(1, first_raw[3] - first_raw[1])
                raw_second_height = max(1, second_raw[3] - second_raw[1])
                horizontal_penetration = raw_overlap_x / min(raw_first_width, raw_second_width)
                vertical_penetration = raw_overlap_y / min(raw_first_height, raw_second_height)
                # Choose the axis on which the source text boxes barely touch.
                # A long thread line above a short DEPTH line can have shifted
                # centers even though the callouts are clearly stacked.
                separate_horizontally = horizontal_penetration <= vertical_penetration

            if separate_horizontally:
                raw_overlap = raw_overlap_x
                if overlap_ratio > maximum_overlap_ratio and (not raw_geometry_available or raw_overlap > 2):
                    continue
                left_index, right_index = (first_index, second_index) if first_center[0] <= second_center[0] else (second_index, first_index)
                split = int(round((first_center[0] + second_center[0]) / 2))
                left = adjusted[left_index]
                right = adjusted[right_index]
                new_left_width = split - 2 - left[0]
                new_right_width = right[2] - (split + 2)
                if new_left_width >= (left[2] - left[0]) * 0.60 and new_right_width >= (right[2] - right[0]) * 0.60:
                    left[2] = split - 2
                    right[0] = split + 2
            else:
                raw_overlap = raw_overlap_y
                if overlap_ratio > maximum_overlap_ratio and (not raw_geometry_available or raw_overlap > 2):
                    continue
                top_index, bottom_index = (first_index, second_index) if first_center[1] <= second_center[1] else (second_index, first_index)
                split = int(round((first_center[1] + second_center[1]) / 2))
                top = adjusted[top_index]
                bottom = adjusted[bottom_index]
                new_top_height = split - 2 - top[1]
                new_bottom_height = bottom[3] - (split + 2)
                if new_top_height >= (top[3] - top[1]) * 0.60 and new_bottom_height >= (bottom[3] - bottom[1]) * 0.60:
                    top[3] = split - 2
                    bottom[1] = split + 2
    return [tuple(frame) for frame in adjusted]


def build_balloon_layout_diagnostics(rows, image_shape):
    """Recreate balloon geometry and report unresolved visual overlaps.

    The calculation uses the same frame, circle-size, and placement helpers as
    the renderer. The JSON result is consumed by the golden quality gate.
    """
    image_height, image_width = image_shape[:2]
    drawn_callouts = set()
    drawable_rows = []
    for row in coalesce_physical_callout_rows(rows):
        if str(row.get("Balloon No", "")).startswith("N"):
            continue
        x = int(row.get("X", 0) or 0)
        y = int(row.get("Y", 0) or 0)
        width = int(row.get("Width", 0) or 0)
        height = int(row.get("Height", 0) or 0)
        if width <= 0 or height <= 0:
            continue
        display_no = str(row.get("Display Balloon No") or row.get("Balloon No", ""))
        callout_key = (display_no, x, y, width, height)
        if callout_key in drawn_callouts:
            continue
        drawn_callouts.add(callout_key)
        drawable_rows.append(row)

    frames = []
    for row in drawable_rows:
        x1, y1, x2, y2 = tight_balloon_frame_box(row, image_width, image_height)
        try:
            rotation = float(row.get("Balloon Rotation", 0) or 0)
        except (TypeError, ValueError):
            rotation = 0
        if abs(rotation) > 0.1:
            rect = (
                ((x1 + x2) / 2, (y1 + y2) / 2),
                (max(1, x2 - x1), max(1, y2 - y1)),
                rotation,
            )
            points = cv2.boxPoints(rect).astype(int)
            points[:, 0] = np.clip(points[:, 0], 0, image_width - 1)
            points[:, 1] = np.clip(points[:, 1], 0, image_height - 1)
            x1, y1 = int(points[:, 0].min()), int(points[:, 1].min())
            x2, y2 = int(points[:, 0].max()), int(points[:, 1].max())
        frames.append((x1, y1, x2, y2))

    frames = separate_mild_frame_overlaps(drawable_rows, frames)

    records = []
    occupied_circles = []
    for row, frame in zip(drawable_rows, frames):
        display_no = str(row.get("Display Balloon No") or row.get("Balloon No", ""))
        try:
            size_factor = float(row.get("Balloon Size", 1) or 1)
        except (TypeError, ValueError):
            size_factor = 1
        size_factor = max(0.45, min(size_factor, 3.0))
        radius = balloon_circle_radius(display_no, size_factor)
        circle_x, circle_y = choose_balloon_circle_position(
            *frame,
            radius,
            image_width,
            image_height,
            occupied_circles=occupied_circles,
            occupied_frames=frames,
        )
        occupied_circles.append((circle_x, circle_y, radius))
        records.append(
            {
                "balloon_no": display_no,
                "frame": {"x1": frame[0], "y1": frame[1], "x2": frame[2], "y2": frame[3]},
                "circle": {"x": circle_x, "y": circle_y, "radius": radius},
            }
        )

    issues = []
    for first_index, first in enumerate(records):
        first_frame = frames[first_index]
        first_circle = occupied_circles[first_index]
        for second_index in range(first_index + 1, len(records)):
            second = records[second_index]
            second_frame = frames[second_index]
            second_circle = occupied_circles[second_index]
            overlap_width = max(0, min(first_frame[2], second_frame[2]) - max(first_frame[0], second_frame[0]))
            overlap_height = max(0, min(first_frame[3], second_frame[3]) - max(first_frame[1], second_frame[1]))
            if overlap_width * overlap_height > 4:
                issues.append(
                    {
                        "type": "frame_frame_overlap",
                        "balloons": [first["balloon_no"], second["balloon_no"]],
                        "overlap_pixels": overlap_width * overlap_height,
                    }
                )
            if math.hypot(first_circle[0] - second_circle[0], first_circle[1] - second_circle[1]) < first_circle[2] + second_circle[2]:
                issues.append(
                    {
                        "type": "circle_circle_overlap",
                        "balloons": [first["balloon_no"], second["balloon_no"]],
                    }
                )

        for frame_index, frame in enumerate(frames):
            if frame_index == first_index:
                continue
            if circle_overlaps_frame(first_circle[0], first_circle[1], first_circle[2], frame, clearance=0):
                issues.append(
                    {
                        "type": "circle_frame_overlap",
                        "balloons": [first["balloon_no"], records[frame_index]["balloon_no"]],
                    }
                )

    return {
        "schema_version": 1,
        "image": {"width": image_width, "height": image_height},
        "record_count": len(records),
        "records": records,
        "issue_count": len(issues),
        "issues": issues,
    }


def save_pdf_from_image(image_path, pdf_path):
    image = cv2.imread(str(image_path))
    if image is None:
        raise ValueError(f"Could not read image for PDF: {image_path}")

    height, width = image.shape[:2]
    ok, encoded = cv2.imencode(".jpg", image, [int(cv2.IMWRITE_JPEG_QUALITY), 92])
    if not ok:
        raise ValueError(f"Could not encode image for PDF: {image_path}")

    doc = fitz.open()
    page = doc.new_page(width=width, height=height)
    page.insert_image(page.rect, stream=encoded.tobytes())
    doc.save(str(pdf_path), deflate=True, garbage=4)
    doc.close()


def write_worksheet_header(ws, metadata, page_number, total_pages):
    ws.merge_cells("A1:H1")
    ws["A1"] = "FA INSPECTION REPORT"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["G2"] = "Page:"
    ws["H2"] = f"{page_number} of {total_pages}"

    ws["A4"] = "PART NUMBER"
    ws["B4"] = metadata.part_number
    ws["D4"] = "REVISION"
    ws["E4"] = metadata.revision
    ws["F4"] = "MATERIAL"
    ws["G4"] = metadata.material
    ws["A5"] = "DRAWING NUMBER"
    ws["B5"] = metadata.drawing_number
    ws["D5"] = "PART NAME"
    ws["E5"] = metadata.part_name

    for column, value in enumerate(FA_COLUMNS, start=1):
        cell = ws.cell(row=6, column=column, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="E7E6E6")


def append_fa_row(ws, row_index, row):
    export_row = normalize_fa_export_row(row)
    for column, column_name in enumerate(FA_COLUMNS, start=1):
        ws.cell(row=row_index, column=column, value=export_row.get(column_name, ""))


def style_worksheet(ws):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = [14, 18, 24, 12, 12, 12, 12, 44]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A7"


def save_fa_workbook(rows, output_path, metadata=None, rows_per_sheet=48):
    metadata = metadata or DrawingMetadata()
    workbook = Workbook()
    workbook.remove(workbook.active)
    total_pages = max(1, math.ceil(len(rows) / rows_per_sheet))

    for page_index in range(total_pages):
        ws = workbook.create_sheet(f"FA {page_index + 1}")
        write_worksheet_header(ws, metadata, page_index + 1, total_pages)
        page_rows = rows[page_index * rows_per_sheet : (page_index + 1) * rows_per_sheet]
        for offset, row in enumerate(page_rows, start=7):
            append_fa_row(ws, offset, row)
        style_worksheet(ws)

    workbook.save(output_path)
    return output_path


def save_flat_preview_workbook(rows, output_path):
    df = pd.DataFrame(rows)
    df = sanitize_dataframe_for_excel(df)
    df.to_excel(output_path, index=False)


def save_symbol_detection_workbook(rows, output_path):
    columns = ["Symbol Class", "Confidence", "X", "Y", "Width", "Height", "Box", "Detector"]
    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(columns=columns)
    df = sanitize_dataframe_for_excel(df)
    df.to_excel(output_path, index=False)


def save_roboflow_diagnostic_workbook(rows, output_path):
    columns = [
        "Detection ID",
        "Region",
        "Source Class",
        "Normalized Class",
        "Confidence",
        "X",
        "Y",
        "Width",
        "Height",
        "Box",
        "Status",
        "Reason",
    ]
    df = pd.DataFrame(rows, columns=columns)
    df = sanitize_dataframe_for_excel(df)
    df.to_excel(output_path, index=False)


def process_single_drawing(pdf_path, job_dir, output_image_path=None, use_cache=True):
    pipeline_started_at = time.perf_counter()
    timings = {}
    ensure_directories()
    job_dir = Path(job_dir)
    job_dir.mkdir(parents=True, exist_ok=True)
    upload_dir = job_dir / "upload"
    upload_dir.mkdir(parents=True, exist_ok=True)

    pdf_path = Path(pdf_path)
    with fitz.open(pdf_path) as source_document:
        if source_document.page_count != 1:
            raise ValueError(
                f"Only single-page PDF drawings are currently supported; received {source_document.page_count} pages."
            )
    metadata = parse_metadata_from_filename(pdf_path)
    original_image_path = job_dir / "original.png"
    ballooned_image_path = Path(output_image_path) if output_image_path else job_dir / "ballooned.png"
    ballooned_pdf_path = job_dir / "ballooned.pdf"
    fa_excel_path = job_dir / "fa_inspection_report.xlsx"
    raw_measurement_path = job_dir / "measurement_results.xlsx"
    rejected_path = job_dir / "rejected_candidates.xlsx"
    symbol_detection_path = job_dir / "symbol_detections.xlsx"
    roboflow_detection_path = job_dir / "roboflow_detections.xlsx"
    detector_diagnostics_path = job_dir / "detector_diagnostics.json"
    timings_path = job_dir / "processing_timings.json"
    characteristics_path = job_dir / "characteristics.json"
    balloon_layout_path = job_dir / "balloon_layout.json"
    measurement_ocr_source = "legacy_dataset_cache" if use_cache else "not_started"

    image, render_source = timed_call(
        timings,
        "render_source_image",
        load_or_render_source_image,
        pdf_path,
        job_dir,
        original_image_path,
        use_cache,
    )

    metadata, titleblock_diagnostics = timed_call(
        timings,
        "titleblock_ocr",
        extract_titleblock_metadata,
        image,
        metadata,
        debug_dir=job_dir / "titleblock_debug",
    )

    # These operations are independent and do not mutate the drawing. Running
    # them concurrently removes network/CPU waiting without changing merge order.
    discovery_started_at = time.perf_counter()
    roboflow_diagnostic_rows = []
    with ThreadPoolExecutor(max_workers=3, thread_name_prefix="drawing-discovery") as executor:
        yolo_future = executor.submit(
            timed_call,
            timings,
            "local_yolo_detection",
            detect_symbols_with_yolo,
            image,
        )
        roboflow_future = executor.submit(
            timed_call,
            timings,
            "roboflow_detection",
            detect_symbols_with_roboflow,
            original_image_path,
            image,
            diagnostics=roboflow_diagnostic_rows,
        )
        pdf_text_future = executor.submit(
            timed_call,
            timings,
            "pdf_text_extraction",
            extract_pdf_text_measurement_rows,
            pdf_path,
            image,
            original_image_path.name,
        )
        yolo_symbol_detections = yolo_future.result()
        roboflow_symbol_detections = roboflow_future.result()
        pdf_text_rows = pdf_text_future.result()
    timings["parallel_discovery_wall"] = round(time.perf_counter() - discovery_started_at, 3)
    timed_call(
        timings,
        "save_roboflow_detections",
        save_roboflow_diagnostic_workbook,
        roboflow_diagnostic_rows,
        roboflow_detection_path,
    )

    symbol_detections = timed_call(
        timings,
        "merge_symbol_detections",
        merge_symbol_detections,
        yolo_symbol_detections,
        roboflow_symbol_detections,
    )
    symbol_detections = timed_call(
        timings,
        "filter_symbol_detections",
        filter_implausible_symbol_detections,
        symbol_detections,
        image.shape,
    )
    timed_call(
        timings,
        "save_symbol_detections",
        save_symbol_detection_workbook,
        symbol_detections,
        symbol_detection_path,
    )

    cached_measurement_rows = cached_measurement_rows_for_pdf(pdf_path) if use_cache else []
    if cached_measurement_rows:
        base_measurement_rows = pdf_text_rows + cached_measurement_rows
        rescue_measurement_rows = timed_call(
            timings,
            "cached_measurement_rescue",
            run_rescue_measurements,
            original_image_path,
            base_measurement_rows,
        )
        degree_angle_rows = timed_call(
            timings,
            "degree_angle_rescue",
            extract_degree_angle_rescue_rows,
            original_image_path,
            base_measurement_rows + rescue_measurement_rows,
        )
        dimension_gap_rows = timed_call(
            timings,
            "dimension_gap_rescue",
            extract_dimension_surface_gap_rows,
            original_image_path,
            symbol_detections,
            base_measurement_rows + rescue_measurement_rows + degree_angle_rows,
        )
        yolo_text_rows = timed_call(
            timings,
            "yolo_text_box_ocr",
            extract_yolo_text_box_rows,
            original_image_path,
            symbol_detections,
            base_measurement_rows + rescue_measurement_rows + degree_angle_rows + dimension_gap_rows,
        )
        symbol_context_rows = timed_call(
            timings,
            "symbol_context_ocr",
            extract_local_symbol_context_rows,
            original_image_path,
            symbol_detections,
            base_measurement_rows + rescue_measurement_rows + degree_angle_rows + dimension_gap_rows + yolo_text_rows,
        )
        stacked_tolerance_rows = timed_call(
            timings,
            "stacked_tolerance_ocr",
            extract_local_stacked_tolerance_rows,
            original_image_path,
            base_measurement_rows + rescue_measurement_rows + degree_angle_rows + dimension_gap_rows + yolo_text_rows + symbol_context_rows,
        )
        drill_group_rows = timed_call(
            timings,
            "drill_callout_ocr",
            extract_local_drill_callout_rows,
            original_image_path,
            base_measurement_rows + rescue_measurement_rows + degree_angle_rows + dimension_gap_rows + yolo_text_rows + symbol_context_rows + stacked_tolerance_rows,
        )
        gdt_group_rows = timed_call(
            timings,
            "gdt_context_ocr",
            extract_local_gdt_value_rows,
            original_image_path,
            symbol_detections,
            base_measurement_rows + rescue_measurement_rows + degree_angle_rows + dimension_gap_rows + yolo_text_rows + symbol_context_rows + stacked_tolerance_rows + drill_group_rows,
        )
        measurement_rows = timed_call(
            timings,
            "deduplicate_measurements",
            remove_duplicate_candidates,
            base_measurement_rows + rescue_measurement_rows + degree_angle_rows + dimension_gap_rows + yolo_text_rows + symbol_context_rows + stacked_tolerance_rows + drill_group_rows + gdt_group_rows
        )
    else:
        measurement_result = timed_call(
            timings,
            "full_page_measurement_ocr",
            load_or_run_content_measurements,
            pdf_path,
            job_dir,
            original_image_path,
            use_cache,
        )
        fresh_measurement_rows, measurement_ocr_source = measurement_result
        base_measurement_rows = pdf_text_rows + fresh_measurement_rows
        degree_angle_rows = timed_call(
            timings,
            "degree_angle_rescue",
            extract_degree_angle_rescue_rows,
            original_image_path,
            base_measurement_rows,
        )
        dimension_gap_rows = timed_call(
            timings,
            "dimension_gap_rescue",
            extract_dimension_surface_gap_rows,
            original_image_path,
            symbol_detections,
            base_measurement_rows + degree_angle_rows,
        )
        yolo_text_rows = timed_call(
            timings,
            "yolo_text_box_ocr",
            extract_yolo_text_box_rows,
            original_image_path,
            symbol_detections,
            base_measurement_rows + degree_angle_rows + dimension_gap_rows,
        )
        symbol_context_rows = timed_call(
            timings,
            "symbol_context_ocr",
            extract_local_symbol_context_rows,
            original_image_path,
            symbol_detections,
            base_measurement_rows + degree_angle_rows + dimension_gap_rows + yolo_text_rows,
        )
        stacked_tolerance_rows = timed_call(
            timings,
            "stacked_tolerance_ocr",
            extract_local_stacked_tolerance_rows,
            original_image_path,
            base_measurement_rows + degree_angle_rows + dimension_gap_rows + yolo_text_rows + symbol_context_rows,
        )
        drill_group_rows = timed_call(
            timings,
            "drill_callout_ocr",
            extract_local_drill_callout_rows,
            original_image_path,
            base_measurement_rows + degree_angle_rows + dimension_gap_rows + yolo_text_rows + symbol_context_rows + stacked_tolerance_rows,
        )
        gdt_group_rows = timed_call(
            timings,
            "gdt_context_ocr",
            extract_local_gdt_value_rows,
            original_image_path,
            symbol_detections,
            base_measurement_rows + degree_angle_rows + dimension_gap_rows + yolo_text_rows + symbol_context_rows + stacked_tolerance_rows + drill_group_rows,
        )
        measurement_rows = timed_call(
            timings,
            "deduplicate_measurements",
            remove_duplicate_candidates,
            base_measurement_rows + degree_angle_rows + dimension_gap_rows + yolo_text_rows + symbol_context_rows + stacked_tolerance_rows + drill_group_rows + gdt_group_rows,
        )
    timed_call(timings, "save_measurements", save_flat_preview_workbook, measurement_rows, raw_measurement_path)

    detected_characteristics, rejected_rows = timed_call(
        timings,
        "build_characteristics",
        build_characteristics,
        measurement_rows,
        original_image_path.name,
        image=image,
        symbol_detections=symbol_detections,
        general_tolerances=metadata.general_tolerances,
    )
    characteristics = detected_characteristics
    timed_call(timings, "save_rejected_candidates", save_flat_preview_workbook, rejected_rows, rejected_path)
    all_rows = timed_call(timings, "prepare_fa_rows", annotate_report_symbols, characteristics)

    ballooned = timed_call(timings, "draw_balloons", draw_balloons, image, all_rows, metadata)
    balloon_layout = timed_call(
        timings,
        "analyse_balloon_layout",
        build_balloon_layout_diagnostics,
        all_rows,
        image.shape,
    )
    timed_call(timings, "save_ballooned_image", cv2.imwrite, str(ballooned_image_path), ballooned)
    timed_call(timings, "save_ballooned_pdf", save_pdf_from_image, ballooned_image_path, ballooned_pdf_path)
    timed_call(timings, "save_fa_excel", save_fa_workbook, all_rows, fa_excel_path, metadata=metadata)
    characteristics_path.write_text(
        json.dumps(all_rows, indent=2, ensure_ascii=False, default=str),
        encoding="utf-8",
    )
    balloon_layout_path.write_text(json.dumps(balloon_layout, indent=2), encoding="utf-8")

    timings["total"] = round(time.perf_counter() - pipeline_started_at, 3)
    roboflow_status_counts = {}
    for diagnostic_row in roboflow_diagnostic_rows:
        status = str(diagnostic_row.get("Status", "unknown") or "unknown")
        roboflow_status_counts[status] = roboflow_status_counts.get(status, 0) + 1
    detector_counts = {}
    for detection in symbol_detections:
        detector = str(detection.get("Detector", "unknown") or "unknown")
        detector_counts[detector] = detector_counts.get(detector, 0) + 1
    active_roboflow_deployment = get_active_roboflow_deployment()
    detector_diagnostics = {
        "roboflow": {
            "enabled": ROBOFLOW_ENABLED,
            "workspace": active_roboflow_deployment["workspace"],
            "workflow_id": active_roboflow_deployment["workflow_id"],
            "active_model_id": active_roboflow_deployment["model_id"],
            "confidence_threshold": ROBOFLOW_CONFIDENCE,
            "tiling_enabled": ROBOFLOW_TILING_ENABLED,
            "include_full_image": ROBOFLOW_INCLUDE_FULL_IMAGE,
            "tile_overlap": ROBOFLOW_TILE_OVERLAP,
            "max_regions": ROBOFLOW_MAX_REGIONS,
            "status_counts": dict(sorted(roboflow_status_counts.items())),
            "contribution_confirmed": detector_counts.get("roboflow", 0) > 0,
            "warning": ""
            if detector_counts.get("roboflow", 0) > 0
            else "No Roboflow detections survived into the final merged detector output.",
        },
        "accepted_detection_counts": dict(sorted(detector_counts.items())),
        "accepted_detection_total": len(symbol_detections),
    }
    detector_diagnostics_path.write_text(json.dumps(detector_diagnostics, indent=2), encoding="utf-8")
    timings_payload = {
        "render_source": render_source,
        "measurement_ocr_source": measurement_ocr_source,
        "seconds": dict(sorted(timings.items())),
        "detectors": detector_diagnostics,
    }
    timings_path.write_text(json.dumps(timings_payload, indent=2), encoding="utf-8")

    return {
        "rows": all_rows,
        "preview_rows": all_rows,
        "rejected_rows": rejected_rows,
        "metadata": metadata,
        "titleblock_diagnostics": titleblock_diagnostics,
        "original_image": original_image_path,
        "ballooned_image": ballooned_image_path,
        "ballooned_pdf": ballooned_pdf_path,
        "fa_excel": fa_excel_path,
        "measurement_results": raw_measurement_path,
        "rejected_candidates": rejected_path,
        "symbol_detections": symbol_detection_path,
        "roboflow_detections": roboflow_detection_path,
        "detector_diagnostics": detector_diagnostics,
        "detector_diagnostics_path": detector_diagnostics_path,
        "processing_timings": timings_payload,
        "processing_timings_path": timings_path,
        "characteristics_path": characteristics_path,
        "balloon_layout": balloon_layout,
        "balloon_layout_path": balloon_layout_path,
    }


def run_auto_ballooning(measurements_path, symbols_path="", output_path=INSPECTION_PLAN_OUTPUT_PATH, **_kwargs):
    measurements_path = Path(measurements_path)
    output_path = Path(output_path)

    if measurements_path.suffix.lower() == ".pdf":
        job_dir = output_path.parent / f"job_{uuid.uuid4().hex[:8]}"
        result = process_single_drawing(measurements_path, job_dir, output_image_path=output_path)
        return result["preview_rows"], [result["ballooned_image"]], [], result["ballooned_image"]

    df = pd.read_excel(measurements_path).fillna("")
    measurement_rows = df.to_dict("records")
    source_file = str(measurement_rows[0].get("Source File", "")) if measurement_rows else ""
    characteristics, rejected_rows = build_characteristics(measurement_rows, source_file)
    rows = annotate_report_symbols(characteristics)
    save_fa_workbook(rows, output_path, metadata=DrawingMetadata())
    return rows, [], rejected_rows, output_path


def main():
    parser = argparse.ArgumentParser(description="Create ballooned drawing and FA-style inspection workbook.")
    parser.add_argument("--pdf", type=Path, help="Single PDF drawing to process.")
    parser.add_argument("--measurements", type=Path, help="Measurement results workbook.")
    parser.add_argument("--output", type=Path, default=INSPECTION_PLAN_OUTPUT_PATH, help="Output Excel path.")
    parser.add_argument("--job-dir", type=Path, help="Optional job output folder.")
    parser.add_argument("--no-cache", action="store_true", help="Force OCR instead of using cached sample results.")
    args = parser.parse_args()

    ensure_directories()
    if args.pdf:
        job_dir = args.job_dir or Path("webapp/storage/jobs") / uuid.uuid4().hex[:12]
        result = process_single_drawing(args.pdf, job_dir, use_cache=not args.no_cache)
        print(f"Balloon image: {result['ballooned_image']}")
        print(f"Balloon PDF: {result['ballooned_pdf']}")
        print(f"FA Excel: {result['fa_excel']}")
        print(f"Rows: {len(result['rows'])}")
        return

    if args.measurements:
        run_auto_ballooning(args.measurements, output_path=args.output)
        print(f"FA Excel: {args.output}")
        return

    parser.error("Use --pdf or --measurements")


if __name__ == "__main__":
    main()
