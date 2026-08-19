"""Strict, all-drawing quality gate for OCR/model candidates.

This module never trains or activates a model. It only reads approved golden
artifacts and candidate outputs, writes a report, and returns pass/reject.
"""

import argparse
import hashlib
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from compare_golden_output import (
    FA_COLUMNS,
    clean,
    compare_characteristic_values,
    compare_rows,
    compare_characteristic_geometry,
    duplicate_balloon_identities,
    load_characteristic_records,
    read_fa_document,
)


REQUIRED_METADATA = ("part_number", "drawing_number", "revision", "material", "part_name")


def invalid_duplicate_balloon_identities(excel_rows, characteristic_rows):
    """Reject duplicate IDs unless records declare one compound balloon."""
    duplicates = duplicate_balloon_identities(excel_rows)
    invalid = []
    for identity in duplicates:
        report_count = sum(
            1
            for row in excel_rows or []
            if clean(row.get("BALLOON NO", row.get("Balloon No", ""))) == identity
        )
        matching = [
            row
            for row in characteristic_rows or []
            if clean(row.get("Balloon No", row.get("BALLOON NO", ""))) == identity
        ]
        try:
            declared_counts = {int(row.get("Subrow Count", 0) or 0) for row in matching}
            indexes = {int(row.get("Subrow Index", 0) or 0) for row in matching}
        except (TypeError, ValueError):
            invalid.append(identity)
            continue
        valid_compound = (
            report_count > 1
            and len(matching) == report_count
            and declared_counts == {report_count}
            and indexes == set(range(1, report_count + 1))
        )
        if not valid_compound:
            invalid.append(identity)
    return invalid


def comparable_metadata_value(field, value):
    """Normalize harmless punctuation spacing without hiding real changes."""
    normalized = clean(value)
    if field == "part_name":
        normalized = re.sub(r"\s*,\s*", ",", normalized)
    return normalized


def sha256_file(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def load_json(path):
    return json.loads(Path(path).read_text(encoding="utf-8"))


def load_manifests(golden_root):
    manifests = []
    for path in Path(golden_root).glob("*_golden_manifest.json"):
        payload = load_json(path)
        payload["_manifest_path"] = str(path)
        manifests.append(payload)
    return sorted(manifests, key=lambda item: int(item.get("golden_test_number", 9999)))


def check(name, passed, expected="", actual="", detail="", skipped=False):
    return {
        "name": name,
        "status": "SKIP" if skipped else ("PASS" if passed else "FAIL"),
        "expected": expected,
        "actual": actual,
        "detail": detail,
    }


def resolve_candidate_file(root, drawing_number, filename):
    if root is None:
        return None
    return Path(root) / drawing_number / filename


def verify_manifest(manifest, golden_root):
    checks = []
    root = Path(golden_root)
    checks.append(check("manifest approved", manifest.get("status") == "approved", "approved", manifest.get("status", "")))
    for field in ("source_pdf", "expected_pdf", "expected_excel", "expected_characteristics"):
        record = manifest.get(field) or {}
        artifact = root / str(record.get("path", ""))
        exists = artifact.is_file()
        checks.append(check(f"{field} exists", exists, "file present", str(artifact)))
        if exists:
            actual_hash = sha256_file(artifact)
            expected_hash = clean(record.get("sha256", "")).upper()
            checks.append(check(f"{field} hash", actual_hash == expected_hash, expected_hash, actual_hash))

    excel_record = manifest.get("expected_excel") or {}
    excel_path = root / str(excel_record.get("path", ""))
    if excel_path.is_file():
        document = read_fa_document(excel_path)
        expected_count = int(manifest.get("approved_characteristics", 0) or 0)
        checks.append(check("approved characteristic count", len(document["rows"]) == expected_count, expected_count, len(document["rows"])))
        characteristic_record = manifest.get("expected_characteristics") or {}
        characteristic_path = root / str(characteristic_record.get("path", ""))
        characteristic_rows = (
            load_characteristic_records(characteristic_path)
            if characteristic_path.is_file()
            else []
        )
        duplicates = invalid_duplicate_balloon_identities(
            document["rows"], characteristic_rows
        )
        checks.append(check("golden balloon identities unique", not duplicates, "none", ", ".join(duplicates)))
        if manifest.get("main_balloons") is not None:
            main_identities = {
                identity.split(".", 1)[0]
                for identity in (clean(row.get("BALLOON NO", "")) for row in document["rows"])
                if identity.split(".", 1)[0].isdigit()
            }
            expected_main = int(manifest.get("main_balloons", 0) or 0)
            checks.append(check("main balloon count", len(main_identities) == expected_main, expected_main, len(main_identities)))
    return checks


def layout_signature(layout):
    records = layout.get("records", []) if isinstance(layout, dict) else []
    selected = []
    for record in records:
        selected.append(
            {
                "balloon_no": clean(record.get("balloon_no", "")),
                "frame": record.get("frame", {}),
                "circle": record.get("circle", {}),
            }
        )
    return selected


def load_seconds(path):
    if path is None or not Path(path).is_file():
        return None
    payload = load_json(path)
    seconds = payload.get("seconds", {}).get("total")
    try:
        return float(seconds)
    except (TypeError, ValueError):
        return None


def compare_candidate(manifest, golden_root, candidate_root, repeat_root=None, baseline_root=None, promotion=False):
    drawing_number = manifest["drawing_number"]
    expected_excel = Path(golden_root) / manifest["expected_excel"]["path"]
    candidate_excel = resolve_candidate_file(candidate_root, drawing_number, "fa_inspection_report.xlsx")
    candidate_layout = resolve_candidate_file(candidate_root, drawing_number, "balloon_layout.json")
    candidate_characteristics = resolve_candidate_file(candidate_root, drawing_number, "characteristics.json")
    candidate_pdf = resolve_candidate_file(candidate_root, drawing_number, "ballooned.pdf")
    candidate_timings = resolve_candidate_file(candidate_root, drawing_number, "processing_timings.json")
    result = {
        "golden_test_number": manifest.get("golden_test_number"),
        "drawing_number": drawing_number,
        "checks": [],
        "differences": [],
        "geometry_differences": [],
        "baseline_differences": None,
    }
    checks = result["checks"]

    if not candidate_excel or not candidate_excel.is_file():
        checks.append(check("candidate Excel exists", False, "file present", str(candidate_excel)))
        return result

    checks.append(check("candidate Excel exists", True, "file present", str(candidate_excel)))
    expected = read_fa_document(expected_excel)
    current = read_fa_document(candidate_excel)
    checks.append(check("characteristic count", len(current["rows"]) == len(expected["rows"]), len(expected["rows"]), len(current["rows"])))
    expected_characteristics = Path(golden_root) / manifest["expected_characteristics"]["path"]
    expected_geometry = load_characteristic_records(expected_characteristics)
    if candidate_characteristics and candidate_characteristics.is_file():
        current_geometry = load_characteristic_records(candidate_characteristics)
        duplicates = invalid_duplicate_balloon_identities(
            current["rows"], current_geometry
        )
        checks.append(check("balloon identities unique", not duplicates, "none", ", ".join(duplicates)))
        comparable_fields = {
            "Dimension",
            "Tolerance -",
            "Tolerance +",
            "MIN",
            "MAX",
            "Measurement Type",
            "Report Symbol",
            "Symbol",
        }
        has_expected_values = any(
            clean(row.get(field, ""))
            for row in expected_geometry
            for field in comparable_fields
        )
        has_current_values = any(
            clean(row.get(field, ""))
            for row in current_geometry
            for field in comparable_fields
        )
        if has_expected_values and has_current_values:
            differences = compare_characteristic_values(expected_geometry, current_geometry)
        else:
            # Some legacy positional references contain only box geometry.
            # They cannot prove that the generated inspection values match,
            # so the approved and candidate Excel rows remain the authority.
            differences = compare_rows(expected["rows"], current["rows"])
        result["differences"] = differences
        checks.append(check("all characteristic values", not differences, "0 positional differences", len(differences)))
        geometry_differences = compare_characteristic_geometry(expected_geometry, current_geometry)
        result["geometry_differences"] = geometry_differences
        checks.append(check("characteristic geometry", not geometry_differences, "0 differences", len(geometry_differences), json.dumps(geometry_differences[:5], ensure_ascii=False)))
    else:
        duplicates = duplicate_balloon_identities(current["rows"])
        checks.append(check("balloon identities unique", not duplicates, "none", ", ".join(duplicates)))
        differences = compare_rows(expected["rows"], current["rows"])
        result["differences"] = differences
        checks.append(check("all characteristic values", not differences, "0 differences", len(differences)))
        checks.append(check("candidate characteristics exist", False, "characteristics.json", str(candidate_characteristics)))

    checks.append(check("candidate ballooned PDF exists", bool(candidate_pdf and candidate_pdf.is_file()), "ballooned.pdf", str(candidate_pdf)))

    for field in REQUIRED_METADATA:
        expected_value = comparable_metadata_value(field, expected["metadata"].get(field, ""))
        current_value = comparable_metadata_value(field, current["metadata"].get(field, ""))
        checks.append(check(f"metadata {field}", current_value == expected_value, expected_value, current_value))

    current_layout_payload = None
    if candidate_layout and candidate_layout.is_file():
        current_layout_payload = load_json(candidate_layout)
        issues = current_layout_payload.get("issues", [])
        checks.append(check("balloon layout has no overlap", not issues, "0 issues", len(issues), json.dumps(issues[:5], ensure_ascii=False)))
        expected_main = int(manifest.get("main_balloons", 0) or 0)
        actual_main = int(current_layout_payload.get("record_count", len(current_layout_payload.get("records", []))) or 0)
        checks.append(check("balloon layout count", actual_main == expected_main, expected_main, actual_main))
    else:
        checks.append(check("balloon layout diagnostics exist", not promotion, "balloon_layout.json", str(candidate_layout), skipped=not promotion))

    if repeat_root is None:
        checks.append(check("repeat run is identical", not promotion, "repeat output", "not supplied", skipped=not promotion))
    else:
        repeat_excel = resolve_candidate_file(repeat_root, drawing_number, "fa_inspection_report.xlsx")
        repeat_layout = resolve_candidate_file(repeat_root, drawing_number, "balloon_layout.json")
        if repeat_excel and repeat_excel.is_file():
            repeated = read_fa_document(repeat_excel)
            repeat_differences = compare_rows(current["rows"], repeated["rows"])
            metadata_same = current["metadata"] == repeated["metadata"]
            layout_same = True
            if current_layout_payload is not None:
                layout_same = bool(repeat_layout and repeat_layout.is_file()) and layout_signature(current_layout_payload) == layout_signature(load_json(repeat_layout))
            checks.append(check("repeat run is identical", not repeat_differences and metadata_same and layout_same, "identical", f"row differences={len(repeat_differences)}, metadata_same={metadata_same}, layout_same={layout_same}"))
        else:
            checks.append(check("repeat run is identical", False, "repeat output", str(repeat_excel)))

    if baseline_root is None:
        checks.append(check("no regression versus active model", not promotion, "active baseline", "not supplied", skipped=not promotion))
        checks.append(check("processing time increase", not promotion, "<= 10%", "not supplied", skipped=not promotion))
    else:
        baseline_excel = resolve_candidate_file(baseline_root, drawing_number, "fa_inspection_report.xlsx")
        baseline_characteristics = resolve_candidate_file(baseline_root, drawing_number, "characteristics.json")
        baseline_timings = resolve_candidate_file(baseline_root, drawing_number, "processing_timings.json")
        if baseline_excel and baseline_excel.is_file():
            baseline = read_fa_document(baseline_excel)
            if baseline_characteristics and baseline_characteristics.is_file():
                baseline_differences = compare_characteristic_values(
                    expected_geometry,
                    load_characteristic_records(baseline_characteristics),
                )
            else:
                baseline_differences = compare_rows(expected["rows"], baseline["rows"])
            result["baseline_differences"] = len(baseline_differences)
            checks.append(check("no regression versus active model", len(differences) <= len(baseline_differences), f"<= {len(baseline_differences)} differences", len(differences)))
        else:
            checks.append(check("no regression versus active model", False, "active baseline Excel", str(baseline_excel)))

        candidate_seconds = load_seconds(candidate_timings)
        baseline_seconds = load_seconds(baseline_timings)
        if candidate_seconds is not None and baseline_seconds is not None:
            limit = baseline_seconds * 1.10
            checks.append(check("processing time increase", candidate_seconds <= limit, f"<= {limit:.3f}s", f"{candidate_seconds:.3f}s"))
        else:
            checks.append(check("processing time increase", False, "candidate and baseline timings", f"candidate={candidate_seconds}, baseline={baseline_seconds}"))

    return result


def overall_status(drawings, promotion=False):
    failed = any(item["status"] == "FAIL" for drawing in drawings for item in drawing["checks"])
    skipped = any(item["status"] == "SKIP" for drawing in drawings for item in drawing["checks"])
    if failed or (promotion and skipped):
        return "REJECTED"
    return "PASSED"


def write_reports(report_dir, payload):
    report_dir = Path(report_dir)
    report_dir.mkdir(parents=True, exist_ok=True)
    json_path = report_dir / "golden_gate_report.json"
    json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Golden gate", payload["status"]])
    summary.append(["Created UTC", payload["created_at"]])
    summary.append([])
    summary.append(["Drawing", "Gate", "Status", "Expected", "Actual", "Detail"])
    for drawing in payload["drawings"]:
        for item in drawing["checks"]:
            summary.append([drawing["drawing_number"], item["name"], item["status"], item["expected"], item["actual"], item["detail"]])

    differences_sheet = workbook.create_sheet("Differences")
    differences_sheet.append(["Drawing", "Expected row", "Current row", "Column", "Expected", "Current"])
    for drawing in payload["drawings"]:
        for item in drawing.get("differences", []):
            differences_sheet.append(
                [drawing["drawing_number"], item["EXPECTED ROW"], item["CURRENT ROW"], item["COLUMN"], item["EXPECTED"], item["CURRENT"]]
            )
    geometry_sheet = workbook.create_sheet("Geometry Differences")
    geometry_sheet.append(["Drawing", "Balloon", "Type", "Expected", "Current", "Detail"])
    for drawing in payload["drawings"]:
        for item in drawing.get("geometry_differences", []):
            geometry_sheet.append(
                [
                    drawing["drawing_number"],
                    item.get("balloon_no", ""),
                    item.get("type", ""),
                    json.dumps(item.get("expected", ""), ensure_ascii=False),
                    json.dumps(item.get("current", ""), ensure_ascii=False),
                    json.dumps(item, ensure_ascii=False),
                ]
            )
    xlsx_path = report_dir / "golden_gate_report.xlsx"
    workbook.save(xlsx_path)
    return json_path, xlsx_path


def run_gate(golden_root, candidate_root=None, repeat_root=None, baseline_root=None, report_dir=None, verify_only=False, promotion=False, candidate_id=""):
    manifests = load_manifests(golden_root)
    if len(manifests) != 5:
        raise ValueError(f"Expected exactly 5 approved golden manifests, found {len(manifests)}.")

    drawings = []
    for manifest in manifests:
        manifest_checks = verify_manifest(manifest, golden_root)
        if verify_only:
            drawings.append({"golden_test_number": manifest.get("golden_test_number"), "drawing_number": manifest.get("drawing_number", ""), "checks": manifest_checks, "differences": []})
        else:
            evaluated = compare_candidate(manifest, golden_root, candidate_root, repeat_root, baseline_root, promotion)
            evaluated["checks"] = manifest_checks + evaluated["checks"]
            drawings.append(evaluated)

    if baseline_root is not None and promotion:
        improved = any(
            drawing.get("baseline_differences") is not None
            and len(drawing.get("differences", [])) < drawing["baseline_differences"]
            for drawing in drawings
        )
        for drawing in drawings:
            drawing["checks"].append(check("candidate improves at least one known error", improved, "at least one improvement", improved))

    payload = {
        "schema_version": 1,
        "created_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "mode": "verify_only" if verify_only else ("promotion" if promotion else "candidate_check"),
        "status": overall_status(drawings, promotion=promotion),
        "golden_count": len(manifests),
        "candidate_root": str(candidate_root or ""),
        "repeat_root": str(repeat_root or ""),
        "baseline_root": str(baseline_root or ""),
        "candidate_id": clean(candidate_id),
        "drawings": drawings,
    }
    if report_dir:
        json_path, xlsx_path = write_reports(report_dir, payload)
        payload["report_json"] = str(json_path)
        payload["report_excel"] = str(xlsx_path)
    return payload


def main():
    parser = argparse.ArgumentParser(description="Run the strict five-drawing OCR golden quality gate.")
    parser.add_argument("--golden-root", default=str(Path(__file__).resolve().parents[1] / "golden_tests"))
    parser.add_argument("--candidate-root")
    parser.add_argument("--repeat-root")
    parser.add_argument("--baseline-root")
    parser.add_argument("--report-dir", default=str(Path(__file__).resolve().parents[1] / "golden_tests" / "comparison_reports" / "latest"))
    parser.add_argument("--verify-only", action="store_true")
    parser.add_argument("--promotion", action="store_true", help="Require layout, repeat-run, active-baseline, improvement, and timing gates.")
    parser.add_argument("--candidate-id", default="", help="Candidate model/version ID recorded in the gate report.")
    args = parser.parse_args()
    if not args.verify_only and not args.candidate_root:
        parser.error("--candidate-root is required unless --verify-only is used.")
    if args.promotion and (not args.repeat_root or not args.baseline_root):
        parser.error("--promotion requires --repeat-root and --baseline-root.")

    try:
        payload = run_gate(
            args.golden_root,
            candidate_root=args.candidate_root,
            repeat_root=args.repeat_root,
            baseline_root=args.baseline_root,
            report_dir=args.report_dir,
            verify_only=args.verify_only,
            promotion=args.promotion,
            candidate_id=args.candidate_id,
        )
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        print(f"GOLDEN GATE ERROR: {exc}", file=sys.stderr)
        return 2

    print(f"Golden drawings: {payload['golden_count']}")
    print(f"Status: {payload['status']}")
    print(f"JSON report: {payload.get('report_json', '')}")
    print(f"Excel report: {payload.get('report_excel', '')}")
    return 0 if payload["status"] == "PASSED" else 1


if __name__ == "__main__":
    raise SystemExit(main())
