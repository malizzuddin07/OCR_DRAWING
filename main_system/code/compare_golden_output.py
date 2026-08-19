import argparse
import json
import math
from pathlib import Path

from openpyxl import Workbook, load_workbook

from analyze_golden_misses import (
    group_by_box,
    group_content_similarity,
    group_values,
    match_groups,
)


FA_COLUMNS = [
    "BALLOON NO",
    "SYMBOL",
    "VALUE",
    "-",
    "+",
    "MIN",
    "MAX",
    "REMARKS",
]

METADATA_LABELS = {
    "PART NUMBER": "part_number",
    "DRAWING NUMBER": "drawing_number",
    "REVISION": "revision",
    "MATERIAL": "material",
    "PART NAME": "part_name",
}


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def find_header_row(ws):
    for row_index in range(1, ws.max_row + 1):
        values = [clean(ws.cell(row=row_index, column=col).value).upper() for col in range(1, ws.max_column + 1)]
        if "BALLOON NO" in values and "VALUE" in values:
            return row_index, values
    return None, []


def read_fa_document(path):
    workbook = load_workbook(path, data_only=True)
    rows = []
    metadata = {}

    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell_index, cell in enumerate(row):
                label = clean(cell.value).upper()
                field = METADATA_LABELS.get(label)
                if field and field not in metadata and cell_index + 1 < len(row):
                    metadata[field] = clean(row[cell_index + 1].value)

        header_row, headers = find_header_row(ws)
        if not header_row:
            continue

        column_map = {}
        for index, header in enumerate(headers, start=1):
            if header in FA_COLUMNS:
                column_map[header] = index

        for row_index in range(header_row + 1, ws.max_row + 1):
            row = {
                column: clean(ws.cell(row=row_index, column=column_map[column]).value)
                for column in FA_COLUMNS
                if column in column_map
            }
            if any(row.values()):
                rows.append(row)

    return {"rows": rows, "metadata": metadata}


def read_fa_rows(path):
    return read_fa_document(path)["rows"]


def balloon_identity(row):
    return clean(row.get("BALLOON NO", row.get("Balloon No", "")))


def duplicate_balloon_identities(rows):
    counts = {}
    for row in rows or []:
        identity = balloon_identity(row)
        if identity:
            counts[identity] = counts.get(identity, 0) + 1
    return sorted(identity for identity, count in counts.items() if count > 1)


def row_position(row):
    try:
        x = float(row.get("X", row.get("x", "")))
        y = float(row.get("Y", row.get("y", "")))
        width = float(row.get("Width", row.get("width", 0)) or 0)
        height = float(row.get("Height", row.get("height", 0)) or 0)
    except (TypeError, ValueError):
        return None
    return x + width / 2, y + height / 2


def load_characteristic_records(path):
    """Read the positional characteristic format written by the OCR pipeline."""
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict) and isinstance(payload.get("records"), list):
        return payload["records"]
    raise ValueError(f"Invalid characteristic reference: {path}")


def characteristic_balloon_identity(row):
    return clean(row.get("Balloon No", row.get("BALLOON NO", "")))


def characteristic_box(row):
    try:
        x = float(row.get("X", row.get("x", "")))
        y = float(row.get("Y", row.get("y", "")))
        width = float(row.get("Width", row.get("width", "")))
        height = float(row.get("Height", row.get("height", "")))
    except (TypeError, ValueError):
        return None
    if width <= 0 or height <= 0:
        return None
    return x, y, width, height


def _group_balloon(group):
    return ", ".join(
        identity
        for identity in (characteristic_balloon_identity(row) for row in group.get("rows", []))
        if identity
    )


def compare_characteristic_values(expected_rows, current_rows):
    """Compare values after matching characteristic boxes by drawing position.

    Balloon numbering is intentionally not used to establish identity. A
    missing early characteristic can renumber every later balloon, but it must
    not turn one real error into dozens of false missing/extra errors.
    """
    expected_groups, invalid_expected = group_by_box(expected_rows)
    current_groups, invalid_current = group_by_box(current_rows)
    matches, missing, extra = match_groups(expected_groups, current_groups)
    differences = []

    for match in matches:
        if group_content_similarity(match["expected"], match["current"]) == 1.0:
            continue
        differences.append(
            {
                "EXPECTED ROW": _group_balloon(match["expected"]),
                "CURRENT ROW": _group_balloon(match["current"]),
                "COLUMN": "CHARACTERISTIC",
                "EXPECTED": group_values(match["expected"]),
                "CURRENT": group_values(match["current"]),
                "MATCH": {
                    "iou": round(match["iou"], 4),
                    "center_distance": round(match["center_distance"], 4),
                    "confidence": match["matching_confidence"],
                },
            }
        )

    for group in missing:
        differences.append(
            {
                "EXPECTED ROW": _group_balloon(group),
                "CURRENT ROW": "",
                "COLUMN": "ROW",
                "EXPECTED": group_values(group),
                "CURRENT": "<MISSING>",
            }
        )
    for group in extra:
        differences.append(
            {
                "EXPECTED ROW": "",
                "CURRENT ROW": _group_balloon(group),
                "COLUMN": "ROW",
                "EXPECTED": "<MISSING>",
                "CURRENT": group_values(group),
            }
        )
    for row in invalid_expected:
        differences.append(
            {
                "EXPECTED ROW": characteristic_balloon_identity(row),
                "CURRENT ROW": "",
                "COLUMN": "POSITION",
                "EXPECTED": "valid approved box",
                "CURRENT": "<INVALID EXPECTED BOX>",
            }
        )
    for row in invalid_current:
        differences.append(
            {
                "EXPECTED ROW": "",
                "CURRENT ROW": characteristic_balloon_identity(row),
                "COLUMN": "POSITION",
                "EXPECTED": "valid box",
                "CURRENT": "<INVALID CURRENT BOX>",
            }
        )
    return differences


def compare_characteristic_geometry(
    expected_rows,
    current_rows,
    *,
    center_tolerance=40.0,
    size_tolerance=30.0,
):
    """Compare approved boxes and balloon settings by drawing position.

    Pixel tolerances allow small OCR-box variations while still rejecting a
    partial crop, a moved balloon, or a materially different frame size.
    """
    expected_groups, invalid_expected = group_by_box(expected_rows)
    current_groups, invalid_current = group_by_box(current_rows)
    matches, missing, extra = match_groups(expected_groups, current_groups)
    # Preserve the useful legacy diagnostic for a box that kept the same
    # balloon identity but moved too far to qualify as a spatial match.
    remaining_missing = []
    remaining_extra = list(extra)
    for expected_group in missing:
        expected_identity = _group_balloon(expected_group)
        extra_index = next(
            (
                index
                for index, current_group in enumerate(remaining_extra)
                if expected_identity and _group_balloon(current_group) == expected_identity
            ),
            None,
        )
        if extra_index is None:
            remaining_missing.append(expected_group)
            continue
        current_group = remaining_extra.pop(extra_index)
        matches.append({"expected": expected_group, "current": current_group})
    missing = remaining_missing
    extra = remaining_extra
    issues = []
    for match in matches:
        expected_group = match["expected"]
        current_group = match["current"]
        expected = expected_group["rows"][0]
        current = current_group["rows"][0]
        identity = _group_balloon(expected_group)
        current_identity = _group_balloon(current_group)
        expected_box = expected_group["box"]
        current_box = current_group["box"]
        ex, ey, ew, eh = expected_box
        cx, cy, cw, ch = current_box
        center_distance = math.hypot((ex + ew / 2) - (cx + cw / 2), (ey + eh / 2) - (cy + ch / 2))
        if center_distance > center_tolerance:
            issues.append(
                {
                    "balloon_no": identity,
                    "current_balloon_no": current_identity,
                    "type": "position_changed",
                    "expected": [ex, ey, ew, eh],
                    "current": [cx, cy, cw, ch],
                    "center_distance": round(center_distance, 2),
                }
            )
        if abs(ew - cw) > size_tolerance or abs(eh - ch) > size_tolerance:
            issues.append(
                {
                    "balloon_no": identity,
                    "current_balloon_no": current_identity,
                    "type": "box_size_changed",
                    "expected": [ew, eh],
                    "current": [cw, ch],
                }
            )
        for field in ("Balloon Size", "Balloon Rotation"):
            expected_value = clean(expected.get(field, ""))
            current_value = clean(current.get(field, ""))
            # Fresh pipeline records do not always carry review-only balloon
            # settings. Compare the setting when both artifacts provide it;
            # layout overlap/count checks cover generated balloon placement.
            if expected_value and current_value and expected_value != current_value:
                issues.append(
                    {
                        "balloon_no": identity,
                        "current_balloon_no": current_identity,
                        "type": "balloon_setting_changed",
                        "field": field,
                        "expected": expected_value,
                        "current": current_value,
                    }
                )

    for group in missing:
        issues.append({"balloon_no": _group_balloon(group), "type": "missing_geometry"})
    for group in extra:
        issues.append({"balloon_no": _group_balloon(group), "type": "unexpected_geometry"})
    for row in invalid_expected:
        issues.append({"balloon_no": characteristic_balloon_identity(row), "type": "invalid_box"})
    for row in invalid_current:
        issues.append(
            {
                "balloon_no": characteristic_balloon_identity(row),
                "type": "unexpected_invalid_box",
            }
        )
    return issues


def compare_rows(expected_rows, current_rows):
    unmatched_current = set(range(len(current_rows)))
    pairs = []

    def content_signature(row):
        return tuple(clean(row.get(column, "")) for column in FA_COLUMNS if column != "BALLOON NO")

    for expected_index, expected in enumerate(expected_rows):
        expected_balloon = clean(expected.get("BALLOON NO", ""))
        same_balloon = [
            current_index
            for current_index in unmatched_current
            if expected_balloon and clean(current_rows[current_index].get("BALLOON NO", "")) == expected_balloon
        ]
        if same_balloon:
            if len(same_balloon) == 1:
                current_index = same_balloon[0]
            else:
                expected_position = row_position(expected)
                positioned = [
                    (math.dist(expected_position, row_position(current_rows[index])), index)
                    for index in same_balloon
                    if expected_position is not None and row_position(current_rows[index]) is not None
                ]
                current_index = min(positioned)[1] if positioned else same_balloon[0]
        else:
            signature = content_signature(expected)
            same_content = [
                current_index
                for current_index in unmatched_current
                if signature and content_signature(current_rows[current_index]) == signature
            ]
            current_index = same_content[0] if len(same_content) == 1 else None

        if current_index is not None:
            unmatched_current.remove(current_index)
        pairs.append((expected_index, current_index))

    pairs.extend((None, current_index) for current_index in sorted(unmatched_current))
    differences = []

    for expected_index, current_index in pairs:
        expected = expected_rows[expected_index] if expected_index is not None else {}
        current = current_rows[current_index] if current_index is not None else {}

        if expected_index is None or current_index is None:
            differences.append(
                {
                    "EXPECTED ROW": "" if expected_index is None else expected_index + 1,
                    "CURRENT ROW": "" if current_index is None else current_index + 1,
                    "COLUMN": "ROW",
                    "EXPECTED": "<MISSING>" if expected_index is None else clean(expected.get("BALLOON NO", "")),
                    "CURRENT": "<MISSING>" if current_index is None else clean(current.get("BALLOON NO", "")),
                }
            )
            continue

        for column in FA_COLUMNS:
            expected_value = clean(expected.get(column, ""))
            current_value = clean(current.get(column, ""))
            if expected_value != current_value:
                differences.append(
                    {
                        "EXPECTED ROW": expected_index + 1,
                        "CURRENT ROW": current_index + 1,
                        "COLUMN": column,
                        "EXPECTED": expected_value,
                        "CURRENT": current_value,
                    }
                )

    return differences


def write_report(output_path, expected_rows, current_rows, differences):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Metric", "Value"])
    summary.append(["Expected rows", len(expected_rows)])
    summary.append(["Current rows", len(current_rows)])
    summary.append(["Different cells", len(differences)])

    details = workbook.create_sheet("Differences")
    details.append(["EXPECTED ROW", "CURRENT ROW", "COLUMN", "EXPECTED", "CURRENT"])
    for item in differences:
        details.append(
            [
                item["EXPECTED ROW"],
                item["CURRENT ROW"],
                item["COLUMN"],
                item["EXPECTED"],
                item["CURRENT"],
            ]
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_json_report(output_path, payload):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(description="Compare current FA Excel output against a manually verified golden Excel.")
    parser.add_argument("--expected", required=True, help="Path to manually verified expected FA Excel.")
    parser.add_argument("--current", required=True, help="Path to current generated FA Excel.")
    parser.add_argument("--output", required=True, help="Path to write comparison report Excel.")
    args = parser.parse_args()

    expected_path = Path(args.expected)
    current_path = Path(args.current)
    output_path = Path(args.output)

    expected_rows = read_fa_rows(expected_path)
    current_rows = read_fa_rows(current_path)
    differences = compare_rows(expected_rows, current_rows)
    write_report(output_path, expected_rows, current_rows, differences)

    print(f"Expected rows: {len(expected_rows)}")
    print(f"Current rows: {len(current_rows)}")
    print(f"Different cells: {len(differences)}")
    print(f"Report: {output_path}")


if __name__ == "__main__":
    main()
