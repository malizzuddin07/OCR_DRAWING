"""Create a human-reviewable missed-characteristic evidence package.

Approved characteristic boxes are the source of truth. Current detections are
matched by page position, not spreadsheet row order or balloon numbering.
"""

import argparse
import json
import math
import re
from collections import Counter
from pathlib import Path

import cv2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_GOLDEN_ROOT = PROJECT_ROOT / "golden_tests"

COLORS = {
    "missing": (0, 0, 255),       # red in BGR
    "incorrect": (0, 165, 255),  # orange
    "extra": (255, 0, 255),      # purple
}


def load_json(path):
    return json.loads(Path(path).read_text(encoding="utf-8"))


def clean(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def item_box(item):
    try:
        x = float(item.get("X", ""))
        y = float(item.get("Y", ""))
        width = float(item.get("Width", ""))
        height = float(item.get("Height", ""))
    except (TypeError, ValueError):
        return None
    if width <= 0 or height <= 0:
        return None
    return x, y, width, height


def group_by_box(rows):
    grouped = {}
    invalid = []
    for row in rows or []:
        box = item_box(row)
        if box is None:
            invalid.append(row)
            continue
        # Approved subrows share the same integer crop. Rounding also removes
        # harmless JSON float formatting differences.
        key = tuple(round(value, 2) for value in box)
        grouped.setdefault(key, []).append(row)
    return [{"box": key, "rows": values} for key, values in grouped.items()], invalid


def box_iou(first, second):
    ax, ay, aw, ah = first
    bx, by, bw, bh = second
    x1, y1 = max(ax, bx), max(ay, by)
    x2, y2 = min(ax + aw, bx + bw), min(ay + ah, by + bh)
    intersection = max(0.0, x2 - x1) * max(0.0, y2 - y1)
    union = aw * ah + bw * bh - intersection
    return intersection / union if union > 0 else 0.0


def normalized_center_distance(first, second):
    ax, ay, aw, ah = first
    bx, by, bw, bh = second
    distance = math.hypot((ax + aw / 2) - (bx + bw / 2), (ay + ah / 2) - (by + bh / 2))
    return distance / max(20.0, math.hypot(aw, ah))


def row_signature(row):
    return tuple(
        clean(row.get(field, "")).upper()
        for field in (
            "Report Symbol",
            "Symbol",
            "Dimension",
            "Tolerance -",
            "Tolerance +",
            "MIN",
            "MAX",
            "Measurement Type",
        )
    )


def group_content_similarity(first, second):
    expected = Counter(row_signature(row) for row in first["rows"])
    current = Counter(row_signature(row) for row in second["rows"])
    common = sum((expected & current).values())
    total = max(sum(expected.values()), sum(current.values()), 1)
    return common / total


def match_groups(expected_groups, current_groups):
    """Return stable one-to-one spatial matches plus missing and extra groups."""
    candidates = []
    for expected_index, expected in enumerate(expected_groups):
        for current_index, current in enumerate(current_groups):
            overlap = box_iou(expected["box"], current["box"])
            distance = normalized_center_distance(expected["box"], current["box"])
            if overlap < 0.10 and distance > 0.45:
                continue
            content = group_content_similarity(expected, current)
            score = overlap * 5.0 + content * 0.5 - distance
            candidates.append((score, overlap, distance, content, expected_index, current_index))

    candidates.sort(reverse=True)
    used_expected = set()
    used_current = set()
    matches = []
    for score, overlap, distance, content, expected_index, current_index in candidates:
        if expected_index in used_expected or current_index in used_current:
            continue
        used_expected.add(expected_index)
        used_current.add(current_index)
        confidence = "HIGH" if overlap >= 0.40 else ("MEDIUM" if overlap >= 0.15 else "LOW")
        matches.append(
            {
                "expected": expected_groups[expected_index],
                "current": current_groups[current_index],
                "iou": overlap,
                "center_distance": distance,
                "content_similarity": content,
                "matching_confidence": confidence,
            }
        )

    missing = [group for index, group in enumerate(expected_groups) if index not in used_expected]
    extra = [group for index, group in enumerate(current_groups) if index not in used_current]
    return matches, missing, extra


def group_is_correct(match):
    expected = Counter(row_signature(row) for row in match["expected"]["rows"])
    current = Counter(row_signature(row) for row in match["current"]["rows"])
    return expected == current


def group_balloons(group):
    return ", ".join(clean(row.get("Balloon No", "")) for row in group["rows"])


def group_values(group):
    values = []
    for row in group["rows"]:
        symbol = clean(row.get("Report Symbol", row.get("Symbol", "")))
        dimension = clean(row.get("Dimension", ""))
        minus = clean(row.get("Tolerance -", ""))
        plus = clean(row.get("Tolerance +", ""))
        text = " ".join(part for part in (symbol, dimension) if part)
        if minus or plus:
            text += f" [-{minus} / +{plus}]"
        values.append(text)
    return " | ".join(values)


def group_types(group):
    return ", ".join(sorted({clean(row.get("Measurement Type", "")) for row in group["rows"] if clean(row.get("Measurement Type", ""))}))


def safe_name(value):
    return re.sub(r"[^0-9A-Za-z_.-]+", "_", clean(value)).strip("_") or "unknown"


def integer_box(box, image_shape):
    height, width = image_shape[:2]
    x, y, box_width, box_height = box
    x1 = max(0, min(width - 1, int(round(x))))
    y1 = max(0, min(height - 1, int(round(y))))
    x2 = max(x1 + 1, min(width, int(round(x + box_width))))
    y2 = max(y1 + 1, min(height, int(round(y + box_height))))
    return x1, y1, x2, y2


def draw_issue(image, issue_type, group, label):
    color = COLORS[issue_type]
    x1, y1, x2, y2 = integer_box(group["box"], image.shape)
    thickness = max(3, round(max(image.shape[:2]) / 1200))
    cv2.rectangle(image, (x1, y1), (x2, y2), color, thickness)
    font_scale = max(0.65, max(image.shape[:2]) / 4200)
    text_size, _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, font_scale, thickness)
    label_y1 = max(0, y1 - text_size[1] - 12)
    label_x2 = min(image.shape[1], x1 + text_size[0] + 12)
    cv2.rectangle(image, (x1, label_y1), (label_x2, y1), color, -1)
    cv2.putText(image, label, (x1 + 5, max(text_size[1] + 2, y1 - 5)), cv2.FONT_HERSHEY_SIMPLEX, font_scale, (255, 255, 255), thickness, cv2.LINE_AA)


def save_issue_crop(source_image, output_path, issue_type, group, label, padding=100):
    x1, y1, x2, y2 = integer_box(group["box"], source_image.shape)
    crop_x1 = max(0, x1 - padding)
    crop_y1 = max(0, y1 - padding)
    crop_x2 = min(source_image.shape[1], x2 + padding)
    crop_y2 = min(source_image.shape[0], y2 + padding)
    crop = source_image[crop_y1:crop_y2, crop_x1:crop_x2].copy()
    local_group = {"box": (x1 - crop_x1, y1 - crop_y1, x2 - x1, y2 - y1), "rows": group["rows"]}
    draw_issue(crop, issue_type, local_group, label)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if not cv2.imwrite(str(output_path), crop):
        raise OSError(f"Could not write issue crop: {output_path}")


def issue_record(drawing, issue_type, expected=None, current=None, match=None, crop_file=""):
    expected = expected or {"box": ("", "", "", ""), "rows": []}
    current = current or {"box": ("", "", "", ""), "rows": []}
    box = expected["box"] if expected["rows"] else current["box"]
    return {
        "Drawing": drawing,
        "Issue": issue_type.upper(),
        "Severity": "HIGH" if issue_type in {"missing", "incorrect"} else "MEDIUM",
        "Expected Balloon": group_balloons(expected),
        "Current Balloon": group_balloons(current),
        "Expected Type": group_types(expected),
        "Current Type": group_types(current),
        "Expected Value": group_values(expected),
        "Current Value": group_values(current),
        "X": box[0],
        "Y": box[1],
        "Width": box[2],
        "Height": box[3],
        "Position IoU": round(match["iou"], 4) if match else "",
        "Match Confidence": match["matching_confidence"] if match else "",
        "Expected Rows": len(expected["rows"]),
        "Current Rows": len(current["rows"]),
        "Crop File": crop_file,
        "User Confirmation": "PENDING",
        "User Notes": "",
    }


def analyze_drawing(manifest, golden_root, candidate_root, output_root):
    drawing = manifest["drawing_number"]
    expected_path = golden_root / manifest["expected_characteristics"]["path"]
    current_path = candidate_root / drawing / "characteristics.json"
    image_path = candidate_root / drawing / "original.png"
    expected_rows = load_json(expected_path)["records"]
    current_rows = load_json(current_path)
    image = cv2.imread(str(image_path))
    if image is None:
        raise FileNotFoundError(f"Could not read source image: {image_path}")

    expected_groups, invalid_expected = group_by_box(expected_rows)
    current_groups, invalid_current = group_by_box(current_rows)
    if invalid_expected:
        raise ValueError(f"{drawing}: {len(invalid_expected)} approved rows have no valid position.")
    matches, missing, extra = match_groups(expected_groups, current_groups)
    correct = [match for match in matches if group_is_correct(match)]
    incorrect = [match for match in matches if not group_is_correct(match)]

    annotated = image.copy()
    crop_root = output_root / "issue_crops" / drawing
    records = []
    for group in missing:
        balloon = group_balloons(group)
        label = f"MISSING {balloon}"
        draw_issue(annotated, "missing", group, label)
        crop_path = crop_root / f"missing_expected_{safe_name(balloon)}.png"
        save_issue_crop(image, crop_path, "missing", group, label)
        records.append(issue_record(drawing, "missing", expected=group, crop_file=str(crop_path.relative_to(output_root))))
    for match in incorrect:
        expected = match["expected"]
        current = match["current"]
        label = f"WRONG {group_balloons(expected)}"
        draw_issue(annotated, "incorrect", expected, label)
        crop_path = crop_root / f"incorrect_expected_{safe_name(group_balloons(expected))}.png"
        save_issue_crop(image, crop_path, "incorrect", expected, label)
        records.append(issue_record(drawing, "incorrect", expected=expected, current=current, match=match, crop_file=str(crop_path.relative_to(output_root))))
    for group in extra:
        balloon = group_balloons(group)
        label = f"EXTRA {balloon}"
        draw_issue(annotated, "extra", group, label)
        crop_path = crop_root / f"extra_current_{safe_name(balloon)}.png"
        save_issue_crop(image, crop_path, "extra", group, label)
        records.append(issue_record(drawing, "extra", current=group, crop_file=str(crop_path.relative_to(output_root))))

    annotated_path = output_root / "annotated_pages" / f"{drawing}_issues.png"
    annotated_path.parent.mkdir(parents=True, exist_ok=True)
    if not cv2.imwrite(str(annotated_path), annotated):
        raise OSError(f"Could not write annotated page: {annotated_path}")

    return {
        "drawing": drawing,
        "approved_characteristics": len(expected_rows),
        "current_characteristics": len(current_rows),
        "approved_callouts": len(expected_groups),
        "current_callouts": len(current_groups),
        "correct_callouts": len(correct),
        "incorrect_callouts": len(incorrect),
        "missing_callouts": len(missing),
        "extra_callouts": len(extra),
        "invalid_current_rows": len(invalid_current),
        "annotated_page": str(annotated_path.relative_to(output_root)),
        "issues": records,
    }


def write_workbook(path, results):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    headers = [
        "Drawing", "Approved Characteristics", "Current Characteristics",
        "Approved Callouts", "Current Callouts", "Correct Callouts",
        "Incorrect Callouts", "Missing Callouts", "Extra Callouts",
        "Callout Coverage %", "Annotated Page",
    ]
    summary.append(headers)
    summary.freeze_panes = "A2"
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for result in results:
        covered = result["correct_callouts"] + result["incorrect_callouts"]
        coverage = round(100 * covered / max(1, result["approved_callouts"]), 1)
        summary.append(
            [
                result["drawing"], result["approved_characteristics"], result["current_characteristics"],
                result["approved_callouts"], result["current_callouts"], result["correct_callouts"],
                result["incorrect_callouts"], result["missing_callouts"], result["extra_callouts"],
                coverage, result["annotated_page"],
            ]
        )

    issue_headers = [
        "Drawing", "Issue", "Severity", "Expected Balloon", "Current Balloon",
        "Expected Type", "Current Type", "Expected Value", "Current Value",
        "X", "Y", "Width", "Height", "Position IoU", "Match Confidence",
        "Expected Rows", "Current Rows", "Crop File", "User Confirmation", "User Notes",
    ]
    all_issues = [issue for result in results for issue in result["issues"]]
    for issue_type in ("MISSING", "INCORRECT", "EXTRA"):
        sheet = workbook.create_sheet(issue_type.title())
        sheet.append(issue_headers)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:T1"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor={"MISSING": "C00000", "INCORRECT": "F4B183", "EXTRA": "7030A0"}[issue_type])
        for issue in all_issues:
            if issue["Issue"] == issue_type:
                sheet.append([issue[header] for header in issue_headers])
        for column, width in {"A": 25, "B": 12, "D": 18, "E": 18, "F": 20, "G": 20, "H": 35, "I": 35, "R": 55, "S": 20, "T": 35}.items():
            sheet.column_dimensions[column].width = width

    instructions = workbook.create_sheet("Review Instructions")
    instructions.append(["Colour", "Meaning", "User action"])
    instructions.append(["RED", "Approved callout was not detected", "Confirm the red box contains a required dimension/callout"])
    instructions.append(["ORANGE", "A callout was detected but its extracted fields differ", "Compare expected and current values"])
    instructions.append(["PURPLE", "Current OCR detected a box with no approved positional match", "Confirm whether it is false or a matching error"])
    instructions.append([])
    instructions.append(["Set User Confirmation to CONFIRMED or REJECTED. Add a short note only when needed."])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def run_analysis(golden_root, candidate_root, output_root):
    golden_root = Path(golden_root)
    candidate_root = Path(candidate_root)
    output_root = Path(output_root)
    output_root.mkdir(parents=True, exist_ok=True)
    manifests = [load_json(path) for path in golden_root.glob("*_golden_manifest.json")]
    manifests.sort(key=lambda item: int(item.get("golden_test_number", 9999)))
    if len(manifests) != 5:
        raise ValueError(f"Expected five golden manifests, found {len(manifests)}.")
    results = [analyze_drawing(manifest, golden_root, candidate_root, output_root) for manifest in manifests]
    workbook_path = output_root / "missed_characteristics.xlsx"
    write_workbook(workbook_path, results)
    payload = {
        "schema_version": 1,
        "golden_root": str(golden_root.resolve()),
        "candidate_root": str(candidate_root.resolve()),
        "drawing_count": len(results),
        "results": results,
    }
    json_path = output_root / "missed_characteristics.json"
    json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return payload, workbook_path, json_path


def main():
    parser = argparse.ArgumentParser(description="Build missed-characteristic Excel and image evidence.")
    parser.add_argument("--golden-root", default=str(DEFAULT_GOLDEN_ROOT))
    parser.add_argument("--candidate-root", required=True)
    parser.add_argument("--output-root", required=True)
    args = parser.parse_args()
    payload, workbook_path, json_path = run_analysis(args.golden_root, args.candidate_root, args.output_root)
    print(f"Drawings: {payload['drawing_count']}")
    print(f"Missing callouts: {sum(item['missing_callouts'] for item in payload['results'])}")
    print(f"Incorrect callouts: {sum(item['incorrect_callouts'] for item in payload['results'])}")
    print(f"Extra callouts: {sum(item['extra_callouts'] for item in payload['results'])}")
    print(f"Excel: {workbook_path}")
    print(f"JSON: {json_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
